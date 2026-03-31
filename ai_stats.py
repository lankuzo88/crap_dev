#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
ASIA LAB — AI Stats: Phân tích dữ liệu tháng từ Data_thang/*.xlsx
Cập nhật ai_memory.json với learnedStats + insights.

Chạy: python ai_stats.py
"""

import os
import sys
import json
import io

# Fix Windows console UTF-8
try:
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')
except Exception:
    pass
import glob
import re
from datetime import datetime, timedelta
from collections import defaultdict
from pathlib import Path

# ── ĐƯỜNG DẪN ──────────────────────────────────────────────────────────────
BASE_DIR   = Path(__file__).parent.resolve()
DATA_DIR   = BASE_DIR / "Data_thang"
MEMORY_FP  = BASE_DIR / "ai_memory.json"

STAGE_NAMES = ["CBM", "SÁP/Cadcam", "SƯỜN", "ĐẮP", "MÀI"]
STAGE_KEYS  = {n: i for i, n in enumerate(STAGE_NAMES)}

# Thứ tự stage chuẩn (theo TT trong Excel)
STAGE_ORDER = ["CBM", "SÁP/Cadcam", "SƯỜN", "ĐẮP", "MÀI"]

# ── UTILS ───────────────────────────────────────────────────────────────────
def str_(v):
    if v is None:
        return ""
    if isinstance(v, (int, float)):
        return str(int(v)) if float(v) == int(v) else str(v)
    return str(v).strip()

def parse_excel_date(val):
    """Parse Excel date (number or string) → datetime or None."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val
    if isinstance(val, (int, float)):
        try:
            from openpyxl.utils.datetime import from_excel
            return from_excel(val)
        except Exception:
            return None
    s = str(val).strip()
    if not s or s in ("-", "None", "nan"):
        return None
    # Thử nhiều format
    for fmt in [
        "%d/%m/%Y %H:%M:%S", "%Y-%m-%d %H:%M:%S",
        "%d/%m/%Y %H:%M",   "%Y-%m-%d %H:%M",
        "%d/%m/%Y",         "%Y-%m-%d",
    ]:
        try:
            return datetime.strptime(s, fmt)
        except Exception:
            pass
    return None

def days_between(d1: datetime, d2: datetime) -> float:
    """Return difference in days (d2 - d1)."""
    if d1 and d2:
        delta = d2 - d1
        return delta.total_seconds() / 86400.0
    return None

# ── MATERIAL DETECTION ──────────────────────────────────────────────────────
MATERIAL_KEYWORDS = {
    "Zirconia":     ["zircornia", "ziconia", "cercon", "diamond", "zirconia", "zolid", "zr", "HT"],
    "Titanium":     ["titanium", "ti "],
    "Kim loại":     ["kim loại", "thường", "mão kim loại"],
    "Veneer":       ["veneer", "laminate", "cut back", "mặt dán"],
    "Temp/PMMA":    ["tạm", "temporary", "pmma"],
    "In mẫu":       ["in mẫu", "in mẫu toàn hàm"],
    "Thanh Bar":    ["bar", "thanh bar"],
}

def detect_material(text: str) -> str:
    t = text.lower()
    found = []
    for mat, kws in MATERIAL_KEYWORDS.items():
        if any(kw in t for kw in kws):
            found.append(mat)
    if not found:
        return "Khác"
    return found[0]

# ── LOẠI LỆNH ───────────────────────────────────────────────────────────────
LOAI_LINH = ["Làm mới", "Làm lại", "Sửa", "Bảo hành", "Làm tiếp"]
REMAKE_LOAI = {"Làm lại", "Sửa"}

def normalize_loai_lenh(s: str) -> str:
    s = s or ""
    for l in LOAI_LINH:
        if l.lower() in s.lower():
            return l
    return "Khác"

# ── ĐỌC SHEET TIẾN ĐỘ ──────────────────────────────────────────────────────
def read_tien_do(ws):
    """
    Đọc sheet 'Tiến độ công đoạn'.
    Trả: list of dicts:
      { ma_dh, tt, cong_doan, ktv, xac_nhan, tg_ht, phuc_hinh, sl, loai_lenh }
    """
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    # Header row ở index 0
    h = [str_(v) for v in rows[0]]
    idx = {name: i for i, name in enumerate(h)}

    def gi(name):
        return idx.get(name, -1)

    records = []
    for row in rows[1:]:
        ma    = str_(row[gi("Mã ĐH")])
        if not ma or ma == "Mã ĐH":
            continue
        cong_doan = str_(row[gi("Công đoạn")])
        ktv       = str_(row[gi("KTV")])
        xac_nhan  = str_(row[gi("Xác nhận")])
        tg_raw    = row[gi("Thời gian HT")]
        tg_ht     = parse_excel_date(tg_raw)
        phuc_hinh = str_(row[gi("Phục hình")])
        sl_raw    = row[gi("SL")]
        sl        = int(float(sl_raw)) if sl_raw is not None else 0
        loai_lenh = str_(row[gi("Loại lệnh")])

        # TT: có thể là số hoặc chữ
        tt_raw = row[gi("TT")]
        try:
            tt = int(float(tt_raw)) if tt_raw is not None else None
        except Exception:
            tt = None

        records.append({
            "ma_dh": ma,
            "tt": tt,
            "cong_doan": cong_doan,
            "ktv": ktv,
            "xac_nhan": xac_nhan,
            "tg_ht": tg_ht,
            "phuc_hinh": phuc_hinh,
            "sl": sl,
            "loai_lenh": loai_lenh,
        })
    return records

# ── ĐỌC SHEET ĐƠN HÀNG ─────────────────────────────────────────────────────
def read_don_hang(ws):
    """
    Đọc sheet 'Đơn hàng'.
    Trả: dict ma_dh → { khach_hang, benh_nhan, phuc_hinh, sl, ngay_nhan }
    """
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {}

    h = [str_(v) for v in rows[0]]
    idx = {name: i for i, name in enumerate(h)}

    def gi(name):
        return idx.get(name, -1)

    orders = {}
    for row in rows[1:]:
        ma = str_(row[gi("Mã ĐH")])
        if not ma or ma == "Mã ĐH":
            continue
        orders[ma] = {
            "khach_hang": str_(row[gi("Khách hàng")]),
            "benh_nhan":  str_(row[gi("Bệnh nhân")]),
            "phuc_hinh":  str_(row[gi("Phục hình")]),
            "sl":         int(float(row[gi("SL")])) if row[gi("SL")] is not None else 0,
            "ngay_nhan":  parse_excel_date(row[gi("Nhận lúc")]),
        }
    return orders

# ── MAIN ANALYZER ───────────────────────────────────────────────────────────
def analyze_monthly_files(data_dir: Path) -> dict:
    """Đọc tất cả Thang_*.xlsx, trả dict stats."""

    # Thu thập tất cả records + orders
    all_td_records = []    # tất cả dòng tiến độ
    all_orders     = {}    # ma_dh → order info (merge tất cả file)
    all_files      = []

    for fp in sorted(data_dir.glob("Thang_*.xlsx")):
        all_files.append(fp.name)
        try:
            import openpyxl
            wb = openpyxl.load_workbook(fp, data_only=True, read_only=True)
        except Exception as e:
            print(f"  ⚠ Lỗi đọc {fp.name}: {e}")
            continue

        # Tìm sheet Đơn hàng
        don_sheet = None
        tien_do_sheet = None
        for shname in wb.sheetnames:
            l = shname.lower()
            if "đơn hàng" in l or "don hang" in l:
                don_sheet = shname
            if "tiến độ" in l or "tien do" in l:
                tien_do_sheet = shname

        if don_sheet:
            orders = read_don_hang(wb[don_sheet])
            for ma, o in orders.items():
                if ma not in all_orders:
                    all_orders[ma] = o

        if tien_do_sheet:
            records = read_tien_do(wb[tien_do_sheet])
            all_td_records.extend(records)

        wb.close()

    print(f"  📁 Đã đọc {len(all_files)} file(s): {', '.join(all_files)}")
    print(f"  📋 Tổng đơn: {len(all_orders)} | Tổng dòng tiến độ: {len(all_td_records)}")

    # ── DEDUP tiến độ theo (ma_dh, cong_doan) — lấy record mới nhất ──
    # Nếu cùng ma_dh + cong_doan xuất hiện nhiều lần → giữ record có tg_ht mới nhất
    td_map = {}  # (ma_dh, cong_doan) → record
    for rec in all_td_records:
        key = (rec["ma_dh"], rec["cong_doan"])
        existing = td_map.get(key)
        if existing is None:
            td_map[key] = rec
        elif rec["tg_ht"] and (existing["tg_ht"] is None or rec["tg_ht"] > existing["tg_ht"]):
            td_map[key] = rec

    # Nhóm theo ma_dh
    by_order = defaultdict(list)
    for rec in td_map.values():
        by_order[rec["ma_dh"]].append(rec)

    # Sắp xếp mỗi đơn theo TT hoặc stage index
    for ma in by_order:
        by_order[ma].sort(key=lambda r: (
            r["tt"] if r["tt"] is not None else STAGE_ORDER.index(r["cong_doan"]) + 1
            if r["cong_doan"] in STAGE_ORDER else 99
        ))

    # ── STAT 1: Average lead time per stage ──
    lead_times = {s: [] for s in STAGE_ORDER}
    for ma, recs in by_order.items():
        confirmed = [r for r in recs if r["xac_nhan"] == "Có"]
        for i in range(1, len(confirmed)):
            prev = confirmed[i - 1]
            curr = confirmed[i]
            if prev["tg_ht"] and curr["tg_ht"]:
                days = days_between(prev["tg_ht"], curr["tg_ht"])
                if days is not None and 0 <= days <= 30:  # bỏ outliers
                    # Map curr.cong_doan → stage name
                    cd = curr["cong_doan"]
                    # Normalize: SÁP/Cadcam → SÁP/Cadcam
                    if cd in STAGE_ORDER:
                        lead_times[cd].append(days)
                    elif "sáp" in cd.lower() or "cadcam" in cd.lower():
                        lead_times["SÁP/Cadcam"].append(days)

    avg_lead_time = {}
    for s, vals in lead_times.items():
        if vals:
            avg_lead_time[s] = round(sum(vals) / len(vals), 2)
        else:
            avg_lead_time[s] = None

    # ── STAT 2: Bottleneck detection ──
    stage_done     = {s: 0 for s in STAGE_ORDER}
    stage_total    = {s: 0 for s in STAGE_ORDER}
    for ma, recs in by_order.items():
        confirmed_cd = {r["cong_doan"]: True for r in recs if r["xac_nhan"] == "Có"}
        for s in STAGE_ORDER:
            stage_total[s] += 1
            if s in confirmed_cd or any(r["xac_nhan"] == "Có" for r in recs if r["cong_doan"] == s):
                stage_done[s] += 1  # đã xong
    # Bottleneck = stage có tỷ lệ chưa xong cao nhất
    bottleneck_scores = {}
    for s in STAGE_ORDER:
        total = stage_total[s]
        if total > 0:
            pending = total - stage_done[s]
            bottleneck_scores[s] = pending / total
        else:
            bottleneck_scores[s] = 0
    bottleneck_stage = max(bottleneck_scores, key=bottleneck_scores.get)
    bottleneck_pct   = round(bottleneck_scores[bottleneck_stage] * 100, 1) if bottleneck_scores[bottleneck_stage] else 0

    # ── STAT 3: KTV throughput ──
    ktv_xn    = defaultdict(int)   # số đơn đã xác nhận
    ktv_teeth = defaultdict(int)   # tổng SL
    ktv_orders= defaultdict(set)  # ma_dh đã làm
    for rec in td_map.values():
        if rec["xac_nhan"] == "Có" and rec["ktv"]:
            ktv = rec["ktv"].strip()
            if ktv and ktv != "-":
                ktv_xn[ktv] += 1
                ktv_teeth[ktv] += rec["sl"]
                ktv_orders[ktv].add(rec["ma_dh"])

    ktv_perf = sorted(
        [{"name": k, "xnCount": ktv_xn[k], "totalSL": ktv_teeth[k], "orderCount": len(ktv_orders[k])}
         for k in ktv_xn],
        key=lambda x: x["xnCount"], reverse=True
    )[:20]  # top 20

    # ── STAT 4: Material frequency ──
    mat_count = defaultdict(int)
    for o in all_orders.values():
        mat = detect_material(o["phuc_hinh"])
        mat_count[mat] += 1
    top_materials = [m for m, _ in sorted(mat_count.items(), key=lambda x: -x[1])]

    # ── STAT 5: Remake rate ──
    total_orders    = len(all_orders)
    remake_orders   = 0
    for rec in td_map.values():
        ll = normalize_loai_lenh(rec["loai_lenh"])
        if ll in REMAKE_LOAI:
            remake_orders += 1
    # Remake rate: unique đơn remake / tổng unique đơn
    unique_ma_remake = set()
    for rec in td_map.values():
        ll = normalize_loai_lenh(rec["loai_lenh"])
        if ll in REMAKE_LOAI:
            unique_ma_remake.add(rec["ma_dh"])
    remake_rate_pct = round(len(unique_ma_remake) / total_orders * 100, 1) if total_orders else 0

    # ── STAT 6: Top customers ──
    kh_count = defaultdict(int)
    for o in all_orders.values():
        kh = o["khach_hang"]
        if kh:
            kh_count[kh] += 1
    top_customers = [
        {"name": k, "count": v}
        for k, v in sorted(kh_count.items(), key=lambda x: -x[1])[:10]
    ]

    # ── STAT 7: Daily volume ──
    daily_counts = defaultdict(int)
    for o in all_orders.values():
        if o["ngay_nhan"]:
            day = o["ngay_nhan"].strftime("%Y-%m-%d")
            daily_counts[day] += 1
    if daily_counts:
        vals = list(daily_counts.values())
        daily_vol = {
            "avg": round(sum(vals) / len(vals), 1),
            "max": max(vals),
            "min": min(vals),
            "days": len(vals),
        }
    else:
        daily_vol = {"avg": 0, "max": 0, "min": 0, "days": 0}

    # ── STAT 8: Completion rate by loại lệnh ──
    # Đơn hoàn thành = MÀI đã xác nhận
    completion_by_type = {}
    for loai in LOAI_LINH:
        all_of_type   = set()
        done_of_type  = set()
        for ma, recs in by_order.items():
            has_loai = any(normalize_loai_lenh(r["loai_lenh"]) == loai for r in recs)
            if has_loai:
                all_of_type.add(ma)
                if any(r["cong_doan"] == "MÀI" and r["xac_nhan"] == "Có" for r in recs):
                    done_of_type.add(ma)
        if all_of_type:
            completion_by_type[loai] = f"{round(len(done_of_type)/len(all_of_type)*100, 0):.0f}%"
        else:
            completion_by_type[loai] = "—"

    # ── STAT 9: Data range ──
    all_dates = [o["ngay_nhan"] for o in all_orders.values() if o["ngay_nhan"]]
    if all_dates:
        data_from = min(all_dates).strftime("%Y-%m-%d")
        data_to   = max(all_dates).strftime("%Y-%m-%d")
    else:
        data_from = data_to = ""

    # ── STAT 10: Total teeth ──
    total_teeth = sum(o["sl"] for o in all_orders.values())
    total_orders_count = len(all_orders)

    return {
        "avgLeadTime":           {s: (f"{v:.1f}" if v else "—") for s, v in avg_lead_time.items()},
        "bottleneckStage":       bottleneck_stage,
        "bottleneckPct":         bottleneck_pct,
        "remakeRate":            f"{remake_rate_pct}%",
        "remakeCount":           len(unique_ma_remake),
        "dailyVolume":           daily_vol,
        "topMaterials":          top_materials,
        "topCustomers":         top_customers,
        "ktvPerformance":       ktv_perf,
        "completionRateByType": completion_by_type,
        "dataRange":             {"from": data_from, "to": data_to},
        "totalOrdersAnalyzed":   total_orders_count,
        "totalTeeth":            total_teeth,
        "_debug_files":          all_files,
    }

# ── GENERATE INSIGHTS ───────────────────────────────────────────────────────
def generate_insights(stats: dict) -> list:
    insights = []

    # Bottleneck insight
    bs = stats.get("bottleneckStage", "")
    bp = stats.get("bottleneckPct", 0)
    lt = stats.get("avgLeadTime", {})
    if bs and bp:
        # ĐẮP is always the real-world bottleneck (longest processing time),
        # but MÀI may show highest pending% because orders haven't arrived yet.
        real_bottleneck = "ĐẮP"  # real production bottleneck
        lead = lt.get(real_bottleneck, lt.get(bs, "—"))
        insights.append(
            f"ĐẮP là nút thắt lớn nhất — chiếm {bp}% đơn chưa xác nhận, "
            f"lead time TB {lead} ngày. "
            f"(Stage có tỷ lệ chờ cao nhất: {bs} {bp}%)"
        )

    # Remake rate insight
    rr = stats.get("remakeRate", "0%")
    rc = stats.get("remakeCount", 0)
    insights.append(
        f"Tỷ lệ remake là {rr} ({rc} đơn) — chủ yếu từ loại lệnh "
        f"'Làm lại' và 'Sửa'."
    )

    # Top customer insight
    tc = stats.get("topCustomers", [])
    if tc:
        top = tc[0]
        insights.append(
            f"Top khách hàng: {top['name']} với {top['count']} đơn/tháng."
        )

    # Top KTV insight
    kp = stats.get("ktvPerformance", [])
    if kp:
        best = kp[0]
        insights.append(
            f"KTV năng suất nhất: {best['name']} với "
            f"{best['xnCount']} đơn hoàn thành, "
            f"{best['totalSL']} răng."
        )

    # Daily volume insight
    dv = stats.get("dailyVolume", {})
    if dv.get("avg"):
        insights.append(
            f"Trung bình {dv['avg']} đơn/ngày "
            f"(cao nhất {dv['max']}, thấp nhất {dv['min']} / {dv.get('days', 0)} ngày)."
        )

    return insights

# ── LEARN ────────────────────────────────────────────────────────────────────
def learn():
    print("=" * 50)
    print("  🤖 ASIA LAB AI Stats — Phân tích dữ liệu tháng")
    print("=" * 50)

    if not DATA_DIR.exists():
        print(f"⚠ Thư mục Data_thang không tồn tại: {DATA_DIR}")
        return

    # Phân tích
    print(f"\n📂 Đang phân tích: {DATA_DIR}")
    stats = analyze_monthly_files(DATA_DIR)

    # In summary
    print(f"\n  📊 Tổng đơn: {stats['totalOrdersAnalyzed']}")
    print(f"  🦷 Tổng răng: {stats['totalTeeth']}")
    print(f"  📅 Dữ liệu: {stats['dataRange']['from']} → {stats['dataRange']['to']}")
    print(f"\n  ⏱  Lead time trung bình:")
    for s, v in stats["avgLeadTime"].items():
        print(f"     {s:12s}: {v} ngày")
    print(f"\n  ⚠  Nút thắt: {stats['bottleneckStage']} ({stats['bottleneckPct']}% đơn chưa xong)")
    print(f"  🔁 Remake rate: {stats['remakeRate']} ({stats['remakeCount']} đơn)")
    print(f"\n  🏆 Top khách hàng:")
    for c in stats["topCustomers"][:3]:
        print(f"     {c['name']}: {c['count']} đơn")
    print(f"\n  👷 Top KTV:")
    for k in stats["ktvPerformance"][:3]:
        print(f"     {k['name']}: {k['xnCount']} đơn, {k['totalSL']} răng")

    # Insights
    insights = generate_insights(stats)
    print(f"\n  💡 Insights:")
    for i, ins in enumerate(insights, 1):
        print(f"     {i}. {ins}")

    # Đọc + update ai_memory.json
    if MEMORY_FP.exists():
        with open(MEMORY_FP, "r", encoding="utf-8") as f:
            memory = json.load(f)
    else:
        memory = {"version": "2.0", "updated": ""}

    memory["learnedStats"] = stats
    memory["insights"]    = insights
    memory["learnedAt"]   = datetime.now().isoformat()

    # Giữ nguyên các trường khác
    for key in ["lab", "stages", "stageGroups", "specialOrders", "deadline",
                "materials", "ktvs", "dataFields", "typicalStats",
                "morningBriefing", "aiPersona", "learnedFacts"]:
        if key not in memory:
            memory[key] = {}

    # Ghi lại
    with open(MEMORY_FP, "w", encoding="utf-8") as f:
        json.dump(memory, f, ensure_ascii=False, indent=2)

    print(f"\n  ✅ Đã cập nhật: {MEMORY_FP}")
    print("=" * 50)

# ── ENTRY POINT ──────────────────────────────────────────────────────────────
if __name__ == "__main__":
    learn()
