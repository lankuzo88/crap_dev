#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
ASIA LAB AI Cowork Sandbox — Tool Registry
15+ tools: file, Excel, analysis, report, memory, system
"""
import os, sys, json, re, glob
from datetime import datetime
from pathlib import Path

BASE_DIR  = Path(__file__).parent.resolve()
PARENT_DIR = BASE_DIR.parent          # = crap_dev/

# ── Safety helpers ─────────────────────────────────────────────────────────
def safe_path(path: str, base: Path = BASE_DIR) -> Path:
    """
    Ensure resolved path is inside base directory.
    Raises PermissionError on path traversal attempt.
    """
    abs_path = (base / path).resolve()
    abs_base = base.resolve()
    if not str(abs_path).startswith(str(abs_base)):
        raise PermissionError(f"Path traversal blocked: {path}")
    return abs_path


def safe_read(path: str, base: Path = BASE_DIR) -> str:
    fp = safe_path(path, base)
    with open(fp, "r", encoding="utf-8") as f:
        return f.read()


def safe_write(path: str, content: str, mode: str = "overwrite") -> dict:
    """
    Write file inside ai_workspace/ only.
    mode: overwrite | append
    """
    fp = safe_path(path, BASE_DIR)
    os.makedirs(fp.parent, exist_ok=True)
    if mode == "append":
        with open(fp, "a", encoding="utf-8") as f:
            f.write(content)
    else:
        with open(fp, "w", encoding="utf-8") as f:
            f.write(content)
    return {"ok": True, "bytes_written": len(content.encode("utf-8")), "path": path}


def parent_safe_read(relative_path: str) -> str:
    """Read file from parent directory (crap_dev/)."""
    fp = (PARENT_DIR / relative_path).resolve()
    if not str(fp).startswith(str(PARENT_DIR)):
        raise PermissionError("Access denied")
    with open(fp, "r", encoding="utf-8") as f:
        return f.read()


# ══════════════════════════════════════════════════════════════════════════
# TOOL REGISTRY
# ══════════════════════════════════════════════════════════════════════════
TOOL_REGISTRY: dict[str, callable] = {}


# ── 1. File Tools ─────────────────────────────────────────────────────────

def read_file(path: str) -> str:
    """Read a text file inside ai_workspace/."""
    return safe_read(path, BASE_DIR)


def write_file(path: str, content: str, mode: str = "overwrite") -> dict:
    """Write text content to a file inside ai_workspace/. mode: overwrite|append."""
    return safe_write(path, content, mode)


def list_dir(path: str = ".") -> dict:
    """List files and directories at a path inside ai_workspace/."""
    fp = safe_path(path, BASE_DIR)
    if not fp.exists():
        return {"error": f"Path not found: {path}"}
    files, dirs = [], []
    for item in sorted(fp.iterdir()):
        (dirs if item.is_dir() else files).append(item.name)
    return {"path": path, "files": files, "dirs": dirs}


def find_files(pattern: str, root: str = ".") -> dict:
    """Find files matching a glob pattern inside ai_workspace/."""
    fp = safe_path(root, BASE_DIR)
    results = sorted([str(p.relative_to(fp)) for p in fp.rglob(pattern)])
    return {"pattern": pattern, "root": root, "files": results}


def read_json(path: str) -> dict:
    """Read and parse a JSON file inside ai_workspace/."""
    content = safe_read(path, BASE_DIR)
    return json.loads(content)


def read_excel(path: str, sheet: str = None) -> dict:
    """
    Read Excel file from File_sach/ (parent dir) and return as dict.
    path: relative to File_sach/ (e.g. 'Thang_04_2026.xlsx')
    sheet: optional sheet name filter
    """
    import openpyxl
    fp = PARENT_DIR / "File_sach" / path
    if not fp.exists():
        return {"error": f"File not found in File_sach/: {path}"}
    wb = openpyxl.load_workbook(fp, data_only=True, read_only=True)

    if sheet:
        names = [n for n in wb.sheetnames if sheet.lower() in n.lower()]
        sheets_to_read = names or [wb.sheetnames[0]]
    else:
        sheets_to_read = wb.sheetnames[:3]   # limit to first 3 sheets

    result = {}
    for shname in sheets_to_read:
        rows = list(wb[shname].iter_rows(values_only=True))
        if not rows:
            continue
        headers = [str(v).strip() if v is not None else "" for v in rows[0]]
        result[shname] = {
            "headers": headers,
            "rows": [[str(v) if v is not None else "" for v in row] for row in rows[1:20]],  # limit 20 rows
        }
    wb.close()
    return {"file": path, "sheets": list(result.keys()), "data": result}

TOOL_REGISTRY["read_file"]  = read_file
TOOL_REGISTRY["write_file"] = write_file
TOOL_REGISTRY["list_dir"]   = list_dir
TOOL_REGISTRY["find_files"] = find_files
TOOL_REGISTRY["read_json"]  = read_json
TOOL_REGISTRY["read_excel"] = read_excel


# ── Monthly / Data_thang Tools ────────────────────────────────────────────────

def _str(v):
    return str(v).strip() if v is not None else ""


def read_monthly_excel(file_name: str = None, sheet: str = None) -> dict:
    """
    Doc file Excel tu Data_thang/ (parent dir).
    file_name: ten file (vd: 'Thang_04_2026.xlsx'). Bo trong = file moi nhat.
    sheet: ten sheet tuy chon.
    """
    import openpyxl
    from glob import glob

    data_dir = PARENT_DIR / "Data_thang"
    if not data_dir.exists():
        return {"error": f"Data_thang/ not found at {data_dir}"}

    if file_name:
        fp = data_dir / file_name
        files = [fp] if fp.exists() else []
    else:
        files = sorted(data_dir.glob("*.xlsx"), key=lambda p: p.stat().st_mtime, reverse=True)
        files = files[:1] if files else []

    if not files:
        return {"error": f"No .xlsx files in Data_thang/"}

    fp = files[0]
    try:
        wb = openpyxl.load_workbook(fp, data_only=True, read_only=True)
    except Exception as e:
        return {"error": f"Cannot open {fp.name}: {e}"}

    if sheet:
        names = [n for n in wb.sheetnames if sheet.lower() in n.lower()]
        sheets_to_read = names or [wb.sheetnames[0]]
    else:
        sheets_to_read = wb.sheetnames[:3]

    result = {}
    for shname in sheets_to_read:
        rows = list(wb[shname].iter_rows(values_only=True))
        if not rows:
            continue
        headers = [_str(v) for v in rows[0]]
        result[shname] = {
            "headers": headers,
            "rows": [[_str(v) for v in row] for row in rows[1:50]],  # limit 50 rows
        }
    wb.close()
    return {"file": fp.name, "sheets": list(result.keys()), "data": result}


def analyze_monthly(file_name: str = None) -> dict:
    """
    Phan tich chi tiet mot file Data_thang/.
    Bao gom: tong don, loai lenh, vat lieu, top KH, top KTV.
    file_name: vd 'Thang_04_2026.xlsx'. Bo trong = file moi nhat.
    """
    import openpyxl
    from collections import Counter

    data_dir = PARENT_DIR / "Data_thang"
    if file_name:
        fp = data_dir / file_name
        files = [fp] if fp.exists() else []
    else:
        files = sorted(data_dir.glob("*.xlsx"), key=lambda p: p.stat().st_mtime, reverse=True)
        files = files[:1] if files else []

    if not files:
        return {"error": "No files in Data_thang/"}

    fp = files[0]
    wb = openpyxl.load_workbook(fp, data_only=True, read_only=True)

    # Sheet Don hang
    don_sheet = next((n for n in wb.sheetnames
                       if "đơn hàng" in n.lower() or "don hang" in n.lower()), None)
    td_sheet = next((n for n in wb.sheetnames
                     if "tiến độ" in n.lower() or "tien do" in n.lower()), None)

    orders = {}
    lk_count = Counter()
    kh_count = Counter()
    mat_count = Counter()
    sl_total = 0

    if don_sheet:
        rows = list(wb[don_sheet].iter_rows(values_only=True))
        h = [_str(v) for v in rows[0]]
        idx = {name: i for i, name in enumerate(h)}

        for row in rows[1:]:
            ma = _str(row[idx.get("Mã ĐH", -1)])
            if not ma or ma == "Mã ĐH":
                continue
            if ma in orders:
                continue

            kh = _str(row[idx.get("Khách hàng", -1)])
            ph = _str(row[idx.get("Phục hình", -1)])
            sl = int(row[idx.get("SL", -1)]) if _str(row[idx.get("SL", -1)]).isdigit() else 0
            gc = _str(row[idx.get("Ghi chú ĐP", -1) if "Ghi chú ĐP" in h else -1])

            orders[ma] = {"kh": kh, "ph": ph, "sl": sl, "gc": gc}
            sl_total += sl
            if kh:
                kh_count[kh] += 1

            gc_lower = gc.lower()
            if "sửa" in gc_lower or "làm tiếp" in gc_lower:
                lk_count["Sửa/Làm tiếp"] += 1
            elif "bh" in gc_lower:
                lk_count["Bảo hành"] += 1
            elif "ts" in gc_lower or "thử sườn" in gc_lower:
                lk_count["Thử sườn"] += 1
            else:
                lk_count["Làm mới"] += 1

            ph_lower = ph.lower()
            if "zirc" in ph_lower:
                mat_count["Zirconia"] += sl
            elif "kim loại" in ph_lower:
                mat_count["Kim loại"] += sl
            elif "vene" in ph_lower:
                mat_count["Veneer"] += sl
            elif "temp" in ph_lower or "pmma" in ph_lower:
                mat_count["Temp/PMMA"] += sl
            else:
                mat_count["Khác"] += sl

    # Sheet Tien do — top KTV
    ktv_count = Counter()
    if td_sheet:
        rows = list(wb[td_sheet].iter_rows(values_only=True))
        h = [_str(v) for v in rows[0]]
        idx = {name: i for i, name in enumerate(h)}
        for row in rows[1:]:
            ma = _str(row[idx.get("Mã ĐH", -1)])
            ktv = _str(row[idx.get("KTV", -1)]).replace("-", "").strip()
            xn = _str(row[idx.get("Xác nhận", -1)]) == "Có"
            if ma and ma != "Mã ĐH" and xn and ktv:
                ktv_count[ktv] += 1

    wb.close()

    total = len(orders)
    return {
        "ok": True,
        "file": fp.name,
        "total_orders": total,
        "total_teeth": sl_total,
        "order_types": [{"type": k, "count": v, "pct": round(v * 100 / total, 1)} for k, v in lk_count.most_common()],
        "materials": [{"material": k, "teeth": v} for k, v in mat_count.most_common()],
        "top_customers": [{"name": k, "count": v} for k, v in kh_count.most_common(10)],
        "top_ktv": [{"name": k, "count": v} for k, v in ktv_count.most_common(10)],
    }


def compare_months(month1: str = "Thang_03_2026.xlsx",
                   month2: str = "Thang_04_2026.xlsx") -> dict:
    """
    So sanh 2 thang tu Data_thang/.
    Tra ve: don, rang, loai lenh, vat lieu, top KH thay doi.
    """
    r1 = analyze_monthly(month1)
    r2 = analyze_monthly(month2)

    if not r1.get("ok") or not r2.get("ok"):
        return {"error": "Could not load one or both months"}

    def diff(name, v1, v2):
        d = v2 - v1
        pct = round(d / v1 * 100, 1) if v1 else 0
        return {"name": name, "before": v1, "after": v2, "diff": d, "diff_pct": pct}

    changes = {
        "total_orders": diff("Đơn hàng", r1["total_orders"], r2["total_orders"]),
        "total_teeth": diff("Răng", r1["total_teeth"], r2["total_teeth"]),
    }

    # Top KH changes
    kh1 = {x["name"]: x["count"] for x in r1.get("top_customers", [])}
    kh2 = {x["name"]: x["count"] for x in r2.get("top_customers", [])}
    all_kh = set(kh1) | set(kh2)
    kh_changes = []
    for kh in sorted(all_kh, key=lambda x: -(kh2.get(x, 0) - kh1.get(x, 0))):
        c1 = kh1.get(kh, 0)
        c2 = kh2.get(kh, 0)
        if c2 != c1:
            kh_changes.append({"name": kh, "before": c1, "after": c2, "diff": c2 - c1})
    kh_changes.sort(key=lambda x: -x["diff"])

    return {
        "ok": True,
        "month1": {"file": r1["file"], "orders": r1["total_orders"], "teeth": r1["total_teeth"]},
        "month2": {"file": r2["file"], "orders": r2["total_orders"], "teeth": r2["total_teeth"]},
        "changes": changes,
        "kh_changes": kh_changes[:10],
        "month1_order_types": r1["order_types"],
        "month2_order_types": r2["order_types"],
    }


# ── 2. Analysis Tools ──────────────────────────────────────────────────────

def analyze_excel(file_path: str = None, all_files: bool = False) -> dict:
    """
    Run full Excel analysis on File_sach/*.xlsx files.
    Reuses ai_stats.py logic. Returns structured stats dict.
    """
    import openpyxl
    from collections import defaultdict

    data_dir = PARENT_DIR / "File_sach"

    if all_files:
        files = sorted(data_dir.glob("*.xlsx"), key=lambda p: p.stat().st_mtime, reverse=True)[:3]
    elif file_path:
        fp = data_dir / file_path
        files = [fp] if fp.exists() else []
    else:
        files = [max(data_dir.glob("*.xlsx"), default=None, key=lambda p: p.stat().st_mtime)]
        files = [f for f in files if f]

    if not files:
        return {"error": "No Excel files found in File_sach/"}

    def str_(v):
        return str(v).strip() if v is not None else ""

    # Parse all files
    all_orders = {}
    stage_map = defaultdict(lambda: defaultdict(list))

    for fp in files:
        try:
            wb = openpyxl.load_workbook(fp, data_only=True, read_only=True)

            # Sheet Don hang
            don_sheet = next((n for n in wb.sheetnames
                             if "đơn hàng" in n.lower() or "don hang" in n.lower()), None)
            if don_sheet:
                rows = list(wb[don_sheet].iter_rows(values_only=True))
                h = [str_(v) for v in rows[0]]
                idx = {name: i for i, name in enumerate(h)}
                for row in rows[1:]:
                    ma = str_(row[idx.get("Mã ĐH", -1)])
                    if not ma or ma == "Mã ĐH":
                        continue
                    if ma in all_orders:
                        continue
                    kh = str_(row[idx.get("Khách", -1)])
                    bn = str_(row[idx.get("ệnh nhân", -1)])
                    ph = str_(row[idx.get("Phục hình", -1)])
                    sl = int(row[idx.get("SL", -1)]) if str_(row[idx.get("SL", -1)]).isdigit() else 0
                    gc = str_(row[idx.get("Ghi chú", -1)])
                    lk = str_(row[idx.get("Loại lệnh", -1)])
                    all_orders[ma] = {"kh": kh, "bn": bn, "ph": ph, "sl": sl, "gc": gc, "lk": lk}

            # Sheet Tien do
            td_sheet = next((n for n in wb.sheetnames
                            if "tiến độ" in n.lower() or "tien do" in n.lower()), None)
            if td_sheet:
                rows = list(wb[td_sheet].iter_rows(values_only=True))
                h = [str_(v) for v in rows[0]]
                idx = {name: i for i, name in enumerate(h)}
                for row in rows[1:]:
                    ma = str_(row[idx.get("Mã ĐH", -1)])
                    if not ma or ma == "Mã ĐH":
                        continue
                    cd = str_(row[idx.get("Công đoạn", -1)])
                    ktv = str_(row[idx.get("KTV", -1)]).replace("-", "").strip()
                    xn = str_(row[idx.get("Xác nhận", -1)]) == "Có"
                    tg = str_(row[idx.get("Thời gian", -1)]).replace("-", "").strip()
                    if ma not in all_orders:
                        all_orders[ma] = {"kh": "", "bn": "", "ph": "", "sl": 0, "gc": "", "lk": ""}
                    stage_map[ma][cd] = {"ktv": ktv, "xn": xn, "tg": tg}

            wb.close()
        except Exception as e:
            continue

    # Build stats
    orders = list(all_orders.keys())
    total_orders = len(orders)
    total_teeth = sum(all_orders[m]["sl"] for m in orders)

    # Bottleneck
    STAGE_NAMES = ["CBM", "SÁP/Cadcam", "SƯỜN", "ĐẮP", "MÀI"]
    stage_counts = defaultdict(int)
    for ma in orders:
        sm = stage_map.get(ma, {})
        confirmed_stage = None
        for cd in STAGE_NAMES:
            if sm.get(cd, {}).get("xn"):
                confirmed_stage = cd
        if confirmed_stage:
            stage_idx = STAGE_NAMES.index(confirmed_stage)
            next_stage = STAGE_NAMES[stage_idx + 1] if stage_idx + 1 < len(STAGE_NAMES) else None
            if next_stage:
                stage_counts[next_stage] += 1

    bottleneck = max(stage_counts, key=stage_counts.get, default="MÀI")
    bottleneck_pct = round(stage_counts[bottleneck] / total_orders * 100, 1) if total_orders else 0

    # Top KTV
    ktv_counts = defaultdict(int)
    for ma in orders:
        sm = stage_map.get(ma, {})
        for cd in STAGE_NAMES:
            s = sm.get(cd, {})
            if s.get("xn") and s.get("ktv"):
                ktv_counts[s["ktv"]] += 1
    top_ktv = sorted(ktv_counts.items(), key=lambda x: -x[1])[:5]

    # Top customers
    kh_counts = defaultdict(int)
    for ma in orders:
        kh = all_orders[ma]["kh"]
        if kh:
            kh_counts[kh] += 1
    top_kh = sorted(kh_counts.items(), key=lambda x: -x[1])[:5]

    return {
        "ok": True,
        "files_analyzed": [str(f.name) for f in files],
        "total_orders": total_orders,
        "total_teeth": total_teeth,
        "bottleneck": bottleneck,
        "bottleneck_pct": bottleneck_pct,
        "pending_by_stage": dict(stage_counts),
        "top_ktv": [{"name": k, "count": v} for k, v in top_ktv],
        "top_customers": [{"name": k, "count": v} for k, v in top_kh],
    }


def analyze_lead_times(file_path: str = None) -> dict:
    """
    Analyze lead time (thoi gian) per stage from Excel.
    Returns avg days per stage and overdue orders.
    """
    import openpyxl
    from datetime import datetime

    data_dir = PARENT_DIR / "File_sach"
    files = sorted(data_dir.glob("*.xlsx"), key=lambda p: p.stat().st_mtime, reverse=True)
    if file_path:
        fp = data_dir / file_path
        files = [fp] if fp.exists() else []

    def str_(v):
        return str(v).strip() if v is not None else ""

    def parse_dt(s):
        if not s:
            return None
        for fmt in ["%Y-%m-%d %H:%M:%S", "%d/%m/%Y %H:%M:%S", "%Y-%m-%d"]:
            try:
                return datetime.strptime(s.strip()[:19], fmt)
            except Exception:
                pass
        return None

    stage_times = {"CBM": [], "SÁP/Cadcam": [], "SƯỜN": [], "ĐẮP": [], "MÀI": []}
    overdue = []

    for fp in files:
        try:
            wb = openpyxl.load_workbook(fp, data_only=True, read_only=True)
            td_sheet = next((n for n in wb.sheetnames
                            if "tiến độ" in n.lower()), None)
            if td_sheet:
                rows = list(wb[td_sheet].iter_rows(values_only=True))
                h = [str_(v) for v in rows[0]]
                idx = {name: i for i, name in enumerate(h)}
                for row in rows[1:]:
                    ma  = str_(row[idx.get("Mã ĐH", -1)])
                    cd  = str_(row[idx.get("Công đoạn", -1)])
                    tg  = parse_dt(str_(row[idx.get("Thời gian", -1)]))
                    if ma and ma != "Mã ĐH" and cd in stage_times and tg:
                        stage_times[cd].append(tg)
            wb.close()
        except Exception:
            continue

    result = {}
    for cd, times in stage_times.items():
        if len(times) >= 2:
            times.sort()
            spans = [(times[i+1] - times[i]).total_seconds() / 86400 for i in range(len(times)-1)]
            result[cd] = {"avg_days": round(sum(spans)/len(spans), 1), "samples": len(spans)}
        else:
            result[cd] = {"avg_days": 0, "samples": len(times)}

    return {"ok": True, "lead_times": result, "overdue_count": len(overdue)}


def analyze_ktv(file_path: str = None) -> dict:
    """Rank KTV performance by confirmed stage count."""
    stats = analyze_excel(file_path=file_path)
    ranked = sorted(stats.get("top_ktv", []), key=lambda x: -x["count"])
    return {"ok": True, "ranked_ktv": ranked}


def analyze_customers(file_path: str = None) -> dict:
    """Rank customers by order volume."""
    stats = analyze_excel(file_path=file_path)
    ranked = sorted(stats.get("top_customers", []), key=lambda x: -x["count"])
    return {"ok": True, "ranked_customers": ranked}


TOOL_REGISTRY["analyze_excel"]       = analyze_excel
TOOL_REGISTRY["analyze_lead_times"]  = analyze_lead_times
TOOL_REGISTRY["analyze_ktv"]         = analyze_ktv
TOOL_REGISTRY["analyze_customers"]   = analyze_customers


# ── 3. Report Tools ────────────────────────────────────────────────────────

def generate_report(report_type: str = "daily_summary",
                    period: str = "latest",
                    file_path: str = None) -> dict:
    """
    Generate a text report and save to analysis/reports/.
    report_type: daily_summary | ktv_report | customer_report | lead_time_report
    """
    import datetime as dt

    # Gather data
    if report_type == "daily_summary":
        stats = analyze_excel(file_path=file_path)
        content = [
            f"=== BAO CAO TONG HOP — {datetime.now().strftime('%d/%m/%Y %H:%M')} ===",
            f"",
            f"Tong don: {stats.get('total_orders','?')}",
            f"Tong rang: {stats.get('total_teeth','?')}",
            f"",
            f"Nut that: {stats.get('bottleneck','?')} ({stats.get('bottleneck_pct','?')}% don chua xong)",
            f"",
            f"Top 5 KTV:",
        ]
        for i, k in enumerate(stats.get("top_ktv", [])[:5], 1):
            content.append(f"  {i}. {k['name']}: {k['count']} don")
        content.append(f"")
        content.append(f"Top 5 Khach hang:")
        for i, c in enumerate(stats.get("top_customers", [])[:5], 1):
            content.append(f"  {i}. {c['name']}: {c['count']} don")
        content.append(f"")
        content.append(f"Nguon: analyze_excel tool | File: {', '.join(stats.get('files_analyzed',[]))}")

    elif report_type == "ktv_report":
        stats = analyze_ktv(file_path=file_path)
        content = [f"=== BAO CAO KTV — {datetime.now().strftime('%d/%m/%Y %H:%M')} ===", ""]
        for i, k in enumerate(stats.get("ranked_ktv", []), 1):
            content.append(f"  {i}. {k['name']}: {k['count']} don xac nhan")
        content.append("")

    elif report_type == "customer_report":
        stats = analyze_customers(file_path=file_path)
        content = [f"=== BAO CAO KHACH HANG — {datetime.now().strftime('%d/%m/%Y %H:%M')} ===", ""]
        for i, c in enumerate(stats.get("ranked_customers", []), 1):
            content.append(f"  {i}. {c['name']}: {c['count']} don")
        content.append("")

    elif report_type == "lead_time_report":
        lt = analyze_lead_times(file_path=file_path)
        content = [f"=== BAO CAO LEAD TIME — {datetime.now().strftime('%d/%m/%Y %H:%M')} ===", ""]
        for cd, data in lt.get("lead_times", {}).items():
            content.append(f"  {cd}: TB {data['avg_days']} ngay ({data['samples']} mau)")
        content.append("")

    else:
        return {"error": f"Unknown report_type: {report_type}"}

    report_text = "\n".join(content)
    report_dir = BASE_DIR / "analysis" / "reports"
    report_dir.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"{report_type}_{ts}.txt"
    saved_path = report_dir / filename
    with open(saved_path, "w", encoding="utf-8") as f:
        f.write(report_text)

    return {
        "ok": True,
        "report_type": report_type,
        "saved_to": str(saved_path.relative_to(BASE_DIR)),
        "preview": report_text[:500],
    }


def export_orders_csv(filter_type: str = "all") -> dict:
    """
    Export current orders as CSV to analysis/exports/.
    filter_type: all | overdue | today
    """
    stats = analyze_excel()
    rows = []
    for c in stats.get("top_customers", []):
        rows.append({"customer": c["name"], "count": c["count"]})

    export_dir = BASE_DIR / "analysis" / "exports"
    export_dir.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"export_{filter_type}_{ts}.csv"
    filepath = export_dir / filename

    lines = ["customer,count"]
    for r in rows:
        lines.append(f'"{r["customer"]}",{r["count"]}')
    with open(filepath, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))

    return {"ok": True, "saved_to": str(filepath.relative_to(BASE_DIR)), "rows": len(rows)}


TOOL_REGISTRY["generate_report"]  = generate_report
TOOL_REGISTRY["export_orders_csv"] = export_orders_csv


# ── 4. Memory & Workspace Tools ───────────────────────────────────────────

def get_lab_state() -> dict:
    """Get full lab state from shared parent files."""
    from sandbox_knowledge import refresh_knowledge, build_lab_context
    state = refresh_knowledge({})
    return {
        "knowledge_summary": build_lab_context(state),
        "loaded_at": state.get("loaded_at"),
        "latest_excel": state.get("latest_excel"),
    }


def add_learned_fact(fact: str, source: str = "sandbox",
                     confidence: str = "medium",
                     tags: list = None) -> dict:
    """Add a learned fact to sandbox_facts.json."""
    from sandbox_memory import add_fact
    entry = add_fact(fact, source, confidence, tags)
    return {"ok": True, "id": entry.get("id"), "fact": fact}


def save_task(project_id: str, description: str,
              priority: str = "medium") -> dict:
    """Create a new task in sandbox_tasks.json."""
    from sandbox_memory import create_task
    task = create_task(project_id or "default", description, priority)
    return {"ok": True, "id": task.get("id"), "description": description}


def list_tasks(project_id: str = None, status: str = None) -> dict:
    """List tasks from sandbox_tasks.json."""
    from sandbox_memory import list_tasks as lt
    tasks = lt(project_id=project_id, status=status)
    return {"ok": True, "tasks": tasks, "count": len(tasks)}


def save_project(name: str, description: str = "", tags: list = None) -> dict:
    """Create a new project in sandbox_projects.json."""
    from sandbox_memory import create_project
    proj = create_project(name, description, tags)
    return {"ok": True, "id": proj.get("id"), "name": name}


def list_projects(active_only: bool = True) -> dict:
    """List projects from sandbox_projects.json."""
    from sandbox_memory import list_projects as lp
    projects = lp(active_only=active_only)
    return {"ok": True, "projects": projects, "count": len(projects)}


TOOL_REGISTRY["get_lab_state"]    = get_lab_state
TOOL_REGISTRY["add_learned_fact"] = add_learned_fact
TOOL_REGISTRY["save_task"]       = save_task
TOOL_REGISTRY["list_tasks"]      = list_tasks
TOOL_REGISTRY["save_project"]    = save_project
TOOL_REGISTRY["list_projects"]   = list_projects


# ── 5. System Tools ────────────────────────────────────────────────────────

def get_status() -> dict:
    """Get sandbox status: uptime, sessions, memory stats."""
    from sandbox_memory import list_sessions, count_pending_tasks
    sessions = list_sessions()
    pending = count_pending_tasks()
    reports_dir = BASE_DIR / "analysis" / "reports"
    exports_dir  = BASE_DIR / "analysis" / "exports"
    return {
        "ok": True,
        "workspace": str(BASE_DIR.name),
        "parent_dir": str(PARENT_DIR.name),
        "active_sessions": len([s for s in sessions if s["msg_count"] > 0]),
        "total_sessions": len(sessions),
        "pending_tasks": pending["pending"],
        "done_tasks": pending["done"],
        "reports_count": len(list(reports_dir.glob("*.txt"))) if reports_dir.exists() else 0,
        "exports_count": len(list(exports_dir.glob("*.csv"))) if exports_dir.exists() else 0,
    }


def log_activity(action: str, detail: str = "") -> dict:
    """Log a custom action to activity.log."""
    from sandbox_logging import log_action
    log_action(action, {"detail": detail})
    return {"ok": True}


def search_logs(query: str = "", limit: int = 50) -> dict:
    """Search activity logs."""
    from sandbox_logging import search_logs
    entries = search_logs(query, limit)
    return {"ok": True, "entries": entries, "count": len(entries)}


def get_logs(limit: int = 20) -> dict:
    """Get recent activity log entries."""
    from sandbox_logging import get_recent_logs
    entries = get_recent_logs(limit)
    return {"ok": True, "entries": entries, "count": len(entries)}


TOOL_REGISTRY["get_status"]   = get_status
TOOL_REGISTRY["log_activity"] = log_activity
TOOL_REGISTRY["search_logs"]  = search_logs
TOOL_REGISTRY["get_logs"]     = get_logs
TOOL_REGISTRY["read_monthly_excel"] = read_monthly_excel
TOOL_REGISTRY["analyze_monthly"]     = analyze_monthly
TOOL_REGISTRY["compare_months"]      = compare_months
