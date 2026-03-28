"""
laboasia_gui_scraper_tkinter.py
────────────────────────────────
GUI app: scrape order progress from laboasia.com.vn, merge into Excel workbook.

Usage:
    python laboasia_gui_scraper_tkinter.py

Credentials are read from environment variables (recommended):
    export LABO_USER1=lanhn
    export LABO_PASS1=796803
    export LABO_USER2=hyct
    export LABO_PASS2=336876
    ...or use the GUI to enter them directly.

Requirements:
    pip install pandas openpyxl playwright
    playwright install chromium
"""

from __future__ import annotations

import json
import math
import os
import queue
import random
import re
import shutil
import subprocess
import sys
import threading
import time
from dataclasses import asdict, dataclass
from pathlib import Path
from typing import Optional

import requests

import pandas as pd
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from openpyxl import load_workbook
from playwright.sync_api import TimeoutError as PlaywrightTimeoutError
from playwright.sync_api import sync_playwright


# ═══════════════════════════════════════════════════════════════════════════════
# CONFIG FILE PATH
# ═══════════════════════════════════════════════════════════════════════════════

def _get_base_dir() -> Path:
    """Get the directory of this script, with fallback for edge cases."""
    try:
        return Path(__file__).parent.resolve()
    except NameError:
        # __file__ not defined (e.g., python -c)
        return Path.cwd().resolve()


BASE_DIR = _get_base_dir()
CONFIG_FILE = BASE_DIR / "labo_config.json"

# ── Hardcoded directories ──────────────────────────────────────────────────────
# File gốc (chỉ đọc, không xuất ra)
INPUT_DIR  = BASE_DIR / "Excel"
# File đã scrape + JSON + failed (trung gian)
DATA_DIR   = BASE_DIR / "Data"
# File sạch (sau khi chạy labo_cleaner)
CLEAN_DIR  = BASE_DIR / "File_sach"
# File tích lũy theo tháng (chu kỳ 26→25)
DATA_THANG_DIR = BASE_DIR / "Data_thang"

# Tự tạo thư mục nếu chưa có
DATA_DIR.mkdir(exist_ok=True)
CLEAN_DIR.mkdir(exist_ok=True)
DATA_THANG_DIR.mkdir(exist_ok=True)

# Endpoint JSON API của laboasia
API_ENDPOINT = "https://laboasia.com.vn/empconnect/api_handler/"


# ═══════════════════════════════════════════════════════════════════════════════
# DATA MODEL
# ═══════════════════════════════════════════════════════════════════════════════

@dataclass
class ProgressRow:
    ma_dh: str
    thu_tu: int
    cong_doan: str
    ten_ktv: str
    xac_nhan: str
    thoi_gian_hoan_thanh: str
    raw_row_text: str = ""
    tai_khoan_cao: str = ""


# ═══════════════════════════════════════════════════════════════════════════════
# NORMALIZE HELPERS
# ═══════════════════════════════════════════════════════════════════════════════

def normalize_text(s: Optional[str]) -> str:
    """Normalize text: handle None, non-breaking spaces, extra whitespace."""
    if s is None:
        return ""
    s = str(s).replace("\xa0", " ")
    s = re.sub(r"\s+", " ", s).strip()
    return s


def normalize_ma_dh(v) -> str:
    """Normalize an order-ID value (strip, fix float strings like '123.0')."""
    s = normalize_text(v)
    if not s:
        return ""
    if re.fullmatch(r"\d+\.0", s):
        s = s[:-2]
    return s


# ═══════════════════════════════════════════════════════════════════════════════
# SELECTORS
# ═══════════════════════════════════════════════════════════════════════════════

DEFAULT_SELECTORS: dict[str, list[str]] = {
    "username": [
        'input[name="username"]',
        'input[type="text"]',
        'input[placeholder*="user"]',
        'input[placeholder*="User"]',
        'input[placeholder*="tên đăng nhập"]',
        'input[placeholder*="Tên đăng nhập"]',
    ],
    "password": [
        'input[name="password"]',
        'input[type="password"]',
        'input[placeholder*="mật khẩu"]',
        'input[placeholder*="Mật khẩu"]',
        'input[placeholder*="password"]',
        'input[placeholder*="Password"]',
    ],
    "login_button": [
        'button[type="submit"]',
        'button:has-text("Đăng nhập")',
        'button:has-text("Login")',
        'input[type="submit"]',
    ],
    "search_input": [
        'input[name="ma_dh"]',
        'input[placeholder*="Mã"]',
        'input[placeholder*="mã"]',
        'input[placeholder*="đơn"]',
        'input[placeholder*="Đơn"]',
        'input[type="search"]',
        'input[type="text"]',
    ],
    "search_button": [
        'button:has-text("Tìm")',
        'button:has-text("Tra")',
        'button:has-text("Search")',
        'button:has-text("Xem")',
        'button[type="submit"]',
    ],
    "reset_button": [
        'button:has-text("Xóa")',
        'button:has-text("Reset")',
        'button:has-text("Làm mới")',
    ],
    "table_rows": [
        "table tbody tr",
        ".table tbody tr",
    ],
    "cell_cong_doan": [
        'td[data-label*="Công đoạn"]',
        'td[data-label*="cong đoạn"]',
        'td[data-label*="Công Đoạn"]',
        'td[class*="cong-doan"]',
        'td[class*="step"]',
        "table tbody tr td:nth-child(2)",
        "table tbody tr td:nth-child(3)",
    ],
    "cell_nhan_vien": [
        'td[data-label*="Nhân viên"]',
        'td[data-label*="nhân viên"]',
        'td[data-label*="KTV"]',
        'td[data-label*="ktv"]',
        'td[class*="nhan-vien"]',
        'td[class*="ktv"]',
        "table tbody tr td:nth-child(3)",
        "table tbody tr td:nth-child(4)",
    ],
    "cell_xac_nhan": [
        'td[data-label*="Xác nhận"]',
        'td[data-label*="xác nhận"]',
        'td[data-label*="Trạng thái"]',
        'td[class*="xac-nhan"]',
        'td[class*="status"]',
        "table tbody tr td:nth-child(4)",
        "table tbody tr td:nth-child(5)",
    ],
    "cell_thoi_gian": [
        'td[data-label*="Thời gian"]',
        'td[data-label*="thời gian"]',
        'td[data-label*="Thành"]',
        'td[data-label*="Hoàn thành"]',
        'td[class*="time"]',
        'td[class*="thoi-gian"]',
        "table tbody tr td:nth-last-child(2)",
        "table tbody tr td:nth-last-child(1)",
    ],
}


# ═══════════════════════════════════════════════════════════════════════════════
# SCRAPER CORE
# ═══════════════════════════════════════════════════════════════════════════════

class LaboAsiaAPIClient:
    """
    Scraper mới — dùng HTTP JSON API thay vì Playwright.

    Flow:
      1. Login 1 lần bằng Playwright → lấy JWT từ cookie auth_token → đóng browser.
      2. Dùng requests.Session (Bearer JWT) để POST lên API_ENDPOINT cho từng đơn.
      3. Tự re-login nếu nhận HTTP 401 (token hết hạn sau ~4 giờ).
    """

    def __init__(
        self,
        base_url: str,
        username: str,
        password: str,
        selectors: dict[str, list[str]],
        page_timeout_ms: int = 12000,
        max_retry_per_order: int = 3,
        **_ignored,  # bỏ qua các tham số cũ không còn dùng (wait_after, save_debug…)
    ):
        self.base_url = base_url.strip()
        self.username = username.strip()
        self.password = password
        self.selectors = selectors
        self.page_timeout_ms = page_timeout_ms
        self.max_retry_per_order = max_retry_per_order
        self._token: str = ""
        self._session: Optional[requests.Session] = None

    # ── Playwright login (chỉ dùng 1 lần để lấy JWT) ────────────────────────

    def _first_visible(self, page, selector_list: list[str], timeout: int = 4000):
        """Return the first visible locator from a list of selectors."""
        last_error = None
        for selector in selector_list:
            try:
                loc = page.locator(selector).first
                loc.wait_for(state="visible", timeout=timeout)
                return loc
            except Exception as e:
                last_error = e
        raise RuntimeError(
            f"Khong tim thay selector: {selector_list}. Loi: {last_error}"
        )

    def _playwright_login(self) -> str:
        """
        Mở browser ẩn, đăng nhập, lấy JWT từ cookie auth_token, đóng browser.
        Trả về token string.
        """
        with sync_playwright() as p:
            browser = p.chromium.launch(headless=True, slow_mo=0)
            context = browser.new_context(
                viewport={"width": 1440, "height": 900},
                ignore_https_errors=True,
            )
            page = context.new_page()
            page.set_default_timeout(self.page_timeout_ms)
            try:
                page.goto(self.base_url, wait_until="domcontentloaded")

                user_input = self._first_visible(page, self.selectors["username"], timeout=8000)
                pass_input = self._first_visible(page, self.selectors["password"], timeout=8000)
                user_input.fill(self.username)
                pass_input.fill(self.password)

                btn = self._first_visible(page, self.selectors["login_button"], timeout=8000)
                btn.click()
                page.wait_for_load_state("networkidle")

                # Lấy token từ cookie
                cookies = context.cookies()
                token = next(
                    (c["value"] for c in cookies if c["name"] == "auth_token"),
                    "",
                )
                if not token:
                    raise RuntimeError(
                        "Dang nhap thanh cong nhung khong tim thay cookie auth_token"
                    )
                return token
            finally:
                browser.close()

    # ── HTTP session ─────────────────────────────────────────────────────────

    def _build_session(self, token: str) -> requests.Session:
        """Tạo requests.Session với JWT header."""
        sess = requests.Session()
        sess.headers.update({
            "Authorization": f"Bearer {token}",
            "Content-Type": "application/json",
            "Accept": "*/*",
            "Origin": "https://laboasia.com.vn",
            "Referer": "https://laboasia.com.vn/scan/",
        })
        return sess

    def _relogin(self) -> None:
        """Re-login (token hết hạn) và cập nhật session."""
        self._token = self._playwright_login()
        self._session = self._build_session(self._token)

    # ── API call ─────────────────────────────────────────────────────────────

    def _get_order_info(self, ma_dh: str) -> Optional[dict]:
        """
        POST {"action": "get_order_info", "barcode": ma_dh}.
        Trả về dict JSON, hoặc None nếu 401 (cần re-login).
        """
        assert self._session is not None
        payload = {"action": "get_order_info", "barcode": ma_dh}
        resp = self._session.post(API_ENDPOINT, json=payload, timeout=15)
        if resp.status_code == 401:
            return None  # token hết hạn
        resp.raise_for_status()
        return resp.json()

    # ── JSON → ProgressRow ───────────────────────────────────────────────────

    def _parse_rows(self, ma_dh: str, data: dict) -> list[ProgressRow]:
        """Chuyển SanXuatInfo[] trong JSON thành list[ProgressRow]."""
        rows: list[ProgressRow] = []
        for item in data.get("SanXuatInfo", []):
            tg_ht = normalize_text(item.get("thoigianbatdau", ""))  # web hiển thị thời gian bắt đầu
            sanpham = normalize_text(item.get("sanpham", ""))
            soluong = item.get("soluong", 0)
            rang = normalize_text(item.get("rang", ""))
            loaisanpham = normalize_text(item.get("loaisanpham", ""))
            # raw_row_text chứa đủ thông tin để labo_cleaner trích xuất SL/phục hình
            raw = f"{sanpham} SL:{soluong} Rang:{rang} {loaisanpham}".strip()
            rows.append(ProgressRow(
                ma_dh=ma_dh,
                thu_tu=int(item.get("thutulam", len(rows) + 1)),
                cong_doan=normalize_text(item.get("congdoan", "")),
                ten_ktv=normalize_text(item.get("tennhanvien", "")),
                xac_nhan="Có" if tg_ht and tg_ht not in ("", "-") else "Chưa",
                thoi_gian_hoan_thanh=tg_ht,
                raw_row_text=raw,
                tai_khoan_cao=self.username,
            ))
        return rows

    # ── Retry logic ──────────────────────────────────────────────────────────

    def _scrape_one_order(self, ma_dh: str) -> list[ProgressRow]:
        """Gọi API cho 1 đơn, có retry + re-login khi 401."""
        for attempt in range(1, self.max_retry_per_order + 1):
            try:
                data = self._get_order_info(ma_dh)
                if data is None:
                    # Token hết hạn → re-login và thử lại
                    self._relogin()
                    continue
                return self._parse_rows(ma_dh, data)
            except Exception:
                if attempt == self.max_retry_per_order:
                    raise
                time.sleep(1.5 ** attempt)
        raise RuntimeError(f"That bai sau {self.max_retry_per_order} lan thu")

    # ── Entry point ───────────────────────────────────────────────────────────

    def scrape_order_list(
        self, order_ids: list[str], event_q: queue.Queue, worker_name: str
    ) -> None:
        """
        Cào danh sách đơn hàng qua API JSON. Gửi events lên event_q cho GUI.
        Events:
          - ("log", message)
          - ("order_done", worker_name, ma_dh, row_count, error_or_none)
          - ("worker_finished", worker_name, results, failed, username)
        """
        results: list[ProgressRow] = []
        failed: list[str] = []

        # Bước 1: login 1 lần để lấy JWT
        try:
            self._token = self._playwright_login()
            self._session = self._build_session(self._token)
        except Exception as e:
            event_q.put(("log", f"[{worker_name}] Login that bai: {e}"))
            event_q.put(("worker_finished", worker_name, [], list(order_ids), self.username))
            return

        event_q.put(("log", f"[{worker_name}] Dang nhap OK: {self.username}"))
        total = len(order_ids)

        # Bước 2: cào từng đơn bằng HTTP
        for idx, ma_dh in enumerate(order_ids, start=1):
            event_q.put(("log", f"[{worker_name}] {idx}/{total} → {ma_dh}"))
            try:
                rows = self._scrape_one_order(ma_dh)
                results.extend(rows)
                event_q.put(("order_done", worker_name, ma_dh, len(rows), None))
            except Exception as e:
                failed.append(ma_dh)
                event_q.put(("order_done", worker_name, ma_dh, 0, str(e)))

        event_q.put(("worker_finished", worker_name, results, failed, self.username))


# ═══════════════════════════════════════════════════════════════════════════════
# EXCEL / MERGE
# ═══════════════════════════════════════════════════════════════════════════════

def load_order_ids(xlsx_path: str, sheet_name: str, column_name: str) -> list[str]:
    """
    Load order IDs from an Excel column, deduplicate, and return in order.
    Raises clear ValueError/KeyError if sheet or column not found.
    """
    xl = pd.ExcelFile(xlsx_path)
    if sheet_name not in xl.sheet_names:
        raise ValueError(
            f"Khong tim thay sheet '{sheet_name}' trong file Excel.\n"
            f"Cac sheet co trong file: {xl.sheet_names}"
        )
    df = xl.parse(sheet_name=sheet_name)
    if column_name not in df.columns:
        raise KeyError(
            f"Khong thay cot '{column_name}' trong sheet '{sheet_name}'.\n"
            f"Co cac cot: {list(df.columns)}"
        )

    ids = [normalize_ma_dh(v) for v in df[column_name].tolist()]
    ids = [x for x in ids if x]

    seen: set = set()
    uniq: list[str] = []
    for x in ids:
        if x not in seen:
            uniq.append(x)
            seen.add(x)
    return uniq


def build_progress_df(progress_rows: list[ProgressRow]) -> pd.DataFrame:
    """Convert a list of ProgressRow into a pandas DataFrame."""
    if not progress_rows:
        return pd.DataFrame(columns=[
            "ma_dh", "thu_tu", "cong_doan", "ten_ktv", "xac_nhan",
            "thoi_gian_hoan_thanh", "raw_row_text", "tai_khoan_cao",
        ])
    return pd.DataFrame([asdict(r) for r in progress_rows])


def build_order_summary(progress_df: pd.DataFrame) -> pd.DataFrame:
    """Build a per-order summary DataFrame with aggregated KTV and stages."""
    if progress_df.empty:
        return pd.DataFrame(columns=[
            "ma_dh", "so_cd_cao_web", "ktv_theo_cong_doan_live",
            "ds_cong_doan_live", "lan_cap_nhat_cao_web", "tai_khoan_cao",
        ])

    def agg_ktv(g: pd.DataFrame) -> str:
        pairs: list[str] = []
        for _, r in g.iterrows():
            cd = normalize_text(r.get("cong_doan"))
            tv = normalize_text(r.get("ten_ktv"))
            xn = normalize_text(r.get("xac_nhan"))
            if not cd and not tv:
                continue
            # Có đủ cả KTV + xác nhận "Có" → hiển thị "Có", ngược lại → "Chưa"
            if tv and tv not in ("-", "") and xn == "Có":
                xn_out = "Có"
            else:
                xn_out = "Chưa"
            pairs.append(f"{cd}: {xn_out}".strip(": "))
        seen: set = set()
        out: list[str] = []
        for x in pairs:
            if x and x not in seen:
                out.append(x)
                seen.add(x)
        return " | ".join(out)

    def agg_cd(g: pd.DataFrame) -> str:
        vals: list[str] = []
        for x in g["cong_doan"].tolist():
            x = normalize_text(x)
            if x and x not in vals:
                vals.append(x)
        return " -> ".join(vals)

    def agg_accounts(g: pd.DataFrame) -> str:
        vals: list[str] = []
        for x in g["tai_khoan_cao"].tolist():
            x = normalize_text(x)
            if x and x not in vals:
                vals.append(x)
        return " | ".join(vals)

    summary = progress_df.groupby("ma_dh", as_index=False).apply(
        lambda g: pd.Series({
            "so_cd_cao_web": len(g),
            "ktv_theo_cong_doan_live": agg_ktv(g),
            "ds_cong_doan_live": agg_cd(g),
            "lan_cap_nhat_cao_web": pd.Timestamp.now().strftime("%Y-%m-%d %H:%M:%S"),
            "tai_khoan_cao": agg_accounts(g),
        })
    ).reset_index(drop=True)

    return summary


def _autosize_sheet(ws, max_width: int = 45) -> None:
    """Auto-size worksheet columns based on content."""
    from openpyxl.utils import get_column_letter

    if ws.max_column < 1:
        return

    for col_cells in ws.iter_cols():
        cells = list(col_cells)
        if not cells:
            continue
        letter = get_column_letter(cells[0].column)
        width = 10
        for cell in cells[:200]:  # Sample first 200 rows
            val = "" if cell.value is None else str(cell.value)
            width = max(width, min(max_width, len(val) + 2))
        ws.column_dimensions[letter].width = width


def merge_back_to_workbook(
    input_xlsx: str, output_xlsx: str, progress_df: pd.DataFrame
) -> None:
    """
    Load the original workbook, add/update scraped data sheets,
    write the result to output_xlsx.
    """
    wb = load_workbook(input_xlsx)

    # ── Cao_web_live sheet (full scraped data) ─────────────────────────────
    if "Cao_web_live" in wb.sheetnames:
        del wb["Cao_web_live"]
    ws_raw = wb.create_sheet("Cao_web_live")
    raw_headers = list(progress_df.columns) if not progress_df.empty else [
        "ma_dh", "thu_tu", "cong_doan", "ten_ktv", "xac_nhan",
        "thoi_gian_hoan_thanh", "raw_row_text", "tai_khoan_cao",
    ]
    ws_raw.append(raw_headers)
    if not progress_df.empty:
        for row in progress_df[raw_headers].itertuples(index=False, name=None):
            ws_raw.append(list(row))
    _autosize_sheet(ws_raw)

    # ── Summary columns in Tong_hop_don_hang & Kanban_ready ────────────────
    summary_df = build_order_summary(progress_df)

    target_sheets = [s for s in ["Tong_hop_don_hang", "Kanban_ready"] if s in wb.sheetnames]
    extra_cols = [
        "ktv_theo_cong_doan_live",
        "ds_cong_doan_live",
        "so_cd_cao_web",
        "lan_cap_nhat_cao_web",
        "tai_khoan_cao",
    ]

    for sheet_name in target_sheets:
        ws = wb[sheet_name]
        header_map: dict[str, int] = {}
        for c in range(1, ws.max_column + 1):
            val = ws.cell(1, c).value
            if val:
                header_map[str(val)] = c

        if "ma_dh" not in header_map:
            continue

        # Add missing columns
        for col_name in extra_cols:
            if col_name not in header_map:
                header_map[col_name] = ws.max_column + 1
                ws.cell(1, header_map[col_name]).value = col_name

        lookup: dict = {}
        for _, r in summary_df.iterrows():
            lookup[normalize_ma_dh(r["ma_dh"])] = r.to_dict()

        for row_idx in range(2, ws.max_row + 1):
            order_id = normalize_ma_dh(ws.cell(row_idx, header_map["ma_dh"]).value)
            item = lookup.get(order_id)
            if not item:
                continue
            for col_name in extra_cols:
                ws.cell(row_idx, header_map[col_name]).value = item.get(col_name, "")
        _autosize_sheet(ws)

    # ── Cong_doan_chi_tiet_live sheet ───────────────────────────────────────
    if "Cong_doan_chi_tiet_live" in wb.sheetnames:
        del wb["Cong_doan_chi_tiet_live"]
    ws_cd = wb.create_sheet("Cong_doan_chi_tiet_live")
    live_headers = [
        "ma_dh", "thu_tu", "cong_doan", "ten_ktv", "xac_nhan",
        "thoi_gian_hoan_thanh", "raw_row_text", "tai_khoan_cao",
    ]
    ws_cd.append(live_headers)
    if not progress_df.empty:
        for row in progress_df[live_headers].itertuples(index=False, name=None):
            ws_cd.append(list(row))
    _autosize_sheet(ws_cd)

    wb.save(output_xlsx)


def save_json(progress_df: pd.DataFrame, json_path: str) -> None:
    """Save scraped data as a JSON file."""
    records = progress_df.to_dict(orient="records")
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(records, f, ensure_ascii=False, indent=2)


# ═══════════════════════════════════════════════════════════════════════════════
# CREDENTIAL MANAGEMENT
# ═══════════════════════════════════════════════════════════════════════════════

def _env(key: str, fallback: str = "") -> str:
    """Read from environment variable with fallback."""
    return os.environ.get(key, fallback)


def load_credentials_from_env() -> dict[int, dict[str, str]]:
    """
    Load credentials from environment variables.
    Returns {1: {user, pass}, 2: {...}, ...} for up to 4 accounts.
    Environment variables: LABO_USER1, LABO_PASS1, LABO_USER2, LABO_PASS2, ...
    """
    creds: dict[int, dict[str, str]] = {}
    for i in range(1, 5):
        user = _env(f"LABO_USER{i}").strip()
        pw = _env(f"LABO_PASS{i}").strip()
        if user and pw:
            creds[i] = {"user": user, "pass": pw}
    return creds


# ═══════════════════════════════════════════════════════════════════════════════
# GUI
# ═══════════════════════════════════════════════════════════════════════════════

class App:
    def __init__(self, root: tk.Tk):
        self.root = root
        self.root.title("LaboAsia Scan Scraper GUI")
        self.root.geometry("1080x880")

        # Event queue for thread-safe GUI updates
        self.event_q: queue.Queue[object] = queue.Queue()

        # State
        self.running = False
        self.total_orders = 0
        self.done_orders = 0
        self.ok_orders = 0
        self.fail_orders = 0
        self.worker_progress: dict[str, int] = {}
        self.final_results: list[ProgressRow] = []
        self.final_failed: list[str] = []
        self.finished_workers = 0
        self._cfg: Optional[dict] = None  # set during run_job
        self._last_run_file: Optional[str] = None  # path of last successfully scraped file
        self._auto_scheduled = False  # prevent multiple scheduled loops
        self._server_process: Optional[subprocess.Popen] = None  # node server process

        self._build_ui()
        self._load_config()
        self._load_credentials_from_env()
        self._start_server()
        self.root.after(300, self.process_events)
        self.root.after(800, self._auto_launch)

    # ── UI Build ─────────────────────────────────────────────────────────────

    def _build_ui(self):
        pad = {"padx": 8, "pady": 6}
        frm = ttk.Frame(self.root)
        frm.pack(fill="both", expand=True)

        # ── Config section ────────────────────────────────────────────────────
        cfg = ttk.LabelFrame(frm, text="Cau hinh")
        cfg.pack(fill="x", padx=10, pady=8)

        self.base_url_var = tk.StringVar(value="https://laboasia.com.vn/scan")
        self.excel_path_var = tk.StringVar()
        self.output_path_var = tk.StringVar()
        self.sheet_var = tk.StringVar()
        self.col_var = tk.StringVar()
        self.debug_dir_var = tk.StringVar()

        ttk.Label(cfg, text="Trang web").grid(row=0, column=0, sticky="w", **pad)
        ttk.Entry(cfg, textvariable=self.base_url_var, width=80).grid(
            row=0, column=1, columnspan=3, sticky="ew", **pad
        )

        # Thong tin thu muc (chi hien, khong cho chinh sua tay)
        info_text = (
            "  📂 Excel goc: Excel/  |  "
            "📁 Data: Data/  |  "
            "🧹 File sach: File_sach/"
        )
        ttk.Label(cfg, text=info_text, foreground="gray").grid(
            row=1, column=0, columnspan=4, sticky="w", **pad
        )

        # ── Dem nguoc auto-watch ───────────────────────────────────────────
        self._countdown_seconds = 0
        self._countdown_after_id: Optional[str] = None
        self.countdown_var = tk.StringVar(value="")
        countdown_lbl = ttk.Label(
            cfg, textvariable=self.countdown_var,
            foreground="blue", font=("Arial", 9, "bold"),
        )
        countdown_lbl.grid(row=2, column=0, columnspan=4, sticky="w", **pad)

        for i in range(4):
            cfg.columnconfigure(i, weight=1)

        # ── Account credentials ───────────────────────────────────────────────
        acc = ttk.LabelFrame(frm, text="Tai khoan (nhap hoac de trong neu dung env)")
        acc.pack(fill="x", padx=10, pady=4)

        self.user_vars: list[tk.StringVar] = []
        self.pass_vars: list[tk.StringVar] = []
        # Default users (accounts are lab identifiers, not sensitive)
        default_users = ["kythuat", "lanhn", "", ""]  # Pre-filled credentials
        default_passes = ["670226", "796803", "", ""]  # Pre-filled passwords
        default_labels = ["User 1", "User 2", "User 3", "User 4"]

        apad = {"padx": 6, "pady": 2}
        for i in range(4):
            user_var = tk.StringVar(value=default_users[i])
            pass_var = tk.StringVar(value=default_passes[i])
            self.user_vars.append(user_var)
            self.pass_vars.append(pass_var)

            ttk.Label(acc, text=f"{default_labels[i]}").grid(row=i, column=0, sticky="w", **apad)
            ttk.Entry(acc, textvariable=user_var, width=25).grid(
                row=i, column=1, sticky="ew", **apad
            )
            ttk.Label(acc, text=f"Pass {i+1}").grid(row=i, column=2, sticky="w", **apad)
            ttk.Entry(acc, textvariable=pass_var, width=25, show="*").grid(
                row=i, column=3, sticky="ew", **apad
            )

        for i in range(4):
            acc.columnconfigure(i, weight=1)

        # ── Options ───────────────────────────────────────────────────────────
        opt = ttk.LabelFrame(frm, text="Tuy chon")
        opt.pack(fill="x", padx=10, pady=8)

        self.retry_var = tk.IntVar(value=3)
        self.timeout_var = tk.IntVar(value=12000)
        self.wait_var = tk.DoubleVar(value=0.8)
        self.save_debug_var = tk.BooleanVar(value=True)

        ttk.Label(opt, text="Retry / ma").grid(row=0, column=0, sticky="w", **pad)
        ttk.Entry(opt, textvariable=self.retry_var, width=12).grid(
            row=0, column=1, sticky="w", **pad
        )
        ttk.Label(opt, text="Timeout (ms)").grid(row=0, column=2, sticky="w", **pad)
        ttk.Entry(opt, textvariable=self.timeout_var, width=12).grid(
            row=0, column=3, sticky="w", **pad
        )
        ttk.Label(opt, text="Delay sau tim (s)").grid(row=0, column=4, sticky="w", **pad)
        ttk.Entry(opt, textvariable=self.wait_var, width=12).grid(
            row=0, column=5, sticky="w", **pad
        )
        ttk.Checkbutton(
            opt, text="Luu debug khi loi", variable=self.save_debug_var
        ).grid(row=0, column=6, sticky="w", **pad)

        # ── Progress section ─────────────────────────────────────────────────
        prg = ttk.LabelFrame(frm, text="Tien do")
        prg.pack(fill="x", padx=10, pady=4)

        self.status_var = tk.StringVar(value="San sang")
        self.summary_var = tk.StringVar(value="Chua chay")

        ttk.Label(prg, textvariable=self.status_var).pack(anchor="w", padx=8, pady=2)
        ttk.Label(prg, textvariable=self.summary_var).pack(anchor="w", padx=8, pady=0)

        self.progress_total = ttk.Progressbar(prg, orient="horizontal", mode="determinate")
        self.progress_total.pack(fill="x", padx=8, pady=3)

        # 4 worker progress bars in a 2×2 grid
        wf = ttk.Frame(prg)
        wf.pack(fill="x", padx=8, pady=2)
        wf.columnconfigure(0, weight=1)
        wf.columnconfigure(1, weight=1)

        self.progress_bars: list[ttk.Progressbar] = []
        worker_labels = ["Worker 1", "Worker 2", "Worker 3", "Worker 4"]

        for i in range(4):
            row = i // 2
            col = i % 2
            lbl_y = 0 if i < 2 else 0
            pad_top = 0 if i < 2 else 4
            ttk.Label(wf, text=worker_labels[i], font=("TkDefaultFont", 8)).grid(
                row=row * 2, column=col, sticky="w", pady=(pad_top, 0)
            )
            bar = ttk.Progressbar(wf, orient="horizontal", mode="determinate")
            bar.grid(row=row * 2 + 1, column=col, sticky="ew", padx=(0, 6 if col == 0 else 0), pady=1)
            self.progress_bars.append(bar)

        # ── Buttons ───────────────────────────────────────────────────────────
        btns = ttk.Frame(frm)
        btns.pack(fill="x", padx=10, pady=6)
        self.start_btn = ttk.Button(
            btns,
            text="Bat dau cao + ghep file",
            command=self.start_run,
        )
        self.start_btn.pack(side="left", padx=6, ipadx=10, ipady=4)
        ttk.Button(
            btns,
            text="Chay ngay (file moi nhat)",
            command=self._run_now,
        ).pack(side="left", padx=6, ipadx=10, ipady=4)
        ttk.Button(btns, text="Thoat", command=self.root.destroy).pack(
            side="right", padx=6
        )

        # ── Log ──────────────────────────────────────────────────────────────
        logfrm = ttk.LabelFrame(frm, text="Log")
        logfrm.pack(fill="both", expand=True, padx=10, pady=(0, 8))

        self.log_text = tk.Text(logfrm, wrap="word")
        self.log_text.pack(side="left", fill="both", expand=True)
        scrollbar = ttk.Scrollbar(logfrm, command=self.log_text.yview)
        scrollbar.pack(side="right", fill="y")
        self.log_text.configure(yscrollcommand=scrollbar.set)

    # ── Credential loading ───────────────────────────────────────────────────

    def _load_credentials_from_env(self):
        """Pre-fill credentials from environment variables if available."""
        env_creds = load_credentials_from_env()
        for i, cred in env_creds.items():
            if i <= 4:
                self.user_vars[i - 1].set(cred["user"])

    # ── File dialogs ───────────────────────────────────────────────────────

    def choose_excel(self):
        path = filedialog.askopenfilename(
            title="Chon file Excel goc",
            initialdir=str(INPUT_DIR),
            filetypes=[
                ("Excel files", "*.xlsx *.xlsm *.xls"),
                ("All files", "*.*"),
            ],
        )
        if path:
            self._load_excel_path(path)

    def _load_excel_path(self, path: str):
        """Load file Excel: convert .xls→.xlsx nếu cần, rồi auto-detect sheet/col."""
        import pandas as pd
        import openpyxl
        import datetime as _dt

        # ── Bước 1: chuyển .xls → .xlsx (openpyxl không đọc .xls) ─────────────
        if path.lower().endswith(".xls") and not path.lower().endswith(".xlsx"):
            tmp = DATA_DIR / (Path(path).stem + ".xlsx")  # lưu tạm vào Data/
            converted = False

            # Thử A: xlrd (file .xls binary cũ — Excel 97-2003)
            if not converted:
                try:
                    import xlrd
                    import datetime as _dt
                    wb_in = xlrd.open_workbook(path)
                    wb_out = openpyxl.Workbook()
                    wb_out.remove(wb_out.active)
                    for sheet_idx in range(wb_in.nsheets):
                        ws_in = wb_in.sheet_by_index(sheet_idx)
                        ws_out = wb_out.create_sheet(title=ws_in.name)
                        col = 0
                        for row in range(ws_in.nrows):
                            for col in range(ws_in.ncols):
                                cell = ws_in.cell(row, col)
                                ctype, value = cell.ctype, cell.value
                                if ctype == xlrd.XL_CELL_DATE:
                                    try:
                                        dt_tuple = xlrd.xldate_as_tuple(value, wb_in.datemode)
                                        value = (_dt.time(*dt_tuple[3:]) if dt_tuple[0] == 0
                                                 else _dt.datetime(*dt_tuple))
                                    except Exception:
                                        pass
                                elif ctype == xlrd.XL_CELL_BOOLEAN:
                                    value = bool(value)
                                elif ctype in (xlrd.XL_CELL_ERROR, xlrd.XL_CELL_EMPTY):
                                    value = None
                                ws_out.cell(row=row + 1, column=col + 1, value=value)
                        if col > 0:
                            ws_out.column_dimensions[
                                openpyxl.utils.get_column_letter(col + 1)
                            ].width = 15
                    wb_out.save(tmp)
                    converted = True
                except Exception:
                    pass

            # Thử B: openpyxl trực tiếp (file thực ra là .xlsx đổi tên thành .xls)
            if not converted:
                try:
                    openpyxl.load_workbook(path).save(tmp)
                    converted = True
                except Exception:
                    pass

            # Thử C: pandas + openpyxl engine
            if not converted:
                try:
                    sheets = pd.read_excel(path, sheet_name=None, engine="openpyxl")
                    wb_out = openpyxl.Workbook()
                    wb_out.remove(wb_out.active)
                    for sname, df in sheets.items():
                        ws = wb_out.create_sheet(title=sname)
                        ws.append(list(df.columns))
                        for _, row_data in df.iterrows():
                            ws.append(list(row_data))
                    wb_out.save(tmp)
                    converted = True
                except Exception:
                    pass

            if not converted:
                messagebox.showerror(
                    "Không đọc được file",
                    f"Không thể mở file:\n{Path(path).name}\n\n"
                    "File có thể bị hỏng hoặc định dạng không được hỗ trợ.\n"
                    "Hãy mở file trong Excel và lưu lại dạng .xlsx rồi thử lại."
                )
                return

            self.log(f"[Convert] {Path(path).name} → {tmp.name}")
            path = str(tmp)

        # ── Bước 2: set path và auto-detect sheet / cột mã đơn ─────────────────
        self.excel_path_var.set(path)

        ORDER_PATTERNS = [
            "mã đh", "mã đơn", "mã dh", "mã đơn hàng",
            "ma_dh", "ma_don", "madon", "madonhang",
            "order", "order_id", "orderid", "ma_order",
            "order_no", "orderno", "don_hang",
            "so_dh", "sodh", "so_don", "sodon",
        ]

        try:
            # Luôn dùng openpyxl cho .xlsx để tránh xlrd can thiệp
            xl = pd.ExcelFile(path, engine="openpyxl")
            sheets = xl.sheet_names
            best_sheet, best_col, best_score = None, None, -1

            for sheet_name in sheets:
                df = xl.parse(sheet_name=sheet_name)
                for col in [str(c).strip() for c in df.columns if str(c).strip()]:
                    col_lc = col.lower()
                    for pat in ORDER_PATTERNS:
                        if pat in col_lc:
                            score = 100 if col_lc == pat else 50
                            if score > best_score:
                                best_score, best_sheet, best_col = score, sheet_name, col
                            break

            if best_sheet is None:
                best_sheet = sheets[0] if sheets else "Sheet"
                df = xl.parse(sheet_name=best_sheet)
                cols = [str(c).strip() for c in df.columns if str(c).strip()]
                best_col = next((c for c in cols if c), "ma_dh")

            self.sheet_var.set(best_sheet)
            self.col_var.set(best_col)

        except Exception as e:
            messagebox.showerror("Lỗi đọc file Excel",
                                 f"Đọc file thành công nhưng không đọc được dữ liệu:\n\n{e}")
            return

        p = Path(path)
        # Output: Data/{stem}_scraped.xlsx
        self.output_path_var.set(str(DATA_DIR / (p.stem + "_scraped.xlsx")))
        # Debug: Data/debug_scrape/
        self.debug_dir_var.set(str(DATA_DIR / "debug_scrape"))
        self.log(
            f"[Config] File Excel: {p.name}  →  Sheet: {self.sheet_var.get()} | "
            f"Cot: {self.col_var.get()}"
        )

    def choose_output(self):
        path = filedialog.asksaveasfilename(
            title="Chon file dau ra",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
        )
        if path:
            self.output_path_var.set(path)

    def choose_debug_dir(self):
        path = filedialog.askdirectory(title="Chon thu muc debug")
        if path:
            self.debug_dir_var.set(path)

    # ── Config save/load ───────────────────────────────────────────────────

    def _load_config(self):
        try:
            if CONFIG_FILE.exists():
                data = json.loads(CONFIG_FILE.read_text(encoding="utf-8"))
                self._last_run_file = data.get("last_run_file") or None
        except Exception as e:
            self.log(f"[Config] Khong doc duoc config: {e}")

    def _save_config(self):
        """Save current config to labo_config.json."""
        try:
            data = {}
            if self._last_run_file:
                data["last_run_file"] = self._last_run_file
            CONFIG_FILE.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
        except Exception as e:
            self.log(f"[Config] Khong luu duoc config: {e}")

    # ── Auto-launch ─────────────────────────────────────────────────────────

    def _find_newest_xls(self) -> Optional[str]:
        """Find the newest .xls/.xlsx file in INPUT_DIR, excluding processed files."""
        if not INPUT_DIR.is_dir():
            return None
        candidates = [
            f
            for f in INPUT_DIR.iterdir()
            if f.suffix.lower() in (".xls", ".xlsx")
            and not any(
                tag in f.stem
                for tag in ("_scraped", "_final", "_cleaned")
            )
        ]
        if not candidates:
            return None
        return str(max(candidates, key=lambda f: f.stat().st_mtime))

    def _auto_launch(self):
        """Automatically find and load the newest Excel file, then start."""
        watch_dir = str(INPUT_DIR)
        if not INPUT_DIR.is_dir():
            self.log(f"[Auto] Thu muc Excel khong ton tai: {watch_dir}")
            self._schedule_auto_watch()
            return

        newest = self._find_newest_xls()
        if not newest:
            self.log(f"[Auto] Khong tim thay file .xls/.xlsx trong: {watch_dir}")
            self._schedule_auto_watch()
            return

        self.log(f"[Auto] Tim thay file moi nhat: {Path(newest).name}")

        # So sanh voi file da chay lan truoc
        if newest != self._last_run_file:
            self._cancel_countdown()
            self.log(f"[Auto] Phat hien file moi — bat dau cao ngay!")
            try:
                self._load_excel_path(newest)
            except Exception as e:
                self.log(f"[Auto] Loi khi load file: {e}")
                self._schedule_auto_watch()
                return
            self.root.after(500, self._trigger_start)
        else:
            self.log(f"[Auto] File da duoc chay truoc do — kiem tra lai sau 10 phut.")
            self._schedule_auto_watch()

    # ── Countdown helpers ─────────────────────────────────────────────────────

    def _start_countdown(self, seconds: int = 600):
        """Start a visible countdown timer; cancels any previous one."""
        if self._countdown_after_id:
            self.root.after_cancel(self._countdown_after_id)
        self._countdown_seconds = seconds
        self._update_countdown()

    def _update_countdown(self):
        """Tick the countdown display by 1 second."""
        if self._countdown_seconds <= 0:
            self.countdown_var.set("")
            return
        mins, secs = divmod(self._countdown_seconds, 60)
        self.countdown_var.set(f"  ⏱  Tu dong kiem tra sau: {mins:02d}:{secs:02d}")
        self._countdown_seconds -= 1
        self._countdown_after_id = self.root.after(1000, self._update_countdown)

    def _cancel_countdown(self):
        """Stop and clear the countdown display."""
        if self._countdown_after_id:
            self.root.after_cancel(self._countdown_after_id)
            self._countdown_after_id = None
        self._countdown_seconds = 0
        self.countdown_var.set("")

    # ── Auto-watch ─────────────────────────────────────────────────────────────

    def _schedule_auto_watch(self):
        """Schedule next auto-watch check in 10 minutes (if not already scheduled)."""
        if self._auto_scheduled:
            return
        self._auto_scheduled = True
        self._start_countdown(600)  # bắt đầu đếm ngược 10 phút
        self.root.after(600_000, self._auto_watch_tick)  # 10 min

    def _auto_watch_tick(self):
        """Called every 10 minutes to check for new files."""
        self._auto_scheduled = False
        self._cancel_countdown()
        if self.running:
            # Dang chay, kiem tra lai sau 10 phut
            self._schedule_auto_watch()
            return
        self._auto_launch()

    def _trigger_start(self):
        """Trigger the start button programmatically."""
        if not self.running:
            self.log("[Auto] Tu dong bat dau cao...")
            self.start_run()

    def _run_now(self):
        """Run immediately on the newest file in INPUT_DIR, bypassing countdown."""
        self._cancel_countdown()
        newest = self._find_newest_xls()
        if not newest:
            self.log("[Run Now] Khong tim thay file nao trong thu muc Excel/")
            return
        self.log(f"[Run Now] Bat dau cao file: {Path(newest).name}")
        try:
            self._load_excel_path(newest)
        except Exception as e:
            self.log(f"[Run Now] Loi load file: {e}")
            return
        self.root.after(300, self.start_run)

    # ── Server lifecycle ────────────────────────────────────────────────────

    def _start_server(self):
        """Start node server.js in a background process."""
        server_script = BASE_DIR / "server.js"
        if not server_script.exists():
            self.log("[Server] Khong tim thay server.js — bo qua.")
            return
        try:
            self._server_process = subprocess.Popen(
                ["node", str(server_script)],
                stdout=subprocess.PIPE,
                stderr=subprocess.PIPE,
                cwd=str(BASE_DIR),
            )
            self.log("[Server] Da khoi dong node server.js")
        except Exception as e:
            self.log(f"[Server] Loi khoi dong: {e}")

    def _stop_server(self):
        """Stop the node server process."""
        if self._server_process:
            try:
                self._server_process.terminate()
                self._server_process.wait(timeout=5)
                self.log("[Server] Da tat node server.js")
            except Exception:
                try:
                    self._server_process.kill()
                    self.log("[Server] Da kill node server.js")
                except Exception:
                    self.log("[Server] Khong the tat process")
            self._server_process = None

    # ── Logging ─────────────────────────────────────────────────────────────

    def log(self, text: str):
        """Append text to the log widget (thread-safe: always via root.after)."""
        self.log_text.insert("end", text + "\n")
        self.log_text.see("end")
        self.root.update_idletasks()

    # ── Validation ─────────────────────────────────────────────────────────

    def _validate_inputs(self) -> Optional[dict]:
        """Validate all inputs. Returns config dict or None on error."""
        excel_path = self.excel_path_var.get().strip()
        output_path = self.output_path_var.get().strip()
        base_url = self.base_url_var.get().strip()
        sheet_name = self.sheet_var.get().strip()
        col_name = self.col_var.get().strip()

        if not base_url:
            messagebox.showerror("Thieu du lieu", "Ban chua nhap trang web.")
            return None
        if not excel_path:
            messagebox.showerror("Thieu du lieu", "Ban chua chon file Excel goc.")
            return None
        if not Path(excel_path).exists():
            messagebox.showerror("Loi file", "File Excel goc khong ton tai.")
            return None
        if not output_path:
            messagebox.showerror("Thieu du lieu", "Ban chua chon file dau ra.")
            return None
        if not sheet_name or not col_name:
            messagebox.showerror(
                "Thieu du lieu",
                "Ban chua nhap sheet hoac cot ma don.",
            )
            return None

        # Collect credentials from UI fields
        accounts: list[tuple[str, str]] = []
        for i in range(4):
            u = self.user_vars[i].get().strip()
            p = self.pass_vars[i].get()
            if u and p:
                accounts.append((u, p))

        if not accounts:
            messagebox.showerror(
                "Thieu du lieu",
                "It nhat phai co User 1 / Pass 1.",
            )
            return None

        return {
            "excel_path": excel_path,
            "output_path": output_path,
            "base_url": base_url,
            "sheet_name": sheet_name,
            "col_name": col_name,
            "accounts": accounts,
            "retry": self.retry_var.get(),
            "timeout": self.timeout_var.get(),
            "wait_after": self.wait_var.get(),
            "save_debug": self.save_debug_var.get(),
            "debug_dir": (
                self.debug_dir_var.get().strip()
                or str(Path(output_path).with_suffix("")) + "_debug"
            ),
        }

    # ── Start ───────────────────────────────────────────────────────────────

    def _split_orders(self, order_ids: list[str], num_workers: int) -> list[list[str]]:
        """Distribute order IDs as evenly as possible across workers."""
        groups: list[list[str]] = [[] for _ in range(num_workers)]
        for i, oid in enumerate(order_ids):
            groups[i % num_workers].append(oid)
        return groups

    def start_run(self):
        if self.running:
            messagebox.showinfo("Dang chay", "Tien trinh dang chay.")
            return

        cfg = self._validate_inputs()
        if not cfg:
            return

        try:
            order_ids = load_order_ids(
                cfg["excel_path"], cfg["sheet_name"], cfg["col_name"]
            )
        except Exception as e:
            messagebox.showerror("Loi doc Excel", str(e))
            return

        if not order_ids:
            messagebox.showerror(
                "Khong co du lieu",
                "Khong tim thay ma don hang nao.",
            )
            return

        num_workers = len(cfg["accounts"])
        groups = self._split_orders(order_ids, num_workers)

        # Reset state
        self.running = True
        self.total_orders = len(order_ids)
        self.done_orders = 0
        self.ok_orders = 0
        self.fail_orders = 0
        self.worker_progress = {f"Worker-{i+1}": 0 for i in range(num_workers)}
        self.final_results = []
        self.final_failed = []
        self.finished_workers = 0
        self._cfg = cfg

        # Reset progress bars
        self.progress_total["maximum"] = max(1, self.total_orders)
        self.progress_total["value"] = 0
        for i, bar in enumerate(self.progress_bars):
            bar["maximum"] = max(1, len(groups[i]) if i < num_workers else 1)
            bar["value"] = 0

        self.log_text.delete("1.0", "end")
        self.status_var.set("Dang chay...")
        summary_parts = " | ".join(
            f"Worker{i+1}: {len(groups[i])}" for i in range(num_workers)
        )
        self.summary_var.set(f"Tong ma don: {self.total_orders} | {summary_parts}")
        self.start_btn.configure(state="disabled")

        threading.Thread(
            target=self._run_job,
            args=(cfg, groups),
            daemon=True,
        ).start()

    # ── Worker job ─────────────────────────────────────────────────────────

    def _run_job(self, cfg: dict, groups: list[list[str]]):
        """Run all scraper workers and then merge results."""
        debug_dir = Path(cfg["debug_dir"])
        debug_dir.mkdir(parents=True, exist_ok=True)

        threads: list[threading.Thread] = []
        for i, (account, group) in enumerate(zip(cfg["accounts"], groups), start=1):
            username, password = account
            worker_name = f"Worker-{i}"
            scraper = LaboAsiaAPIClient(
                base_url=cfg["base_url"],
                username=username,
                password=password,
                selectors=DEFAULT_SELECTORS,
                page_timeout_ms=cfg["timeout"],
                max_retry_per_order=cfg["retry"],
                # wait_after_search_sec, save_debug, debug_dir không còn dùng
                # nhưng vẫn truyền vào để tương thích nếu cần
                wait_after_search_sec=cfg["wait_after"],
                save_debug=cfg["save_debug"],
                debug_dir=debug_dir / f"worker{i}",
            )
            t = threading.Thread(
                target=scraper.scrape_order_list,
                args=(group, self.event_q, worker_name),
                daemon=True,
            )
            threads.append(t)

        for t in threads:
            t.start()

        for t in threads:
            t.join()

        # Tất cả worker thread đã kết thúc và đã post worker_finished vào queue.
        # Đặt sentinel _start_merge VÀO CUỐI queue (FIFO).
        # Khi main thread xử lý đến sentinel này, mọi worker_finished trước đó
        # đã được xử lý xong → final_results đầy đủ → an toàn để merge.
        self.event_q.put(("_start_merge", cfg))

    # ── Event processing (always on main thread via root.after) ─────────────

    def _process_event_now(self, event: tuple):
        """Process a single event. Called ONLY from main thread."""
        etype = event[0]

        if etype == "log":
            self.log(event[1])

        elif etype == "order_done":
            _, worker_name, ma_dh, row_count, error = event
            self.done_orders += 1
            self.worker_progress[worker_name] += 1
            self.progress_total["value"] = self.done_orders

            bar_map = {
                "Worker-1": self.progress_bars[0],
                "Worker-2": self.progress_bars[1],
                "Worker-3": self.progress_bars[2],
                "Worker-4": self.progress_bars[3],
            }
            if worker_name in bar_map:
                bar_map[worker_name]["value"] = self.worker_progress[worker_name]

            if error:
                self.fail_orders += 1
                self.final_failed.append(ma_dh)
                self.log(f"[{worker_name}] FAIL {ma_dh}: {error}")
            else:
                self.ok_orders += 1
                self.log(f"[{worker_name}] OK {ma_dh}: {row_count} cong doan")

            self.summary_var.set(
                f"Da xu ly {self.done_orders}/{self.total_orders} | "
                f"OK: {self.ok_orders} | FAIL: {self.fail_orders}"
            )

        elif etype == "worker_finished":
            _, worker_name, results, failed, username = event
            self.finished_workers += 1
            self.final_results.extend(results)
            self.log(
                f"[{worker_name}] Hoan tat. User: {username} | "
                f"Dong cao: {len(results)} | Fail: {len(failed)}"
            )

        elif etype == "merge_start":
            self.status_var.set("Dang ghep du lieu vao file goc...")
            self.log("Bat dau ghep du lieu vao workbook goc...")

        elif etype == "all_done":
            _, output_path, row_count, failed_path = event
            self.running = False
            self.start_btn.configure(state="normal")

            # Ghi nhan file da chay & luu config de lan sau kiem tra
            if self._cfg and self._cfg.get("excel_path"):
                self._last_run_file = self._cfg["excel_path"]
                self._save_config()
                self.log(f"[Auto] Da ghi nhan: {Path(self._last_run_file).name}")

            self.status_var.set("Hoan tat — dang lam sach...")
            self.log(f"Da tao file: {output_path}")
            if failed_path:
                self.log(f"Danh sach ma loi: {failed_path}")
            self._run_auto_clean(output_path)

        elif etype == "_start_merge":
            # Sentinel: tất cả worker_finished đã được xử lý (FIFO queue).
            # Chạy merge trong background thread để không block GUI.
            _, cfg = event
            threading.Thread(
                target=self._do_merge,
                args=(cfg,),
                daemon=True,
            ).start()

        elif etype == "fatal_error":
            self.running = False
            self.start_btn.configure(state="normal")
            self.status_var.set("Loi")
            self.log("LOI: " + event[1])
            messagebox.showerror("Loi", event[1])
            if self._cfg and self._cfg.get("excel_path"):
                self._last_run_file = self._cfg["excel_path"]
                self._save_config()
            self._schedule_auto_watch()

        elif etype == "clean_done":
            _, clean_out, retcode, stdout, stderr = event
            if retcode == 0 and Path(clean_out).exists():
                size_kb = Path(clean_out).stat().st_size / 1024
                self.status_var.set("Hoan tat")
                self.log(
                    f"[Auto-clean] Hoan thanh → {Path(clean_out).name} "
                    f"({size_kb:.1f} KB)"
                )
                self._run_data_thang(clean_out)  # lưu vào Data tháng
                self._schedule_auto_watch()  # bat dau dem nguoc cho lan tiep theo
            else:
                self.status_var.set("Lam sach that bai")
                self.log(f"[Auto-clean] Loi (returncode={retcode})")
                if stderr:
                    for line in stderr.strip().split("\n")[:10]:
                        if line.strip():
                            self.log(f"  {line}")

    def process_events(self):
        """Pump the event queue — runs on main thread via root.after."""
        try:
            while True:
                event = self.event_q.get_nowait()
                self._process_event_now(event)
        except queue.Empty:
            pass
        self.root.after(300, self.process_events)

    # ── Merge (chạy trong background thread sau khi tất cả worker xong) ────────

    def _do_merge(self, cfg: dict) -> None:
        """Merge scrape results vào workbook gốc. Chạy trong background thread."""
        self.event_q.put(("merge_start", cfg["excel_path"], cfg["output_path"]))
        try:
            progress_df = build_progress_df(self.final_results)
            save_json(progress_df, str(Path(cfg["output_path"]).with_suffix(".json")))
            merge_back_to_workbook(
                cfg["excel_path"], cfg["output_path"], progress_df
            )
            failed_path = ""
            if self.final_failed:
                failed_path = str(
                    Path(cfg["output_path"]).with_name(
                        Path(cfg["output_path"]).stem + "_failed.txt"
                    )
                )
                Path(failed_path).write_text(
                    "\n".join(self.final_failed), encoding="utf-8"
                )
            self.event_q.put(
                ("all_done", cfg["output_path"], len(progress_df), failed_path)
            )
        except Exception as e:
            self.event_q.put(("fatal_error", str(e)))

    # ── Auto-clean pipeline ─────────────────────────────────────────────────

    def _run_auto_clean(self, scraped_xlsx: str) -> None:
        """Automatically run labo_cleaner.py after scraping completes."""
        cleaner_path = BASE_DIR / "labo_cleaner.py"
        if not cleaner_path.exists():
            self.log(
                "⚠ [Auto-clean] Khong tim thay labo_cleaner.py — "
                "bo qua buoc lam sach."
            )
            return

        # Output name: *_scraped.xlsx → *_final.xlsx  vào thư mục CLEAN_DIR
        stem = Path(scraped_xlsx).stem
        if stem.endswith("_scraped"):
            stem = stem[: -len("_scraped")]
        clean_out = str(CLEAN_DIR / (stem + "_final.xlsx"))

        self.status_var.set("Dang lam sach du lieu...")
        self.log(
            f"[Auto-clean] Bat dau lam sach: {Path(scraped_xlsx).name} → "
            f"{Path(clean_out).name}"
        )

        def _worker():
            try:
                env = os.environ.copy()
                env["PYTHONIOENCODING"] = "utf-8"
                result = subprocess.run(
                    [sys.executable, str(cleaner_path), scraped_xlsx, clean_out],
                    capture_output=True,
                    text=True,
                    encoding="utf-8",
                    errors="replace",
                    env=env,
                )
                self.event_q.put(
                    ("clean_done", clean_out, result.returncode,
                     result.stdout, result.stderr)
                )
            except Exception as exc:
                self.event_q.put(("clean_done", clean_out, -1, "", str(exc)))

        threading.Thread(target=_worker, daemon=True).start()

    # ── Data Tháng ────────────────────────────────────────────────────────────

    def _determine_month_name(self, clean_xlsx: str) -> Optional[str]:
        """
        Xác định tên tháng từ cột 'Nhận lúc' của dòng mới nhất trong
        sheet 'Đơn hàng' của file _final.xlsx.
        Quy tắc: ngày 26–31 → tháng tiếp theo; ngày 1–25 → tháng hiện tại.
        Trả về 'MM_YYYY' hoặc None nếu lỗi.
        """
        import datetime as _dt
        try:
            df = pd.read_excel(clean_xlsx, sheet_name="Đơn hàng", dtype=str)
        except Exception:
            return None

        nhan_col = None
        for c in df.columns:
            if "nhận lúc" in str(c).lower():
                nhan_col = c
                break
        if not nhan_col:
            return None

        newest: Optional[_dt.datetime] = None
        for val in df[nhan_col]:
            if not val or str(val).strip() in ("", "nan"):
                continue
            for fmt in ("%Y-%m-%d %H:%M:%S", "%d/%m/%Y %H:%M:%S",
                        "%Y-%m-%d", "%d/%m/%Y"):
                try:
                    parsed = _dt.datetime.strptime(str(val).strip(), fmt)
                    if newest is None or parsed > newest:
                        newest = parsed
                    break
                except ValueError:
                    continue

        if newest is None:
            return None

        # Quy tắc 26→25: ngày ≥ 26 → tháng tiếp theo
        if newest.day >= 26:
            # Chuyển sang tháng sau
            month = newest.month + 1
            year = newest.year
            if month > 12:
                month = 1
                year += 1
        else:
            month = newest.month
            year = newest.year

        return f"{month:02d}_{year}"

    def _merge_month_file(self, clean_xlsx: str, month_name: str) -> None:
        """
        Gộp dữ liệu từ clean_xlsx vào file tháng tương ứng trong DATA_THANG_DIR.
        Quy tắc ghi đè: chỉ ghi đè nếu nhan_luc mới hơn.
        """
        target_path = DATA_THANG_DIR / f"Thang_{month_name}.xlsx"

        if not target_path.exists():
            # File tháng chưa có → copy nguyên file clean
            shutil.copy2(clean_xlsx, str(target_path))
            self.log(f"[Data Thang] Tao moi: {target_path.name}")
            return

        # Đọc source & target workbook
        src_wb = load_workbook(clean_xlsx)
        tgt_wb = load_workbook(str(target_path))

        SHEET_ORDERS = "Đơn hàng"
        SHEET_STAGES = "Tiến độ công đoạn"
        SHEET_SUM    = "Tổng hợp"

        src_orders = src_wb[SHEET_ORDERS] if SHEET_ORDERS in src_wb.sheetnames else None
        tgt_orders = tgt_wb[SHEET_ORDERS] if SHEET_ORDERS in tgt_wb.sheetnames else None

        # Chỉ merge sheet Đơn hàng để so sánh & ghi đè
        new_count = 0
        update_count = 0
        existing_count = 0

        if src_orders and tgt_orders:
            import datetime as _dt

            def _parse_date(val) -> Optional[_dt.datetime]:
                if val is None:
                    return None
                for fmt in ("%Y-%m-%d %H:%M:%S", "%d/%m/%Y %H:%M:%S",
                            "%Y-%m-%d", "%d/%m/%Y"):
                    try:
                        return _dt.datetime.strptime(str(val).strip(), fmt)
                    except ValueError:
                        continue
                return None

            # Build lookup: ma_dh → row_index (1-based) trong target
            tgt_lookup: dict[str, int] = {}
            for row in tgt_orders.iter_rows(min_row=2):
                cell_val = row[0].value  # cột A = Mã ĐH
                if cell_val and str(cell_val).strip():
                    tgt_lookup[str(cell_val).strip()] = row[0].row

            # Duyệt từng dòng trong source
            for row in src_orders.iter_rows(min_row=2):
                ma_dh = str(row[0].value).strip() if row[0].value else ""
                if not ma_dh or ma_dh in ("Mã ĐH", "nan"):
                    continue

                src_date = _parse_date(row[1].value)  # cột B = Nhận lúc

                if ma_dh not in tgt_lookup:
                    # Đơn mới → append vào target
                    tgt_max_row = tgt_orders.max_row + 1
                    for col_idx, cell in enumerate(row, start=1):
                        tgt_orders.cell(row=tgt_max_row, column=col_idx).value = cell.value
                    tgt_lookup[ma_dh] = tgt_max_row
                    new_count += 1
                else:
                    tgt_row_idx = tgt_lookup[ma_dh]
                    tgt_date = _parse_date(tgt_orders.cell(tgt_row_idx, 2).value)
                    # Ghi đè nếu nhan_luc mới hơn
                    if src_date and (tgt_date is None or src_date > tgt_date):
                        for col_idx, cell in enumerate(row, start=1):
                            tgt_orders.cell(row=tgt_row_idx, column=col_idx).value = cell.value
                        update_count += 1
                    else:
                        existing_count += 1

        # Tổng hợp: xóa & tính lại
        if SHEET_SUM in tgt_wb.sheetnames:
            del tgt_wb[SHEET_SUM]

        # Lưu target workbook
        tgt_wb.save(str(target_path))

        self.log(
            f"[Data Thang] {target_path.name}: "
            f"+{new_count} mới, ~{update_count} cap nhat, "
            f"{existing_count} giu nguyen"
        )

    def _run_data_thang(self, clean_xlsx: str) -> None:
        """Tự động gộp dữ liệu vào file tháng tương ứng."""
        month_name = self._determine_month_name(clean_xlsx)
        if not month_name:
            self.log("[Data Thang] Khong xac dinh duoc thang — bo qua.")
            return

        self.log(f"[Data Thang] Xac dinh thang: Thang_{month_name}")
        try:
            self._merge_month_file(clean_xlsx, month_name)
        except Exception as e:
            self.log(f"[Data Thang] Loi: {e}")


# ═══════════════════════════════════════════════════════════════════════════════
# ENTRY POINT
# ═══════════════════════════════════════════════════════════════════════════════

def main():
    root = tk.Tk()
    style = ttk.Style()
    try:
        style.theme_use("clam")
    except Exception:
        pass
    app = App(root)

    def on_close():
        if app.running:
            if not messagebox.askyesno(
                "Dang chay",
                "Tien trinh dang chay. Ban co chan muon thoat?\n"
                "(Du lieu co the chua duoc luu)",
            ):
                return
            app.running = False

        # Save any partial results before exiting
        if app.final_results and app._cfg:
            try:
                cfg = app._cfg
                progress_df = build_progress_df(app.final_results)
                save_json(
                    progress_df,
                    str(Path(cfg["output_path"]).with_suffix(".json")),
                )
                merge_back_to_workbook(
                    cfg["excel_path"], cfg["output_path"], progress_df
                )
                if app.final_failed:
                    fp = str(
                        Path(cfg["output_path"]).with_name(
                            Path(cfg["output_path"]).stem + "_failed.txt"
                        )
                    )
                    Path(fp).write_text(
                        "\n".join(app.final_failed), encoding="utf-8"
                    )
                messagebox.showinfo(
                    "Da luu",
                    f"File da duoc luu truoc khi thoat:\n{cfg['output_path']}",
                )
            except Exception as e:
                messagebox.showerror("Loi luu file", str(e))

        # Stop node server — use kill() on Windows for reliability
        if app._server_process:
            try:
                app._server_process.terminate()
                app._server_process.wait(timeout=3)
            except Exception:
                try:
                    app._server_process.kill()
                except Exception:
                    pass
            app._server_process = None

        root.destroy()

    root.protocol("WM_DELETE_WINDOW", on_close)
    root.mainloop()


if __name__ == "__main__":
    main()
