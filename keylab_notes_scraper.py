"""
keylab_notes_scraper.py

Scrape "Ghi chu SX" tu Keylab2022 cho file Excel active.
V2:
- Chi chay khi Excel moi da duoc import vao labo_data.db.
- Loc todo qua DB, bo qua don da co ghi_chu_sx va don da co "In mau" trong phuc_hinh.
- Ghi thang vao DB theo batch, khong dung keylab_notes.json lam trung gian.

Yeu cau: Keylab dang mo, da click "Tim kiem" de load danh sach don.
"""

import io
import json
import os
import sqlite3
import sys
import time
import unicodedata
from datetime import datetime
from pathlib import Path

import win32con
import win32gui
from pywinauto import Desktop
import pywinauto.keyboard as kb

if sys.stdout.encoding != "utf-8":
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

BASE_DIR = Path(__file__).parent
EXCEL_DIR = BASE_DIR / "Excel"
DB_PATH = BASE_DIR / "labo_data.db"
ERROR_FILE = BASE_DIR / "scraper_errors.json"

BATCH_SIZE = 10
FILTER_WAIT_SEC = 0.5
DETAIL_WAIT_SEC = 0.8
CLOSE_WAIT_SEC = 0.3
CLEAR_WAIT_SEC = 0.3
WAIT_FOR_DB_SEC = 420  # 7 phút — > timeout 5 phút của run_scrape.py


def normalize_ascii(value) -> str:
    return (
        unicodedata.normalize("NFD", str(value or "").lower())
        .encode("ascii", "ignore")
        .decode("ascii")
    )


def has_in_mau_ham(value) -> bool:
    n = normalize_ascii(value)
    return "in mau ham" in n or ("in mau" in n and "ham" in n)


def has_in_mau(value) -> bool:
    return "in m" in normalize_ascii(value)


def get_conn() -> sqlite3.Connection:
    con = sqlite3.connect(str(DB_PATH), timeout=30)
    con.row_factory = sqlite3.Row
    con.execute("PRAGMA busy_timeout=30000")
    return con


def wait_for_db_import(excel_filename: str, max_sec: int = WAIT_FOR_DB_SEC) -> bool:
    """Poll import_log đến khi có entry cho file Excel. Dùng trong --new-file mode."""
    stem = Path(excel_filename).stem
    deadline = time.time() + max_sec
    print(f"[--new-file] Cho DB import ({excel_filename}, toi da {max_sec}s)...", flush=True)
    while time.time() < deadline:
        with get_conn() as con:
            row = con.execute(
                "SELECT id FROM import_log WHERE ten_file LIKE ? AND trang_thai='ok' LIMIT 1",
                (f"%{stem}%",),
            ).fetchone()
        if row:
            print(f"[--new-file] DB da xac nhan: {excel_filename}", flush=True)
            return True
        time.sleep(10)
    print(f"[--new-file] WARN: het gio cho DB ({max_sec}s) — luu truc tiep.", flush=True)
    return False


# Excel active ---------------------------------------------------------------

def find_active_excel() -> Path | None:
    if not EXCEL_DIR.is_dir():
        return None
    candidates = [
        p for p in EXCEL_DIR.iterdir()
        if p.suffix.lower() in (".xls", ".xlsx", ".xlsm")
        and not any(tag in p.stem for tag in ("_scraped", "_final", "_cleaned"))
    ]
    return max(candidates, key=lambda p: p.stat().st_mtime) if candidates else None


def pick_sheet_name(sheet_names):
    for name in sheet_names:
        n = normalize_ascii(name).replace(" ", "")
        if any(k in n for k in ("donhang", "order", "sheet")):
            return name
    return sheet_names[0]


def pick_ma_dh_col(headers) -> int:
    for i, header in enumerate(headers):
        n = normalize_ascii(header).replace(" ", "").replace("_", "")
        if "ma" in n and any(k in n for k in ("dh", "don", "order")):
            return i
    for i, header in enumerate(headers):
        if "ma" in normalize_ascii(header):
            return i
    return 0


def clean_ma_dh(value) -> str:
    text = str(value or "").strip()
    if text.endswith(".0"):
        text = text[:-2]
    return text


def get_active_ma_dh():
    f = find_active_excel()
    if not f:
        print("ERR: Khong co file Excel trong Excel/")
        return [], ""

    print(f"Excel: {f.name}")
    ids = []
    try:
        if f.suffix.lower() == ".xls":
            import xlrd

            wb = xlrd.open_workbook(str(f))
            ws = wb.sheet_by_name(pick_sheet_name(wb.sheet_names()))
            headers = [str(ws.cell_value(0, c)).strip() for c in range(ws.ncols)]
            col = pick_ma_dh_col(headers)
            for r in range(1, ws.nrows):
                value = clean_ma_dh(ws.cell_value(r, col))
                if value and normalize_ascii(value) not in ("tong", "ma dh"):
                    ids.append(value)
        else:
            import openpyxl

            wb = openpyxl.load_workbook(str(f), read_only=True, data_only=True)
            ws = wb[pick_sheet_name(wb.sheetnames)]
            rows = ws.iter_rows(values_only=True)
            headers = [str(c or "").strip() for c in next(rows)]
            col = pick_ma_dh_col(headers)
            for row in rows:
                value = clean_ma_dh(row[col] if col < len(row) else "")
                if value and normalize_ascii(value) not in ("tong", "ma dh", "none"):
                    ids.append(value)
    except Exception as exc:
        print(f"ERR doc Excel: {exc}")

    return list(dict.fromkeys(ids)), f.name


# DB gate/filter/save --------------------------------------------------------

def ensure_db_schema():
    if not DB_PATH.exists():
        print(f"ERR: Khong thay DB {DB_PATH}. Hay chay python db_manager.py import-all truoc.")
        sys.exit(1)
    with get_conn() as con:
        cols = {row["name"] for row in con.execute("PRAGMA table_info(don_hang)").fetchall()}
        if "ghi_chu_sx" not in cols:
            con.execute("ALTER TABLE don_hang ADD COLUMN ghi_chu_sx TEXT DEFAULT ''")
        if "routed_to" not in cols:
            con.execute("ALTER TABLE don_hang ADD COLUMN routed_to TEXT DEFAULT NULL")
        con.commit()


def check_excel_in_db(excel_filename) -> bool:
    stem = Path(excel_filename).stem
    with get_conn() as con:
        row = con.execute(
            """
            SELECT ngay_import, so_don_hang
            FROM import_log
            WHERE ten_file LIKE ? AND trang_thai = 'ok'
            ORDER BY id DESC
            LIMIT 1
            """,
            (f"%{stem}%",),
        ).fetchone()

    if not row:
        print(f"WARN: {excel_filename} chua co trong import_log.")
        print("Hay chay: python db_manager.py import-all")
        return False

    print(f"DB: {excel_filename} da import luc {row['ngay_import']} ({row['so_don_hang']} don)")
    return True


def get_todo_from_db(ma_dh_list):
    if not ma_dh_list:
        return []

    with get_conn() as con:
        placeholders = ",".join("?" for _ in ma_dh_list)
        rows = con.execute(
            f"""
            SELECT ma_dh, COALESCE(ghi_chu_sx, '') AS ghi_chu_sx, COALESCE(phuc_hinh, '') AS phuc_hinh
            FROM don_hang
            WHERE ma_dh IN ({placeholders})
            """,
            ma_dh_list,
        ).fetchall()

    by_ma = {row["ma_dh"]: row for row in rows}
    skip_done = set()
    skip_inmau = set()
    missing_db = set()

    for ma_dh in ma_dh_list:
        row = by_ma.get(ma_dh)
        if row is None:
            missing_db.add(ma_dh)
            continue
        if str(row["ghi_chu_sx"] or "").strip():
            skip_done.add(ma_dh)
        elif has_in_mau(row["phuc_hinh"]):
            skip_inmau.add(ma_dh)

    todo = [
        ma_dh for ma_dh in ma_dh_list
        if ma_dh not in skip_done and ma_dh not in skip_inmau
    ]

    print(
        f"Tong: {len(ma_dh_list)} | Da co ghi chu: {len(skip_done)} | "
        f"Co In Mau trong phuc_hinh: {len(skip_inmau)} | Chua co DB: {len(missing_db)} | "
        f"Can cao: {len(todo)}"
    )
    return todo


def save_to_db(results: dict[str, str]) -> int:
    if not results:
        return 0

    with get_conn() as con:
        for ma_dh, note in results.items():
            note = str(note or "").strip()
            if has_in_mau_ham(note):
                con.execute(
                    """
                    UPDATE don_hang
                    SET ghi_chu_sx = ?, routed_to = 'zirco', updated_at = datetime('now','localtime')
                    WHERE ma_dh = ? AND (ghi_chu_sx IS NULL OR ghi_chu_sx = '')
                    """,
                    (note, ma_dh),
                )
            else:
                con.execute(
                    """
                    UPDATE don_hang
                    SET ghi_chu_sx = ?, updated_at = datetime('now','localtime')
                    WHERE ma_dh = ? AND (ghi_chu_sx IS NULL OR ghi_chu_sx = '')
                    """,
                    (note, ma_dh),
                )
        con.commit()
    return len(results)


# Pywinauto helpers ----------------------------------------------------------

def find_keylab():
    for w in Desktop(backend="uia").windows():
        try:
            title = w.window_text().lower()
            # Khớp chính xác Keylab2022: "LAB ASIA - KEYLAB VERSION 2022 ..."
            # Loại trừ cửa sổ terminal/console có "keylab" trong tên tab
            if "keylab" in title and ("version" in title or "lab asia" in title):
                return w
        except Exception:
            pass
    return None


def focus(hwnd):
    win32gui.ShowWindow(hwnd, win32con.SW_RESTORE)
    try:
        win32gui.SetForegroundWindow(hwnd)
    except Exception:
        import ctypes

        ctypes.windll.user32.SetForegroundWindow(hwnd)
    time.sleep(0.2)


def cell_value(item):
    try:
        return item.iface_value.CurrentValue or ""
    except Exception:
        try:
            return item.legacy_properties().get("Value", "") or ""
        except Exception:
            return ""


def safe_descendants(ctrl, **kwargs):
    try:
        if hasattr(ctrl, "wrapper_object"):
            ctrl = ctrl.wrapper_object()
        return ctrl.descendants(**kwargs)
    except Exception:
        return []


def is_ghi_chu_sx_cell(item) -> bool:
    try:
        name = item.element_info.name or ""
        control_type = item.element_info.control_type or ""
    except Exception:
        return False
    n = normalize_ascii(name)
    is_note_col = "ghi chu sx" in n or "ghi chu san xuat" in n
    return is_note_col and ("row" in n or normalize_ascii(control_type) == "dataitem")


def collect_ghi_chu_sx_values(root):
    values = []
    found_cells = 0
    for item in safe_descendants(root):
        if not is_ghi_chu_sx_cell(item):
            continue
        found_cells += 1
        value = cell_value(item).strip()
        if value and value not in values:
            values.append(value)
    return values, found_cells


def describe_controls(root, limit=60):
    lines = []
    for item in safe_descendants(root)[:limit]:
        try:
            info = item.element_info
            ctype = info.control_type or ""
            auto_id = info.automation_id or ""
            name = (info.name or "").replace("\r", "\\r").replace("\n", "\\n")
            if auto_id or name or ctype in ("Table", "DataItem", "Edit", "Pane", "Custom"):
                lines.append(f"{ctype}|{auto_id}|{name[:80]}")
        except Exception:
            continue
    return " ; ".join(lines[:limit]) or "no_descendants"


def close_detail(spec):
    try:
        detail = spec.child_window(auto_id="FormTaoDonHang", control_type="Window")
        detail.child_window(auto_id="btnDongLai", control_type="Button").invoke()
        time.sleep(CLOSE_WAIT_SEC)
    except Exception:
        pass


def read_ghi_chu_sx(detail):
    errors = []
    try:
        panel9 = detail.child_window(auto_id="panelControl9", control_type="Pane")
        grid = panel9.child_window(auto_id="gridControlDonHang", control_type="Table")
        dp = grid.child_window(title="Data Panel", control_type="Custom")
        found_cells = 0
        values = []
        for row in dp.children():
            for item in row.children():
                if item.element_info.name.startswith("Ghi ch\u00fa SX row"):
                    found_cells += 1
                    value = cell_value(item).strip()
                    if value:
                        values.append(value)
        if values or found_cells:
            return values
    except Exception as exc:
        errors.append(str(exc))

    values, found_cells = collect_ghi_chu_sx_values(detail)
    if values or found_cells:
        return values

    summary = describe_controls(detail)
    err = "; ".join(errors) if errors else "Ghi chu SX cell not found"
    err = f"{err}; detail_controls={summary}"
    return [f"[ERR:{err}]"]


def wait_for_detail(spec, ma_dh, timeout=DETAIL_WAIT_SEC):
    deadline = time.time() + timeout
    last_err = ""
    while time.time() < deadline:
        try:
            detail = spec.child_window(auto_id="FormTaoDonHang", control_type="Window")
            actual = detail.child_window(auto_id="textEditMaDonHangUser").window_text().strip()
            if actual == ma_dh:
                return detail, ""
            if actual:
                last_err = f"wrong_order:{actual}"
        except Exception as exc:
            last_err = str(exc)
        time.sleep(0.08)
    return None, last_err or "detail_timeout"


def apply_filter(main_grid, ma_dh):
    fr = main_grid.child_window(title="Filter Row", control_type="Custom")
    fi = fr.child_window(title="M\u00e3 \u0110H filter row", control_type="DataItem")
    fi.click_input()
    time.sleep(0.12)
    kb.send_keys("^a{DELETE}", pause=0.03)
    kb.send_keys(ma_dh, pause=0.02)
    kb.send_keys("{ENTER}", pause=0.05)
    time.sleep(FILTER_WAIT_SEC)


def find_visible_row(main_grid, ma_dh):
    dp = main_grid.child_window(title="Data Panel", control_type="Custom")
    for row in dp.children():
        for item in row.children():
            if "M\u00e3 \u0110H row" in item.element_info.name:
                if cell_value(item).strip() == ma_dh:
                    return row
    return None


def scrape_one(spec, main_grid, ma_dh):
    """Filter grid theo ma_dh, double-click, doc Ghi chu SX. Returns (note, status)."""
    try:
        apply_filter(main_grid, ma_dh)
    except Exception as exc:
        return None, f"filter_err:{exc}"

    try:
        target = find_visible_row(main_grid, ma_dh)
    except Exception as exc:
        return None, f"no_dp:{exc}"

    if target is None:
        return None, "not_visible"

    try:
        target.double_click_input()
    except Exception as exc:
        return None, f"dblclick_err:{exc}"

    detail, err = wait_for_detail(spec, ma_dh)
    if not detail:
        close_detail(spec)
        return None, f"read_err:{err}"

    try:
        values = read_ghi_chu_sx(detail)
        if any(str(value).startswith("[ERR:") for value in values):
            close_detail(spec)
            return None, values[0]
        note = " | ".join(values)
        close_detail(spec)
        return note, "ok"
    except Exception as exc:
        close_detail(spec)
        return None, f"read_err:{exc}"


def clear_filter(main_grid):
    try:
        fr = main_grid.child_window(title="Filter Row", control_type="Custom")
        fi = fr.child_window(title="M\u00e3 \u0110H filter row", control_type="DataItem")
        fi.click_input()
        time.sleep(0.1)
        kb.send_keys("^a{DELETE}{ENTER}", pause=0.03)
        time.sleep(CLEAR_WAIT_SEC)
    except Exception:
        pass


# Main -----------------------------------------------------------------------

def main():
    dry_run = "--dry-run" in sys.argv
    new_file_mode = "--new-file" in sys.argv

    ensure_db_schema()

    ma_dh_list, excel_name = get_active_ma_dh()
    if not ma_dh_list:
        sys.exit(1)

    if new_file_mode:
        print("[--new-file] Mode song song: cao Keylab ngay, luu sau khi DB san sang.", flush=True)
        # Chỉ bỏ qua gate check_excel_in_db; vẫn dùng get_todo_from_db để lọc
        # đơn đã có ghi chú hoặc có "In mau" — đơn mới chưa có trong DB sẽ vào
        # missing_db và được đưa vào todo bình thường.
    else:
        if not check_excel_in_db(excel_name):
            sys.exit(1)

    todo = get_todo_from_db(ma_dh_list)
    if dry_run:
        print("Dry-run: chi kiem tra DB/filter, khong thao tac Keylab.")
        return

    if not todo:
        print("Tat ca da co du lieu. Khong can chay.")
        return

    win = find_keylab()
    if not win:
        print("ERR: Keylab khong mo.")
        sys.exit(1)

    focus(win.handle)
    spec = Desktop(backend="uia").window(handle=win.handle)
    close_detail(spec)
    main_grid = spec.child_window(auto_id="gridControlDonHang", control_type="Table")

    pending_results = {}
    errors = []
    ok_count = 0
    t0 = time.time()

    for idx, ma_dh in enumerate(todo, 1):
        print(f"[{idx:>3}/{len(todo)}] {ma_dh} ...", end=" ", flush=True)
        note, status = scrape_one(spec, main_grid, ma_dh)
        if note is not None:
            pending_results[ma_dh] = note
            ok_count += 1
            print(f"OK  {repr(note[:70])}")
        else:
            errors.append({"ma_dh": ma_dh, "err": status})
            print(f"SKIP ({status})")

        # --new-file: gom tất cả vào memory trước, đợi DB rồi mới lưu một lần
        if not new_file_mode and len(pending_results) >= BATCH_SIZE:
            saved = save_to_db(pending_results)
            print(f"Saved DB batch: {saved}")
            pending_results.clear()

    # --new-file: đợi run_scrape.py import xong DB trước khi lưu
    if new_file_mode and pending_results:
        wait_for_db_import(excel_name)

    if pending_results:
        saved = save_to_db(pending_results)
        print(f"Saved DB batch: {saved}")

    clear_filter(main_grid)
    win32gui.ShowWindow(win.handle, win32con.SW_MINIMIZE)

    if errors:
        payload = {
            "scraped_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "excel_active": excel_name,
            "ok_count": ok_count,
            "skip_count": len(errors),
            "todo_count": len(todo),
            "errors": errors,
        }
        ERROR_FILE.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
        print(f"Errors saved: {ERROR_FILE}")
    elif ERROR_FILE.exists():
        try:
            old_payload = json.loads(ERROR_FILE.read_text(encoding="utf-8"))
            if old_payload.get("excel_active") == excel_name:
                ERROR_FILE.unlink()
                print(f"Cleared old errors: {ERROR_FILE}")
        except Exception:
            pass

    elapsed = time.time() - t0
    per_order = elapsed / len(todo) if todo else 0
    print(f"\nOK: {ok_count} | Skip: {len(errors)} | {elapsed:.0f}s | {per_order:.1f}s/don")
    sys.stdout.flush()
    # Force-exit để tránh comtypes/pywinauto COM threads chặn process thoát
    os._exit(2 if (errors and ok_count == 0) else 0)


if __name__ == "__main__":
    main()
