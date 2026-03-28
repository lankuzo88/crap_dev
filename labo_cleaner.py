"""
labo_cleaner.py
───────────────
Đọc file Excel gốc xuất từ DentalLab (.xls hoặc .xlsx),
làm sạch dữ liệu, xuất ra file Excel chuẩn.

Cách dùng:
    python labo_cleaner.py <input.xls> [output.xlsx]

Yêu cầu:
    pip install pandas openpyxl xlrd
"""

import sys
import re
import os
import math
from collections import Counter, defaultdict

# ── Dependencies ──────────────────────────────────────────────────────────────

try:
    import pandas as pd
except ImportError:
    sys.exit("LOI: Thieu thu vien: pip install pandas")

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("LOI: Thieu thu vien: pip install openpyxl")

# ═══════════════════════════════════════════════════════════════════════════════
# CONSTANTS
# ═══════════════════════════════════════════════════════════════════════════════

ACCOUNTS = ["hyct", "kythuat", "sonnt", "lanhn"]
LOAI_LENH = ["Làm mới", "Làm lại", "Bảo hành", "Làm tiếp", "Làm thêm", "Sửa"]
CONG_DOAN = ["CBM", "SÁP/Cadcam", "SƯỜN", "ĐẮP", "MÀI"]

# ── Sheet 1 headers & widths ─────────────────────────────────────────────────
S1_HEADERS = [
    "Mã ĐH", "Nhận lúc", "Y/c hoàn thành", "Y/c giao",
    "Khách hàng", "Bệnh nhân", "Phục hình", "SL",
    "Ghi chú ĐP", "Trạng thái",
]
S1_WIDTHS = [16, 16, 16, 16, 22, 18, 42, 6, 28, 14]
S1_CENTER_COLS = {1, 2, 3, 4, 8, 10}  # 1-indexed

# ── Sheet 2 headers & widths ─────────────────────────────────────────────────
S2_HEADERS = [
    "Mã ĐH", "TT", "Công đoạn", "KTV", "Xác nhận",
    "Thời gian HT", "Phục hình", "SL", "Loại lệnh", "Tài khoản",
]
S2_WIDTHS = [18, 5, 14, 16, 10, 20, 30, 6, 12, 12]
S2_CENTER_COLS = {1, 2, 6, 10}  # 1-indexed

# ── Sheet 3: Tổng hợp ─────────────────────────────────────────────────────────
S3_COL_WIDTHS = {
    "A": 38, "B": 14, "C": 22, "D": 22, "E": 22,
}
S3_SECTION_COLS = 5  # Công đoạn section

# ── Colors ───────────────────────────────────────────────────────────────────
C_HDR1 = "1F4E79"; C_HDR2 = "375623"; C_HDR3 = "833C00"
C_SUB1 = "2E75B6"; C_SUB2 = "5B8A3C"; C_SUB3 = "C55A11"
C_ALT1 = "F2F7FC"; C_ALT2 = "F4FAF0"; C_ALT3 = "FBE5D6"
C_YES  = "C6EFCE"; C_NO   = "FFEB9C"
C_SL   = "FFF2CC"; C_TOTAL = "FFE699"

LOAI_LENH_COLOR = {
    "Làm mới": None, "Sửa": "FFF2CC", "Làm lại": "FCE4D6",
    "Bảo hành": "EAF1FB", "Làm tiếp": "EDF7ED", "Làm thêm": "F3E8FF",
}

# ═══════════════════════════════════════════════════════════════════════════════
# STYLE HELPERS
# ═══════════════════════════════════════════════════════════════════════════════

def _make_border():
    side = Side(style="thin", color="CCCCCC")
    return Border(left=side, right=side, top=side, bottom=side)


def style_header(cell, bg_color: str, fg_color: str = "FFFFFF"):
    cell.font      = Font(name="Arial", bold=True, color=fg_color, size=10)
    cell.fill      = PatternFill("solid", start_color=bg_color)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border    = _make_border()


def style_data(
    cell,
    bg_color: str = None,
    bold: bool = False,
    center: bool = False,
    text_color: str = "000000",
    size: int = 9,
):
    cell.font      = Font(name="Arial", bold=bold, size=size, color=text_color)
    cell.alignment = Alignment(
        horizontal="center" if center else "left",
        vertical="center",
        wrap_text=True,
    )
    cell.border = _make_border()
    if bg_color:
        cell.fill = PatternFill("solid", start_color=bg_color)


# ═══════════════════════════════════════════════════════════════════════════════
# HELPERS
# ═══════════════════════════════════════════════════════════════════════════════

def safe_str(value) -> str:
    """Convert value to string safely, handling NaN floats."""
    if value is None:
        return ""
    if isinstance(value, float) and math.isnan(value):
        return ""
    return str(value).strip()


def extract_sl(text: str) -> int:
    """Extract total quantity (SL) from text like 'SL: 3'."""
    if not text:
        return 0
    return sum(int(m) for m in re.findall(r"SL:\s*(\d+)", str(text)))


def clean_noise(text: str) -> str:
    """Remove embedded account codes from a string."""
    if not text:
        return ""
    t = str(text)
    for account in ACCOUNTS:
        t = re.sub(re.escape(account), "", t)
    return re.sub(r"\s+", " ", t).strip()


def find_col(df, aliases: list) -> str | None:
    """Find a column name using a list of possible aliases (case-insensitive)."""
    for alias in aliases:
        if alias in df.columns:
            return alias
    # Fallback: case-insensitive search
    col_map = {c.lower(): c for c in df.columns}
    for alias in aliases:
        if alias.lower() in col_map:
            return col_map[alias.lower()]
    return None


def detect_engine(path: str) -> str:
    """Detect the appropriate Excel engine for a file."""
    if path.lower().endswith(".xls") and not path.lower().endswith(".xlsx"):
        try:
            import xlrd  # noqa: F401
            return "xlrd"
        except ImportError:
            sys.exit("LOI: File .xls can cai: pip install xlrd")
    return "openpyxl"


def extract_account(text: str) -> str:
    """Extract account code from text."""
    for account in ACCOUNTS:
        if account in str(text):
            return account
    return ""


def extract_loai_lenh(text: str) -> str:
    """Extract 'loại lệnh' keyword from text."""
    for loai in LOAI_LENH:
        if loai in str(text):
            return loai
    return ""


def safe_int(value, default=None):
    """Safely convert value to int."""
    try:
        return int(value)
    except (ValueError, TypeError):
        return default if default is not None else ""


# ═══════════════════════════════════════════════════════════════════════════════
# PARSER SHEET 1 — Đơn hàng
# ═══════════════════════════════════════════════════════════════════════════════

def parse_sheet1(df) -> list[dict]:
    """
    Parse Sheet 1 (Đơn hàng) from the raw DataFrame.
    Returns a list of row dicts with standardized keys.
    """
    aliases = {
        "ma_dh":      ["Mã ĐH", "ma_dh"],
        "nhan_luc":   ["Nhận lúc", "nhan_luc"],
        "yc_hoan":    ["Y/c hoàn thành", "yc_hoan"],
        "yc_giao":    ["Y/c giao", "yc_giao"],
        "khach_hang": ["Khách hàng", "khach_hang"],
        "benh_nhan":  ["Bệnh nhân", "benh_nhan"],
        "phuc_hinh":  ["Phục hình", "phuc_hinh"],
        "ghi_chu":    ["Ghi chú điều phối", "Ghi chú ĐP", "ghi_chu"],
        "trang_thai": ["Trạng thái", "trang_thai"],
    }
    cols = {key: find_col(df, aliases[key]) for key in aliases}

    rows = []
    for _, row in df.iterrows():
        ma = safe_str(row.get(cols["ma_dh"], "")) if cols["ma_dh"] else ""
        if not ma or ma in ("Mã ĐH", "ma_dh"):
            continue
        rows.append({
            key: safe_str(row.get(cols[key], "")) if cols[key] else ""
            for key in aliases
        })
    return rows


# ═══════════════════════════════════════════════════════════════════════════════
# PARSER SHEET 2 — Tiến độ công đoạn
# ═══════════════════════════════════════════════════════════════════════════════

def parse_sheet2(df) -> list[dict]:
    """
    Parse Sheet 2 (Tiến độ công đoạn) from the raw DataFrame.
    Returns a list of row dicts with standardized keys.
    """
    c_ma  = find_col(df, ["ma_dh", "Mã ĐH"])
    c_tt  = find_col(df, ["thu_tu", "TT"])
    c_cd  = find_col(df, ["cong_doan", "Công đoạn"])
    c_ktv = find_col(df, ["ten_ktv", "KTV"])
    c_tg  = find_col(df, ["thoi_gian_hoan_thanh", "Thời gian HT", "thoi_gian"])
    c_raw = find_col(df, ["raw_row_text"])
    c_acc = find_col(df, ["tai_khoan_cao", "Tài khoản"])
    c_ph  = find_col(df, ["phuc_hinh", "Phục hình"])
    c_ll  = find_col(df, ["loai_lenh", "Loại lệnh"])

    rows = []
    for _, row in df.iterrows():
        ma = safe_str(row.get(c_ma, "")) if c_ma else ""
        if not ma or ma in ("ma_dh", "Mã ĐH"):
            continue

        raw = safe_str(row.get(c_raw, "")) if c_raw else ""

        # Extract account: from raw text first, then from dedicated column
        acc = extract_account(raw)
        if not acc and c_acc:
            acc = safe_str(row.get(c_acc, ""))

        ktv = clean_noise(safe_str(row.get(c_ktv, ""))) if c_ktv else "-"
        tg  = clean_noise(safe_str(row.get(c_tg, ""))) if c_tg else "-"
        da_xn = "Có" if tg and tg not in ("-", "") else "Chưa"

        # Phục hình: try dedicated column first, then parse from raw
        phuc_hinh = ""
        if c_ph:
            phuc_hinh = clean_noise(safe_str(row.get(c_ph, "")))

        if not phuc_hinh and raw:
            loai_in_raw = extract_loai_lenh(raw)
            if loai_in_raw:
                loai_pos = raw.find(loai_in_raw)
                before = raw[:loai_pos]
                xm = re.search(
                    r"xác nhận\s*(?:\d{2}/\d{2}/\d{4}\s+\d{2}:\d{2}:\d{2})?\s*-?\s*",
                    before,
                )
                if xm:
                    # Extract text after the "xác nhận" marker
                    after_marker = before[xm.end():].strip()
                    # Remove any trailing loai_lenh keyword
                    candidate = after_marker.replace(loai_in_raw, "", 1).strip()
                    if candidate:
                        phuc_hinh = clean_noise(candidate)
                    else:
                        phuc_hinh = clean_noise(after_marker)
            else:
                # No loai_lenh found in raw — treat entire prefix as phuc_hinh
                phuc_hinh = clean_noise(before.strip()) if before else ""

        # Loại lệnh
        loai_lenh = extract_loai_lenh(raw)
        if not loai_lenh and c_ll:
            loai_lenh = safe_str(row.get(c_ll, ""))

        thu_tu = safe_int(row.get(c_tt, ""), "")

        rows.append({
            "ma_dh":          ma,
            "thu_tu":         thu_tu,
            "cong_doan":      safe_str(row.get(c_cd, "")) if c_cd else "",
            "ten_ktv":        ktv or "-",
            "da_xac_nhan":    da_xn,
            "thoi_gian":      tg or "-",
            "phuc_hinh":      phuc_hinh,
            "sl":             extract_sl(phuc_hinh),
            "loai_lenh":      loai_lenh,
            "tai_khoan_cao":  acc,
        })

    # Propagate tai_khoan_cao through all rows sharing the same ma_dh
    acc_map = {}
    for r in rows:
        if r["tai_khoan_cao"]:
            acc_map[r["ma_dh"]] = r["tai_khoan_cao"]
    for r in rows:
        if not r["tai_khoan_cao"] and r["ma_dh"] in acc_map:
            r["tai_khoan_cao"] = acc_map[r["ma_dh"]]

    return rows


# ═══════════════════════════════════════════════════════════════════════════════
# READ INPUT EXCEL
# ═══════════════════════════════════════════════════════════════════════════════

def read_input(path: str) -> tuple[list[dict], list[dict]]:
    """
    Read an Excel file and auto-detect + parse Sheet 1 (Đơn hàng)
    and Sheet 2 (Tiến độ công đoạn).
    Returns (s1_rows, s2_rows).
    """
    engine = detect_engine(path)
    print(f"   Engine  : {engine}")

    all_sheets = pd.read_excel(path, sheet_name=None, engine=engine, dtype=str)
    print(f"   Sheets  : {list(all_sheets.keys())}")

    s1_df = None
    s2_df = None

    for name, df in all_sheets.items():
        cols_lower = " ".join(str(c).lower() for c in df.columns)

        # Sheet 2 detection: raw_row_text present OR has progress-tracking columns
        is_s2 = (
            "raw_row_text" in cols_lower
            or ("thu_tu" in cols_lower and "cong_doan" in cols_lower)
            or ("tt" in cols_lower.split() and "công đoạn" in cols_lower)
            or ("xác nhận" in cols_lower and "công đoạn" in cols_lower)
        )

        # Sheet 1 detection: has order info but not Sheet 2
        is_s1 = (
            not is_s2
            and any(k in cols_lower for k in [
                "bệnh nhân", "benh_nhan", "khách hàng",
                "phục hình", "y/c giao",
            ])
        )

        if is_s2 and s2_df is None:
            s2_df = df
            print(f"   Sheet 2 : '{name}' ({len(df)} dòng)")
        elif is_s1 and s1_df is None:
            s1_df = df
            print(f"   Sheet 1 : '{name}' ({len(df)} dòng)")

    # Fallback: use sheet order
    if s1_df is None and s2_df is None:
        sheets = list(all_sheets.values())
        if len(sheets) >= 2:
            s1_df, s2_df = sheets[0], sheets[1]
            print("   [!] Dùng theo thứ tự sheet (không nhận diện được tên)")
        elif sheets:
            s2_df = sheets[0]

    return (
        parse_sheet1(s1_df) if s1_df is not None else [],
        parse_sheet2(s2_df) if s2_df is not None else [],
    )


# ═══════════════════════════════════════════════════════════════════════════════
# SUPPLEMENT SL FROM ORDERS INTO PROGRESS SHEET
# ═══════════════════════════════════════════════════════════════════════════════

def fill_sl_from_orders(s1: list[dict], s2: list[dict]) -> int:
    """
    Backfill phuc_hinh and SL into s2 from s1 when scraped data is incomplete.
    Only fills rows where phuc_hinh is empty or sl is 0.
    Returns the number of rows that were filled.
    """
    # Build lookup: ma_dh → (phuc_hinh, sl) from Sheet 1
    lookup: dict[str, dict] = {}
    for r in s1:
        key = r["ma_dh"].strip()
        if not key:
            continue
        ph = r["phuc_hinh"]
        lookup[key] = {"phuc_hinh": ph, "sl": extract_sl(ph)}

    filled = 0
    for r in s2:
        key = r["ma_dh"].strip()
        if key not in lookup:
            continue
        info = lookup[key]
        changed = False

        if not r["phuc_hinh"]:
            r["phuc_hinh"] = info["phuc_hinh"]
            changed = True
        if r["sl"] == 0 and info["sl"] > 0:
            r["sl"] = info["sl"]
            changed = True

        if changed:
            filled += 1

    return filled


# ═══════════════════════════════════════════════════════════════════════════════
# BUILD OUTPUT EXCEL
# ═══════════════════════════════════════════════════════════════════════════════

def build_excel(s1: list[dict], s2: list[dict], output_path: str):
    """
    Build a 3-sheet Excel workbook:
      - Sheet 1: Đơn hàng
      - Sheet 2: Tiến độ công đoạn
      - Sheet 3: Tổng hợp
    """
    wb = Workbook()

    # ── Sheet 1: Đơn hàng ────────────────────────────────────────────────────

    ws1 = wb.active
    ws1.title = "Đơn hàng"
    ws1.row_dimensions[1].height = 38
    ws1.freeze_panes = "A2"

    for col_idx, (header, width) in enumerate(zip(S1_HEADERS, S1_WIDTHS), start=1):
        style_header(ws1.cell(1, col_idx, header), C_HDR1)
        ws1.column_dimensions[get_column_letter(col_idx)].width = width

    for row_idx, row in enumerate(s1, start=2):
        sl = extract_sl(row["phuc_hinh"])
        bg = C_ALT1 if row_idx % 2 == 0 else None
        values = [
            row["ma_dh"], row["nhan_luc"], row["yc_hoan"], row["yc_giao"],
            row["khach_hang"], row["benh_nhan"], row["phuc_hinh"],
            sl, row["ghi_chu"], row["trang_thai"],
        ]
        for col_idx, value in enumerate(values, start=1):
            cell = ws1.cell(row_idx, col_idx, value)
            if col_idx == 8:  # SL column
                sl_color = "C00000" if sl >= 10 else ("7B3F00" if sl >= 5 else "000000")
                style_data(
                    cell,
                    bg_color=C_SL if sl >= 5 else bg,
                    center=True,
                    bold=sl >= 5,
                    text_color=sl_color,
                )
            else:
                style_data(cell, bg_color=bg, center=(col_idx in S1_CENTER_COLS))
        ws1.row_dimensions[row_idx].height = 20

    # Total row
    TOTAL_ROW_1 = len(s1) + 2
    total_sl_s1 = sum(extract_sl(row["phuc_hinh"]) for row in s1)
    ws1.cell(TOTAL_ROW_1, 7, "TỔNG").font = Font(name="Arial", bold=True, size=9)
    ws1.cell(TOTAL_ROW_1, 7).alignment = Alignment(horizontal="right")
    total_cell = ws1.cell(TOTAL_ROW_1, 8, total_sl_s1)
    total_cell.font = Font(name="Arial", bold=True, size=11, color="C00000")
    total_cell.fill = PatternFill("solid", start_color=C_TOTAL)
    total_cell.alignment = Alignment(horizontal="center")
    total_cell.border = _make_border()

    # ── Sheet 2: Tiến độ công đoạn ───────────────────────────────────────────

    ws2 = wb.create_sheet("Tiến độ công đoạn")
    ws2.row_dimensions[1].height = 38
    ws2.freeze_panes = "A2"

    for col_idx, (header, width) in enumerate(zip(S2_HEADERS, S2_WIDTHS), start=1):
        style_header(ws2.cell(1, col_idx, header), C_HDR2)
        ws2.column_dimensions[get_column_letter(col_idx)].width = width

    prev_ma = None
    shade = False
    for row_idx, row in enumerate(s2, start=2):
        if row["ma_dh"] != prev_ma:
            shade = not shade
            prev_ma = row["ma_dh"]
        bg = C_ALT2 if shade else None
        xn = row["da_xac_nhan"]
        sl = row["sl"]
        values = [
            row["ma_dh"], row["thu_tu"], row["cong_doan"], row["ten_ktv"],
            xn, row["thoi_gian"], row["phuc_hinh"],
            sl if sl > 0 else "", row["loai_lenh"], row["tai_khoan_cao"],
        ]
        for col_idx, value in enumerate(values, start=1):
            cell = ws2.cell(row_idx, col_idx, value)
            if col_idx == 5:  # Xác nhận column
                xn_color = "1F6B2E" if xn == "Có" else "7B5C00"
                style_data(
                    cell,
                    bg_color=C_YES if xn == "Có" else C_NO,
                    center=True,
                    bold=(xn == "Có"),
                    text_color=xn_color,
                )
            elif col_idx == 8:  # SL column
                style_data(
                    cell,
                    bg_color=C_SL if sl > 0 else bg,
                    center=True,
                    bold=True,
                    text_color="C00000" if sl >= 5 else "000000",
                )
            elif col_idx == 9:  # Loại lệnh column
                style_data(
                    cell,
                    bg_color=LOAI_LENH_COLOR.get(row["loai_lenh"]) or bg,
                    center=True,
                )
            else:
                style_data(cell, bg_color=bg, center=(col_idx in S2_CENTER_COLS))
        ws2.row_dimensions[row_idx].height = 18

    # ── Sheet 3: Tổng hợp ─────────────────────────────────────────────────────

    ws3 = wb.create_sheet("Tổng hợp")

    def merged_title(ws, row: int, text: str, bg_color: str, num_cols: int = 4):
        cell = ws.cell(row, 1, text)
        cell.font = Font(name="Arial", bold=True, size=12, color="FFFFFF")
        cell.fill = PatternFill("solid", start_color=bg_color)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.merge_cells(f"A{row}:{get_column_letter(num_cols)}{row}")
        ws.row_dimensions[row].height = 28

    # Set column widths
    for col_letter, width in S3_COL_WIDTHS.items():
        ws3.column_dimensions[col_letter].width = width

    # Section 1: Tiến độ theo công đoạn
    ROW_S1 = 1
    merged_title(ws3, ROW_S1, "Tiến độ theo công đoạn", C_HDR2, S3_SECTION_COLS)
    s1_hdrs = ["Công đoạn", "Đã XN", "Chưa XN", "% Hoàn thành", "Tổng SL đã XN"]
    for col_idx, hdr in enumerate(s1_hdrs, start=1):
        style_header(ws3.cell(ROW_S1 + 1, col_idx, hdr), C_SUB2)

    cd_xn = Counter()
    cd_total = Counter()
    cd_sl = Counter()
    for r in s2:
        cd_total[r["cong_doan"]] += 1
        if r["da_xac_nhan"] == "Có":
            cd_xn[r["cong_doan"]] += 1
            cd_sl[r["cong_doan"]] += r["sl"]

    for row_idx, cd in enumerate(CONG_DOAN, start=ROW_S1 + 2):
        total = cd_total.get(cd, 0)
        done = cd_xn.get(cd, 0)
        bg = C_ALT2 if row_idx % 2 == 0 else None
        pct = f"{done / total * 100:.0f}%" if total > 0 else "0%"
        values = [cd, done, total - done, pct, cd_sl.get(cd, 0)]
        for col_idx, val in enumerate(values, start=1):
            style_data(ws3.cell(row_idx, col_idx, val), bg_color=bg, center=(col_idx > 1))

    # Section 2: Khối lượng KTV
    ROW_S2 = ROW_S1 + 2 + len(CONG_DOAN) + 2
    merged_title(ws3, ROW_S2, "Khối lượng KTV (đã xác nhận)", C_HDR1, 4)
    s2_hdrs = ["KTV", "Số lần XN", "Tổng SL răng", "Công đoạn chính"]
    for col_idx, hdr in enumerate(s2_hdrs, start=1):
        style_header(ws3.cell(ROW_S2 + 1, col_idx, hdr), C_SUB1)

    ktv_cnt = Counter()
    ktv_sl = Counter()
    ktv_cd: dict[str, Counter] = {}
    for r in s2:
        if r["ten_ktv"] and r["ten_ktv"] != "-" and r["da_xac_nhan"] == "Có":
            ktv_cnt[r["ten_ktv"]] += 1
            ktv_sl[r["ten_ktv"]] += r["sl"]
            ktv_cd.setdefault(r["ten_ktv"], Counter())[r["cong_doan"]] += 1

    for row_idx, (ktv, cnt) in enumerate(ktv_cnt.most_common(), start=ROW_S2 + 2):
        main_cd = ktv_cd[ktv].most_common(1)[0][0] if ktv in ktv_cd else ""
        bg = C_ALT1 if row_idx % 2 == 0 else None
        values = [ktv, cnt, ktv_sl[ktv], main_cd]
        for col_idx, val in enumerate(values, start=1):
            style_data(ws3.cell(row_idx, col_idx, val), bg_color=bg, center=(col_idx > 1))

    # Section 3: Tổng SL theo loại phục hình
    ROW_S3 = ROW_S2 + 2 + len(ktv_cnt) + 2
    merged_title(ws3, ROW_S3, "Tổng SL theo loại phục hình", C_HDR3, 3)
    s3_hdrs = ["Loại phục hình", "Số răng"]
    for col_idx, hdr in enumerate(s3_hdrs, start=1):
        style_header(ws3.cell(ROW_S3 + 1, col_idx, hdr), C_SUB3)

    sl_by_type: defaultdict = defaultdict(int)
    for row in s1:
        for chunk in row["phuc_hinh"].split(";"):
            slm = re.search(r"SL:\s*(\d+)", chunk)
            if not slm:
                continue
            part = re.sub(r"\(.*?\)", "", chunk).strip()
            if part:
                sl_by_type[part] += int(slm.group(1))

    for row_idx, (pt, sl) in enumerate(
        sorted(sl_by_type.items(), key=lambda x: -x[1]), start=ROW_S3 + 2
    ):
        bg = C_ALT3 if row_idx % 2 == 0 else None
        style_data(ws3.cell(row_idx, 1, pt), bg_color=bg)
        style_data(ws3.cell(row_idx, 2, sl), bg_color=bg, center=True, bold=True)

    # Total row
    TOTAL_ROW_3 = ROW_S3 + 2 + max(len(sl_by_type), 0)
    total_sl = sum(sl_by_type.values())
    ws3.cell(TOTAL_ROW_3, 1, "TỔNG CỘNG").font = Font(name="Arial", bold=True, size=10)
    total_cell3 = ws3.cell(TOTAL_ROW_3, 2, total_sl)
    total_cell3.font = Font(name="Arial", bold=True, size=11, color="C00000")
    total_cell3.fill = PatternFill("solid", start_color=C_TOTAL)
    total_cell3.alignment = Alignment(horizontal="center")
    total_cell3.border = _make_border()

    wb.save(output_path)


# ═══════════════════════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════════════════════

def main():
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(1)

    inp = sys.argv[1]
    if not os.path.exists(inp):
        sys.exit(f"LOI: Khong tim thay file: {inp}")

    ext = os.path.splitext(inp)[1].lower()
    if ext not in (".xls", ".xlsx"):
        sys.exit(f"LOI: Chi ho tro .xls hoac .xlsx (nhan: {ext})")

    out = sys.argv[2] if len(sys.argv) >= 3 else os.path.splitext(inp)[0] + "_cleaned.xlsx"

    print(f"[DOC] Dang doc: {inp}")
    s1, s2 = read_input(inp)

    # Backfill SL from Sheet 1 into Sheet 2
    filled = fill_sl_from_orders(s1, s2)

    total_sl = sum(extract_sl(r["phuc_hinh"]) for r in s1)
    print(f"\n[OK] Da parse:")
    print(f"   Don hang  : {len(s1)} don")
    print(f"   Cong doan : {len(s2)} dong")
    print(f"   Tong SL   : {total_sl} rang")
    if filled:
        print(f"   Bo sung SL: {filled} dong (tu don hang)")

    print(f"\n[LUU] Dang xuat: {out}")
    build_excel(s1, s2, out)
    print(f"[DONE] Hoan thanh! File da luu tai: {out}")


if __name__ == "__main__":
    main()
