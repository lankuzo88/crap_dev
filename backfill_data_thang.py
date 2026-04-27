#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
backfill_data_thang.py
───────────────────────
Backfill: đưa tất cả file *_final.xlsx trong File_sach/ vào đúng file tháng
trong Data_thang/.

Cách dùng:
    python backfill_data_thang.py

File bỏ qua (EXCLUDE):
    - 2503a8_final.xlsx          → file tổng tháng
    - Bản sao 2403a3_final.xlsx → file trùng

Logic tương tự _merge_month_file() trong laboasia_gui_scraper_tkinter.py:
    - Đơn hàng: merge theo ma_dh, ghi đè nếu nhan_luc mới hơn
    - Tiến độ công đoạn: dedup theo (ma_dh, cong_doan), ghi đè nếu tg_ht mới hơn
    - Tổng hợp: xóa (labo_cleaner sẽ tạo lại khi cần)
"""

from __future__ import annotations

import datetime as _dt
import os
import shutil
from pathlib import Path

try:
    from openpyxl import load_workbook
    import pandas as pd
except ImportError:
    print("LOI: pip install openpyxl pandas")
    exit(1)


# ── Đường dẫn ────────────────────────────────────────────────────────────────
BASE_DIR          = Path(__file__).parent.resolve()
DATA_THANG_DIR    = BASE_DIR / "Data_thang"
CLEAN_DIR         = BASE_DIR / "File_sach"

# File bỏ qua
EXCLUDE = {
    "2503a8_final.xlsx",
    "Bản sao 2403a3_final.xlsx",
}

# Sheet names
SHEET_ORDERS = "Đơn hàng"
SHEET_STAGES = "Tiến độ công đoạn"
SHEET_SUM    = "Tổng hợp"


# ── Helpers ───────────────────────────────────────────────────────────────────

def _parse_date(val) -> _dt.datetime | None:
    """Parse various date string formats."""
    if val is None:
        return None
    if isinstance(val, _dt.datetime):
        return val
    if isinstance(val, _dt.date):
        return _dt.datetime.combine(val, _dt.time())
    for fmt in (
        "%Y-%m-%d %H:%M:%S", "%d/%m/%Y %H:%M:%S",
        "%Y-%m-%d",          "%d/%m/%Y",
    ):
        try:
            return _dt.datetime.strptime(str(val).strip(), fmt)
        except ValueError:
            continue
    return None


def determine_month_name(xlsx_path: str) -> str | None:
    """
    Đọc sheet 'Đơn hàng', tìm ngày lớn nhất trong cột 'Nhận lúc',
    trả về 'MM_YYYY' theo quy tắc 26→25.
    """
    try:
        df = pd.read_excel(xlsx_path, sheet_name=SHEET_ORDERS, dtype=str)
    except Exception:
        return None

    nhan_col = None
    for c in df.columns:
        if "nhận lúc" in str(c).lower():
            nhan_col = c
            break
    if not nhan_col:
        return None

    newest = None
    for val in df[nhan_col]:
        if not val or str(val).strip() in ("", "nan"):
            continue
        dt = _parse_date(val)
        if dt and (newest is None or dt > newest):
            newest = dt

    if newest is None:
        return None

    if newest.day >= 26:
        month = newest.month + 1
        year  = newest.year
        if month > 12:
            month, year = 1, year + 1
    else:
        month, year = newest.month, newest.year

    return f"{month:02d}_{year}"


# ── Merge logic ───────────────────────────────────────────────────────────────

def merge_month_file(src_path: str, month_name: str) -> dict:
    """
    Gộp một file _final.xlsx vào file tháng tương ứng.
    Trả về dict thống kê: {new_orders, updated_orders, new_stages, updated_stages, skipped}
    """
    target_path = DATA_THANG_DIR / f"Thang_{month_name}.xlsx"

    if not target_path.exists():
        shutil.copy2(src_path, str(target_path))
        return {"action": "created", "new_orders": 1, "updated_orders": 0,
                "new_stages": 0, "updated_stages": 0, "skipped": 0}

    src_wb = load_workbook(src_path)
    tgt_wb = load_workbook(str(target_path))

    src_orders = src_wb[SHEET_ORDERS] if SHEET_ORDERS in src_wb.sheetnames else None
    tgt_orders = tgt_wb[SHEET_ORDERS] if SHEET_ORDERS in tgt_wb.sheetnames else None
    src_stages = src_wb[SHEET_STAGES] if SHEET_STAGES in src_wb.sheetnames else None
    tgt_stages = tgt_wb[SHEET_STAGES] if SHEET_STAGES in tgt_wb.sheetnames else None

    new_ord, upd_ord, exist_ord = 0, 0, 0
    new_stg, upd_stg = 0, 0

    # ── Merge Đơn hàng ─────────────────────────────────────────────────────
    if src_orders and tgt_orders:
        tgt_lookup: dict[str, int] = {}
        for row in tgt_orders.iter_rows(min_row=2):
            val = row[0].value
            if val and str(val).strip():
                tgt_lookup[str(val).strip()] = row[0].row

        for row in src_orders.iter_rows(min_row=2):
            ma_dh = str(row[0].value).strip() if row[0].value else ""
            if not ma_dh or ma_dh in ("Mã ĐH", "nan"):
                continue

            src_date = _parse_date(row[1].value)

            if ma_dh not in tgt_lookup:
                tgt_max = tgt_orders.max_row + 1
                for ci, cell in enumerate(row, start=1):
                    tgt_orders.cell(row=tgt_max, column=ci).value = cell.value
                tgt_lookup[ma_dh] = tgt_max
                new_ord += 1
            else:
                tgt_row = tgt_lookup[ma_dh]
                tgt_date = _parse_date(tgt_orders.cell(tgt_row, 2).value)
                if src_date and (tgt_date is None or src_date > tgt_date):
                    for ci, cell in enumerate(row, start=1):
                        tgt_orders.cell(row=tgt_row, column=ci).value = cell.value
                    upd_ord += 1
                else:
                    exist_ord += 1

    # ── Merge Tiến độ công đoạn ──────────────────────────────────────────
    if src_stages and tgt_stages:
        tgt_stage_lookup: dict[tuple[str, str], tuple[int, _dt.datetime | None]] = {}
        for row in tgt_stages.iter_rows(min_row=2):
            ma_dh = str(row[0].value).strip() if row[0].value else ""
            cd    = str(row[2].value).strip() if len(row) > 2 and row[2].value else ""
            tg    = _parse_date(row[5].value) if len(row) > 5 else None
            if ma_dh and cd:
                tgt_stage_lookup[(ma_dh, cd)] = (row[0].row, tg)

        for row in src_stages.iter_rows(min_row=2):
            ma_dh = str(row[0].value).strip() if row[0].value else ""
            cd    = str(row[2].value).strip() if len(row) > 2 and row[2].value else ""
            if not ma_dh or not cd or ma_dh == "Mã ĐH":
                continue

            src_tg = _parse_date(row[5].value) if len(row) > 5 else None
            key    = (ma_dh, cd)

            if key not in tgt_stage_lookup:
                tgt_max = tgt_stages.max_row + 1
                vals    = [c.value for c in row]
                for ci, val in enumerate(vals, start=1):
                    tgt_stages.cell(row=tgt_max, column=ci).value = val
                tgt_stage_lookup[key] = (tgt_max, src_tg)
                new_stg += 1
            else:
                tgt_row, tgt_tg = tgt_stage_lookup[key]
                if src_tg and (tgt_tg is None or src_tg > tgt_tg):
                    vals = [c.value for c in row]
                    for ci, val in enumerate(vals, start=1):
                        tgt_stages.cell(row=tgt_row, column=ci).value = val
                    tgt_stage_lookup[key] = (tgt_row, src_tg)
                    upd_stg += 1

    # ── Xóa Tổng hợp ─────────────────────────────────────────────────────
    if SHEET_SUM in tgt_wb.sheetnames:
        del tgt_wb[SHEET_SUM]

    tgt_wb.save(str(target_path))
    return {
        "action": "merged",
        "new_orders": new_ord, "updated_orders": upd_ord,
        "new_stages": new_stg, "updated_stages": upd_stg,
        "skipped": exist_ord,
    }


# ── Main ─────────────────────────────────────────────────────────────────────

def run():
    print("=" * 55)
    print("  ASIA LAB — Backfill Data Thang")
    print("=" * 55)

    if not CLEAN_DIR.is_dir():
        print(f"! Thu muc File_sach/ khong ton tai: {CLEAN_DIR}")
        return

    DATA_THANG_DIR.mkdir(exist_ok=True)

    # Lấy tất cả *_final.xlsx, sorted theo tên (file name chứa date)
    files = sorted(
        f for f in CLEAN_DIR.iterdir()
        if f.suffix.lower() in (".xlsx", ".xls")
        and "_final" in f.stem
        and f.name not in EXCLUDE
    )

    if not files:
        print("! Khong tim thay file _final.xlsx nao trong File_sach/")
        return

    print(f"\n  File_sach/: {CLEAN_DIR}")
    print(f"  Data_thang/: {DATA_THANG_DIR}")
    print(f"  Tim thay {len(files)} file (da loai {len(EXCLUDE)} file)")
    print(f"\n  Bo qua: {EXCLUDE}")
    print()

    total_new_o, total_upd_o, total_new_s, total_upd_s, total_skip = 0, 0, 0, 0, 0

    for fp in files:
        month_name = determine_month_name(str(fp))
        if not month_name:
            print(f"  [!] Bo qua (khong xac dinh thang): {fp.name}")
            continue

        result = merge_month_file(str(fp), month_name)

        if result["action"] == "created":
            print(f"  [NEW] {fp.name} -> Thang_{month_name}.xlsx (tao moi)")
        else:
            no = result["new_orders"]
            uo = result["updated_orders"]
            ns = result["new_stages"]
            us = result["updated_stages"]
            sk = result["skipped"]
            print(
                f"  [OK] {fp.name}\n"
                f"       Don: +{no} moi, ~{uo} cap nhat, {sk} giu\n"
                f"       Tien do: +{ns} moi, ~{us} cap nhat"
            )
            total_new_o += no
            total_upd_o += uo
            total_new_s += ns
            total_upd_s += us
            total_skip  += sk

    print()
    print(f"  === TONG KET ===")
    print(f"  Don: +{total_new_o} moi, ~{total_upd_o} cap nhat, {total_skip} giu")
    print(f"  Tien do: +{total_new_s} moi, ~{total_upd_s} cap nhat")

    # ── Verify: kiểm tra tổng orders trong file tháng ──────────────────────
    print()
    print("  === VERIFY ===")
    for thang_fp in sorted(DATA_THANG_DIR.glob("Thang_*.xlsx")):
        wb = load_workbook(str(thang_fp), data_only=True, read_only=True)
        orders_ws = wb[SHEET_ORDERS] if SHEET_ORDERS in wb.sheetnames else None
        stages_ws = wb[SHEET_STAGES] if SHEET_STAGES in wb.sheetnames else None
        n_orders  = (orders_ws.max_row - 1) if orders_ws else 0
        n_stages  = (stages_ws.max_row - 1) if stages_ws else 0
        wb.close()
        print(f"  {thang_fp.name}: {n_orders} don, {n_stages} tien do")

    print()
    print("  === XONG ===")


if __name__ == "__main__":
    run()