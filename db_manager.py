"""
ASIA LAB — SQLite data manager
DB file: labo_data.db (same directory)

Schema:
  don_hang   — master order table (one row per ma_dh)
  tien_do    — workflow stages (5 rows per ma_dh)
  import_log — audit trail of imported files

CLI:
  python db_manager.py import-json  <file.json>
  python db_manager.py import-excel <file_final.xlsx>
  python db_manager.py import-all          # toàn bộ lịch sử
  python db_manager.py stats
"""

import sys, os, re, json, sqlite3, zipfile, unicodedata
from pathlib import Path
from datetime import datetime, date

BASE_DIR   = Path(__file__).parent
DB_PATH    = BASE_DIR / 'labo_data.db'
EXCEL_DIR  = BASE_DIR / 'Excel'
DATA_DIR   = BASE_DIR / 'Data'
CLEAN_DIR  = BASE_DIR / 'File_sach'

STAGE_ORDER = ['CBM', 'SÁP/Cadcam', 'SƯỜN', 'ĐẮP', 'MÀI']

# ── Helpers ───────────────────────────────────────────────────────────────────

def parse_ma_dh(ma_dh: str):
    """(ma_dh_goc, so_phu) — so_phu là None nếu không có suffix -N."""
    m = re.match(r'^(.+?)-(\d+)$', str(ma_dh).strip())
    if m:
        return m.group(1), int(m.group(2))
    return str(ma_dh).strip(), None

def norm_date(val) -> str:
    if not val:
        return ''
    s = str(val).strip()
    if not s or s == '-':
        return ''
    # DD/MM/YYYY HH:MM:SS  →  giữ nguyên
    # datetime object
    if hasattr(val, 'strftime'):
        return val.strftime('%d/%m/%Y %H:%M:%S')
    return s

def normalize_ascii(value) -> str:
    return unicodedata.normalize('NFD', str(value or '').lower()).encode('ascii', 'ignore').decode('ascii')

def default_room_for(phuc_hinh) -> str:
    t = str(phuc_hinh or '').lower()
    n = normalize_ascii(phuc_hinh)

    is_zirconia = (
        any(kw in t for kw in ['zircornia', 'zirconia', 'ziconia', 'zir-', 'zolid', 'cercon', 'la va'])
        or 'argen' in n
    )
    is_metal = (
        'kim loai' in n or 'titanium' in t or 'titan' in t
        or 'chrome' in t or 'cobalt' in t or 'cr-co' in t or 'cr co' in t
    )

    if 'cui gia' in n and 'zirconia' in n:
        return 'both'
    if 'in mau' in n or 'mau ham' in n:
        return 'zirco'
    if 'rang tam' in n or 'pmma' in t or 'in resin' in n:
        return 'zirco'

    if 'veneer' in t:
        if is_zirconia:
            return 'zirco'
        if is_metal:
            return 'sap'
        return 'sap'
    if 'mat dan' in t or 'mat dan' in n:
        return 'sap'

    if is_zirconia:
        return 'zirco'
    if is_metal:
        return 'sap'
    return 'sap'

def log(msg: str):
    sys.stdout.buffer.write(f'[db_manager] {msg}\n'.encode('utf-8', errors='replace'))
    sys.stdout.buffer.flush()

# ── Kết nối & khởi tạo schema ─────────────────────────────────────────────────

def get_conn() -> sqlite3.Connection:
    conn = sqlite3.connect(DB_PATH)
    conn.execute('PRAGMA journal_mode=WAL')
    conn.execute('PRAGMA foreign_keys=ON')
    conn.row_factory = sqlite3.Row
    return conn

def init_db():
    conn = get_conn()
    conn.executescript("""
    CREATE TABLE IF NOT EXISTS don_hang (
        id            INTEGER PRIMARY KEY AUTOINCREMENT,
        ma_dh         TEXT    NOT NULL UNIQUE,
        ma_dh_goc     TEXT    NOT NULL,
        so_phu        INTEGER,
        la_don_phu    INTEGER DEFAULT 0,

        nhap_luc      TEXT    DEFAULT '',
        yc_hoan_thanh TEXT    DEFAULT '',
        yc_giao       TEXT    DEFAULT '',
        khach_hang    TEXT    DEFAULT '',
        benh_nhan     TEXT    DEFAULT '',
        phuc_hinh     TEXT    DEFAULT '',
        sl            INTEGER DEFAULT 0,
        loai_lenh     TEXT    DEFAULT '',
        ghi_chu       TEXT    DEFAULT '',
        trang_thai    TEXT    DEFAULT '',
        tai_khoan_cao TEXT    DEFAULT '',
        barcode_labo  TEXT    DEFAULT '',
        routed_to     TEXT    DEFAULT NULL,

        nguon_file    TEXT    DEFAULT '',
        created_at    TEXT    DEFAULT (datetime('now','localtime')),
        updated_at    TEXT    DEFAULT (datetime('now','localtime'))
    );

    CREATE TABLE IF NOT EXISTS tien_do (
        id                    INTEGER PRIMARY KEY AUTOINCREMENT,
        ma_dh                 TEXT    NOT NULL,
        thu_tu                INTEGER NOT NULL,
        cong_doan             TEXT    NOT NULL,
        ten_ktv               TEXT    DEFAULT '',
        xac_nhan              TEXT    DEFAULT 'Chưa',
        thoi_gian_hoan_thanh  TEXT    DEFAULT '',
        raw_row_text          TEXT    DEFAULT '',
        nguon_file            TEXT    DEFAULT '',
        created_at            TEXT    DEFAULT (datetime('now','localtime')),
        updated_at            TEXT    DEFAULT (datetime('now','localtime')),

        UNIQUE(ma_dh, thu_tu),
        FOREIGN KEY (ma_dh) REFERENCES don_hang(ma_dh) ON DELETE CASCADE
    );

    CREATE TABLE IF NOT EXISTS import_log (
        id            INTEGER PRIMARY KEY AUTOINCREMENT,
        ten_file      TEXT    NOT NULL,
        loai_file     TEXT,
        ngay_import   TEXT    DEFAULT (datetime('now','localtime')),
        so_don_hang   INTEGER DEFAULT 0,
        so_cong_doan  INTEGER DEFAULT 0,
        trang_thai    TEXT,
        chi_tiet      TEXT
    );

    CREATE INDEX IF NOT EXISTS idx_don_hang_goc  ON don_hang(ma_dh_goc);
    CREATE INDEX IF NOT EXISTS idx_don_hang_giao ON don_hang(yc_giao);
    CREATE INDEX IF NOT EXISTS idx_tien_do_ma    ON tien_do(ma_dh);
    CREATE INDEX IF NOT EXISTS idx_tien_do_cd    ON tien_do(cong_doan);
    CREATE INDEX IF NOT EXISTS idx_tien_do_ktv   ON tien_do(ten_ktv);
    """)
    cols = {r[1] for r in conn.execute("PRAGMA table_info(don_hang)").fetchall()}
    if "barcode_labo" not in cols:
        conn.execute("ALTER TABLE don_hang ADD COLUMN barcode_labo TEXT DEFAULT ''")
    if "routed_to" not in cols:
        conn.execute("ALTER TABLE don_hang ADD COLUMN routed_to TEXT DEFAULT NULL")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_don_hang_barcode_labo ON don_hang(barcode_labo)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_don_hang_routed_to ON don_hang(routed_to)")
    rows = conn.execute("SELECT ma_dh, phuc_hinh FROM don_hang WHERE routed_to IS NULL").fetchall()
    for row in rows:
        conn.execute("UPDATE don_hang SET routed_to=? WHERE ma_dh=?", (default_room_for(row["phuc_hinh"]), row["ma_dh"]))
    conn.commit()
    conn.close()

# ── Upsert helpers ────────────────────────────────────────────────────────────

def upsert_don_hang(conn: sqlite3.Connection, row: dict):
    ma_goc, so_phu = parse_ma_dh(row['ma_dh'])
    conn.execute("""
        INSERT INTO don_hang
            (ma_dh, ma_dh_goc, so_phu, la_don_phu,
             nhap_luc, yc_hoan_thanh, yc_giao,
             khach_hang, benh_nhan, phuc_hinh, sl,
             loai_lenh, ghi_chu, trang_thai, tai_khoan_cao, barcode_labo, routed_to, nguon_file, updated_at)
        VALUES
            (:ma_dh, :ma_dh_goc, :so_phu, :la_don_phu,
             :nhap_luc, :yc_hoan_thanh, :yc_giao,
             :khach_hang, :benh_nhan, :phuc_hinh, :sl,
             :loai_lenh, :ghi_chu, :trang_thai, :tai_khoan_cao, :barcode_labo, :routed_to, :nguon_file,
             datetime('now','localtime'))
        ON CONFLICT(ma_dh) DO UPDATE SET
            nhap_luc      = CASE WHEN excluded.nhap_luc      != '' THEN excluded.nhap_luc      ELSE nhap_luc      END,
            yc_hoan_thanh = CASE WHEN excluded.yc_hoan_thanh != '' THEN excluded.yc_hoan_thanh ELSE yc_hoan_thanh END,
            yc_giao       = CASE WHEN excluded.yc_giao       != '' THEN excluded.yc_giao       ELSE yc_giao       END,
            khach_hang    = CASE WHEN excluded.khach_hang    != '' THEN excluded.khach_hang    ELSE khach_hang    END,
            benh_nhan     = CASE WHEN excluded.benh_nhan     != '' THEN excluded.benh_nhan     ELSE benh_nhan     END,
            phuc_hinh     = CASE WHEN excluded.phuc_hinh     != '' THEN excluded.phuc_hinh     ELSE phuc_hinh     END,
            sl            = CASE WHEN excluded.sl            >  0  THEN excluded.sl            ELSE sl            END,
            loai_lenh     = CASE WHEN excluded.loai_lenh     != '' THEN excluded.loai_lenh     ELSE loai_lenh     END,
            ghi_chu       = CASE WHEN excluded.ghi_chu       != '' THEN excluded.ghi_chu       ELSE ghi_chu       END,
            trang_thai    = CASE WHEN excluded.trang_thai    != '' THEN excluded.trang_thai    ELSE trang_thai    END,
            tai_khoan_cao = CASE WHEN excluded.tai_khoan_cao != '' THEN excluded.tai_khoan_cao ELSE tai_khoan_cao END,
            barcode_labo  = CASE WHEN excluded.barcode_labo  != '' THEN excluded.barcode_labo  ELSE barcode_labo  END,
            nguon_file    = excluded.nguon_file,
            updated_at    = datetime('now','localtime')
    """, {
        **row,
        'barcode_labo': str(row.get('barcode_labo', '')).strip(),
        'routed_to': default_room_for(row.get('phuc_hinh', '')),
        'ma_dh_goc': ma_goc,
        'so_phu': so_phu,
        'la_don_phu': 1 if so_phu is not None else 0,
    })

def upsert_tien_do(conn: sqlite3.Connection, row: dict):
    conn.execute("""
        INSERT INTO tien_do
            (ma_dh, thu_tu, cong_doan, ten_ktv, xac_nhan,
             thoi_gian_hoan_thanh, raw_row_text, nguon_file, updated_at)
        VALUES
            (:ma_dh, :thu_tu, :cong_doan, :ten_ktv, :xac_nhan,
             :thoi_gian_hoan_thanh, :raw_row_text, :nguon_file,
             datetime('now','localtime'))
        ON CONFLICT(ma_dh, thu_tu) DO UPDATE SET
            ten_ktv              = CASE WHEN excluded.ten_ktv              != '' THEN excluded.ten_ktv              ELSE ten_ktv              END,
            xac_nhan             = excluded.xac_nhan,
            thoi_gian_hoan_thanh = CASE WHEN excluded.thoi_gian_hoan_thanh != '' THEN excluded.thoi_gian_hoan_thanh ELSE thoi_gian_hoan_thanh END,
            raw_row_text         = CASE WHEN excluded.raw_row_text         != '' THEN excluded.raw_row_text         ELSE raw_row_text         END,
            nguon_file           = excluded.nguon_file,
            updated_at           = datetime('now','localtime')
    """, row)

# ── Import từ JSON (_scraped.json) ────────────────────────────────────────────

def import_json(filepath: str) -> dict:
    p = Path(filepath)
    fname = p.name
    try:
        rows = json.loads(p.read_text(encoding='utf-8'))
        if not isinstance(rows, list):
            rows = rows.get('rows', rows.get('data', []))
    except Exception as e:
        return {'ok': False, 'error': str(e)}

    conn = get_conn()
    n_orders = n_stages = 0
    seen_orders = set()

    try:
        with conn:
            for row in rows:
                ma = str(row.get('ma_dh', '')).strip()
                if not ma:
                    continue

                # Upsert order header (từ raw_row_text)
                if ma not in seen_orders:
                    txt = str(row.get('raw_row_text', ''))
                    ph, sl, lk = _parse_raw_row(txt)
                    upsert_don_hang(conn, {
                        'ma_dh': ma,
                        'nhap_luc': '', 'yc_hoan_thanh': '', 'yc_giao': '',
                        'khach_hang': '', 'benh_nhan': '',
                        'phuc_hinh': ph, 'sl': sl,
                        'loai_lenh': str(row.get('loai_lenh', lk)),
                        'ghi_chu': '', 'trang_thai': '',
                        'tai_khoan_cao': str(row.get('tai_khoan_cao', row.get('tai_khoan', ''))),
                        'barcode_labo': str(row.get('barcode_labo', '')).strip(),
                        'routed_to': '',
                        'nguon_file': fname,
                    })
                    seen_orders.add(ma)
                    n_orders += 1

                # Upsert stage
                cd = str(row.get('cong_doan', '')).strip()
                thu_tu = STAGE_ORDER.index(cd) + 1 if cd in STAGE_ORDER else (row.get('thu_tu') or 0)
                upsert_tien_do(conn, {
                    'ma_dh': ma,
                    'thu_tu': thu_tu,
                    'cong_doan': cd,
                    'ten_ktv': str(row.get('ten_ktv', '')),
                    'xac_nhan': str(row.get('xac_nhan', 'Chưa')),
                    'thoi_gian_hoan_thanh': norm_date(row.get('thoi_gian_hoan_thanh', '')),
                    'raw_row_text': str(row.get('raw_row_text', '')),
                    'nguon_file': fname,
                })
                n_stages += 1

        _log_import(fname, 'json', n_orders, n_stages, 'ok')
        return {'ok': True, 'don_hang': n_orders, 'tien_do': n_stages}
    except Exception as e:
        _log_import(fname, 'json', 0, 0, 'error', str(e))
        return {'ok': False, 'error': str(e)}
    finally:
        conn.close()

def _parse_raw_row(txt: str):
    """Tách phuc_hinh, sl, loai_lenh từ raw_row_text."""
    sl_m = re.search(r'SL:(\d+)', txt)
    sl = int(sl_m.group(1)) if sl_m else 0
    sl_idx = txt.find(' SL:')
    ph = txt[:sl_idx].strip() if sl_idx > 0 else ''
    lk_m = re.search(r',\s*(\S[^,]*)$', txt)
    lk = lk_m.group(1).strip() if lk_m else ''
    return ph, sl, lk

# ── Import từ Excel (_final.xlsx) ─────────────────────────────────────────────

def import_excel_final(filepath: str) -> dict:
    try:
        import openpyxl
    except ImportError:
        return {'ok': False, 'error': 'openpyxl not installed'}

    p = Path(filepath)
    fname = p.name
    try:
        wb = openpyxl.load_workbook(str(p), read_only=True, data_only=True)
    except Exception as e:
        return {'ok': False, 'error': str(e)}

    def get_sheet(*keywords):
        for kw in keywords:
            for name in wb.sheetnames:
                if kw.lower() in name.lower():
                    return wb[name]
        return None

    conn = get_conn()
    n_orders = n_stages = 0

    try:
        with conn:
            # ── Sheet "Đơn hàng" → don_hang metadata ─────────────────
            sh1 = get_sheet('Đơn hàng', 'don hang', 'order')
            if sh1:
                rows1 = list(sh1.iter_rows(values_only=True))
                if rows1:
                    h = [str(c or '').strip() for c in rows1[0]]
                    def ci(kw): return next((i for i, x in enumerate(h) if kw.lower() in x.lower()), -1)
                    idx = {
                        'ma': ci('Mã ĐH'), 'nhan': ci('Nhận'),
                        'ht': ci('hoàn thành'), 'giao': ci('giao'),
                        'kh': ci('Khách'), 'bn': ci('ệnh nhân'),
                        'ph': ci('Phục hình'), 'sl': ci('SL'),
                        'gc': ci('Ghi chú'), 'tt': ci('Trạng thái'),
                    }
                    for row in rows1[1:]:
                        ma = str(row[idx['ma']] or '').strip() if idx['ma'] >= 0 else ''
                        if not ma or 'TỔNG' in ma or ma == 'Mã ĐH':
                            continue
                        def v(k): return str(row[idx[k]] or '').strip() if idx.get(k, -1) >= 0 else ''
                        sl_raw = row[idx['sl']] if idx['sl'] >= 0 else 0
                        sl = int(sl_raw) if str(sl_raw).isdigit() else 0
                        upsert_don_hang(conn, {
                            'ma_dh': ma,
                            'nhap_luc': norm_date(row[idx['nhan']] if idx['nhan'] >= 0 else ''),
                            'yc_hoan_thanh': norm_date(row[idx['ht']] if idx['ht'] >= 0 else ''),
                            'yc_giao': norm_date(row[idx['giao']] if idx['giao'] >= 0 else ''),
                            'khach_hang': v('kh'), 'benh_nhan': v('bn'),
                            'phuc_hinh': v('ph'), 'sl': sl,
                            'loai_lenh': '', 'ghi_chu': v('gc'),
                            'trang_thai': v('tt'), 'tai_khoan_cao': '',
                            'barcode_labo': '',
                            'routed_to': '',
                            'nguon_file': fname,
                        })
                        n_orders += 1

            # ── Sheet "Tiến độ" → tien_do ─────────────────────────────
            sh2 = get_sheet('Tiến độ', 'tien do', 'progress')
            if sh2:
                rows2 = list(sh2.iter_rows(values_only=True))
                if rows2:
                    h2 = [str(c or '').strip() for c in rows2[0]]
                    def ci2(kw): return next((i for i, x in enumerate(h2) if kw.lower() in x.lower()), -1)
                    idx2 = {
                        'ma': ci2('Mã ĐH'), 'tt': ci2('TT'),
                        'cd': ci2('Công đoạn'), 'ktv': ci2('KTV'),
                        'xn': ci2('Xác nhận'), 'tg': ci2('Thời gian'),
                        'lk': ci2('Loại lệnh'), 'tk': ci2('Tài khoản'),
                        'ph': ci2('Phục hình'), 'sl2': ci2('SL'),
                    }
                    for row in rows2[1:]:
                        ma = str(row[idx2['ma']] or '').strip() if idx2['ma'] >= 0 else ''
                        if not ma or ma == 'Mã ĐH':
                            continue
                        def v2(k): return str(row[idx2[k]] or '').strip() if idx2.get(k, -1) >= 0 else ''
                        cd = v2('cd')
                        tt_raw = row[idx2['tt']] if idx2['tt'] >= 0 else None
                        try:
                            thu_tu = int(tt_raw) if tt_raw else (STAGE_ORDER.index(cd) + 1 if cd in STAGE_ORDER else 0)
                        except (ValueError, TypeError):
                            thu_tu = STAGE_ORDER.index(cd) + 1 if cd in STAGE_ORDER else 0

                        # Cập nhật loai_lenh và tai_khoan vào don_hang nếu có
                        lk = v2('lk'); tk = v2('tk')
                        if lk or tk:
                            conn.execute("""
                                UPDATE don_hang SET
                                    loai_lenh     = CASE WHEN :lk != '' THEN :lk ELSE loai_lenh     END,
                                    tai_khoan_cao = CASE WHEN :tk != '' THEN :tk ELSE tai_khoan_cao END,
                                    updated_at    = datetime('now','localtime')
                                WHERE ma_dh = :ma
                            """, {'ma': ma, 'lk': lk, 'tk': tk})

                        # Đảm bảo don_hang tồn tại trước khi insert tien_do
                        conn.execute("""
                            INSERT OR IGNORE INTO don_hang
                                (ma_dh, ma_dh_goc, so_phu, la_don_phu, nguon_file)
                            VALUES (:ma, :goc, :sp, :ldp, :fn)
                        """, {
                            'ma': ma, 'fn': fname,
                            **({'goc': parse_ma_dh(ma)[0], 'sp': parse_ma_dh(ma)[1],
                                'ldp': 1 if parse_ma_dh(ma)[1] is not None else 0})
                        })

                        upsert_tien_do(conn, {
                            'ma_dh': ma, 'thu_tu': thu_tu, 'cong_doan': cd,
                            'ten_ktv': v2('ktv'),
                            'xac_nhan': 'Có' if v2('xn') == 'Có' else 'Chưa',
                            'thoi_gian_hoan_thanh': norm_date(row[idx2['tg']] if idx2['tg'] >= 0 else ''),
                            'raw_row_text': '', 'nguon_file': fname,
                        })
                        n_stages += 1

        _log_import(fname, 'excel', n_orders, n_stages, 'ok')
        return {'ok': True, 'don_hang': n_orders, 'tien_do': n_stages}
    except Exception as e:
        _log_import(fname, 'excel', 0, 0, 'error', str(e))
        return {'ok': False, 'error': str(e)}
    finally:
        conn.close()

def _log_import(fname, ftype, n_dh, n_td, status, detail=''):
    try:
        conn = get_conn()
        conn.execute(
            "INSERT INTO import_log (ten_file,loai_file,so_don_hang,so_cong_doan,trang_thai,chi_tiet) VALUES (?,?,?,?,?,?)",
            (fname, ftype, n_dh, n_td, status, detail)
        )
        conn.commit()
        conn.close()
    except Exception:
        pass

# ── Import toàn bộ lịch sử ────────────────────────────────────────────────────

def import_all_historical():
    """
    Thứ tự:
      1. _final.xlsx  (Excel sạch — có metadata đầy đủ)
      2. _scraped.json (override/bổ sung tiến độ realtime)
    Bỏ qua file đã import (dựa vào import_log).
    """
    init_db()

    conn = get_conn()
    already = set(r[0] for r in conn.execute("SELECT ten_file FROM import_log WHERE trang_thai='ok'"))
    conn.close()

    excel_files = sorted(CLEAN_DIR.glob('*_final.xlsx'), key=lambda f: f.stat().st_mtime)
    json_files  = sorted(DATA_DIR.glob('*_scraped.json'), key=lambda f: f.stat().st_mtime)

    total_ok = total_err = 0

    log(f'Tìm thấy {len(excel_files)} Excel final + {len(json_files)} JSON scraped')

    for f in excel_files:
        if f.name in already:
            continue
        r = import_excel_final(str(f))
        if r['ok']:
            log(f'  ✓ {f.name}: {r["don_hang"]} đơn, {r["tien_do"]} công đoạn')
            total_ok += 1
        else:
            log(f'  ✗ {f.name}: {r["error"]}')
            total_err += 1

    for f in json_files:
        if f.name in already:
            continue
        r = import_json(str(f))
        if r['ok']:
            log(f'  ✓ {f.name}: {r["don_hang"]} đơn, {r["tien_do"]} công đoạn')
            total_ok += 1
        else:
            log(f'  ✗ {f.name}: {r["error"]}')
            total_err += 1

    log(f'\nHoàn tất: {total_ok} OK, {total_err} lỗi')
    return total_ok, total_err

# ── Query cho dashboard ───────────────────────────────────────────────────────

def get_dashboard_data() -> dict:
    """
    Trả về dict tương thích với format hiện tại của server.js:
    { source: {db: 'labo_data.db'}, orders: [...] }
    """
    conn = get_conn()
    orders_raw = conn.execute("""
        SELECT d.*, GROUP_CONCAT(
            t.thu_tu||'|'||t.cong_doan||'|'||t.ten_ktv||'|'||t.xac_nhan||'|'||t.thoi_gian_hoan_thanh,
            ';;'
        ) AS stages_raw
        FROM don_hang d
        LEFT JOIN tien_do t ON t.ma_dh = d.ma_dh
        GROUP BY d.ma_dh
        ORDER BY d.yc_giao ASC, d.nhap_luc ASC
    """).fetchall()
    conn.close()

    orders = []
    SKIP_MAP = {
        'Sửa':    {0, 1, 2},   # skip CBM, SÁP, SƯỜN
        'TS':     {3, 4},       # skip ĐẮP, MÀI
        'Làm lại':{},
        'Làm mới':{},
    }

    for row in orders_raw:
        d = dict(row)
        stages_raw = d.pop('stages_raw') or ''
        lk = d.get('loai_lenh', '')
        skip = SKIP_MAP.get(lk, set())

        stages_map = {}
        for part in stages_raw.split(';;'):
            parts = part.split('|', 4)
            if len(parts) >= 5:
                try:
                    thu_tu = int(parts[0])
                    stages_map[thu_tu] = {
                        'n':  parts[1],
                        'k':  parts[2],
                        'x':  parts[3] == 'Có',
                        't':  parts[4],
                        'sk': (thu_tu - 1) in skip,
                    }
                except (ValueError, IndexError):
                    pass

        stages = []
        for i, name in enumerate(STAGE_ORDER):
            s = stages_map.get(i + 1, {'n': name, 'k': '', 'x': False, 't': '', 'sk': i in skip})
            s['n'] = name
            s['sk'] = i in skip
            stages.append(s)

        active = [s for s in stages if not s['sk']]
        done = sum(1 for s in active if s['x'])
        total = len(active)
        pct = round(done / total * 100) if total else 0

        cur_ktv = ''
        for s in reversed(stages):
            if not s['sk'] and s['k']:
                cur_ktv = s['k']
                break
        last_tg = ''
        for s in stages:
            if s['t']:
                last_tg = s['t']

        orders.append({
            'ma_dh':   d['ma_dh'],
            'nhan':    d['nhap_luc'],
            'yc_ht':   d['yc_hoan_thanh'],
            'yc_giao': d['yc_giao'],
            'kh':      d['khach_hang'],
            'bn':      d['benh_nhan'],
            'ph':      d['phuc_hinh'],
            'sl':      d['sl'],
            'gc':      d['ghi_chu'],
            'lk':      d['loai_lenh'],
            'tk':      d['tai_khoan_cao'],
            'stages':  stages,
            'done':    done,
            'total':   total,
            'pct':     pct,
            'curKtv':  cur_ktv,
            'lastTg':  last_tg,
        })

    return {'source': {'db': str(DB_PATH.name)}, 'orders': orders}

# ── CLI ───────────────────────────────────────────────────────────────────────

def cmd_stats():
    init_db()
    conn = get_conn()
    n_dh = conn.execute("SELECT COUNT(*) FROM don_hang").fetchone()[0]
    n_td = conn.execute("SELECT COUNT(*) FROM tien_do").fetchone()[0]
    n_phu = conn.execute("SELECT COUNT(*) FROM don_hang WHERE la_don_phu=1").fetchone()[0]
    n_log = conn.execute("SELECT COUNT(*) FROM import_log WHERE trang_thai='ok'").fetchone()[0]
    top_ktv = conn.execute("""
        SELECT ten_ktv, COUNT(*) as c FROM tien_do
        WHERE ten_ktv != '' AND xac_nhan='Có'
        GROUP BY ten_ktv ORDER BY c DESC LIMIT 5
    """).fetchall()
    conn.close()

    lines = [
        '',
        '=== ASIA LAB DB Stats ===',
        f'  Don hang    : {n_dh} (trong do {n_phu} don phu -N)',
        f'  Cong doan   : {n_td}',
        f'  Files import: {n_log}',
        f'  DB path     : {DB_PATH}',
        '',
        '  Top KTV (da xac nhan):',
    ] + [f'    {row[0]}: {row[1]} cong doan' for row in top_ktv]
    sys.stdout.buffer.write('\n'.join(lines).encode('utf-8', errors='replace') + b'\n')
    sys.stdout.buffer.flush()

if __name__ == '__main__':
    cmd = sys.argv[1] if len(sys.argv) > 1 else 'stats'

    if cmd == 'init':
        init_db()
        log('DB initialized.')

    elif cmd == 'import-json':
        if len(sys.argv) < 3:
            print('Usage: python db_manager.py import-json <file.json>')
            sys.exit(1)
        init_db()
        r = import_json(sys.argv[2])
        print(r)

    elif cmd == 'import-excel':
        if len(sys.argv) < 3:
            print('Usage: python db_manager.py import-excel <file_final.xlsx>')
            sys.exit(1)
        init_db()
        r = import_excel_final(sys.argv[2])
        print(r)

    elif cmd == 'import-all':
        import_all_historical()

    elif cmd == 'stats':
        cmd_stats()

    else:
        print(__doc__)
