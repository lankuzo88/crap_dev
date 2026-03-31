#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
ASIA LAB - AI Chat: Claude Sonnet 4 + NotebookLM-style RAG
- 3 nguon data: realtime (server.js) / learnedStats / knowledge
- Trich dan nguon moi cau tra loi
- Khong hallucinate
"""
import os, sys, json, io, re, urllib.request, urllib.error, uuid
from pathlib import Path
from datetime import datetime

try:
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
except Exception:
    pass


# ── Helpers ──────────────────────────────────────────────────────────────────
import unicodedata

def unaccent(s: str) -> str:
    """
    Chuyen 'Đơn Hàng' -> 'don hang' (ASCII thuan).
    unicodedata NFKD khong xu ly duoc 'đ'/'Đ' (U+0110/U+0111) tren Python 3.14.
    Can replace thu cong cho cac ky tu tieng Viet co diacritics.
    """
    if not s:
        return ""
    # 1. Decompose (xử lý dấu mũ, dấu ngã, dấu hỏi...)
    nfkd = unicodedata.normalize("NFKD", s)
    # 2. Strip combining marks
    ascii_only = "".join(c for c in nfkd if not unicodedata.combining(c))
    # 3. Replace Vietnamese base chars NFKD can't handle
    for old, new in [("Đ","D"),("đ","d"),("Ơ","O"),("ơ","o"),
                     ("Ư","U"),("ư","u"),("Ô","O"),("ô","o"),
                     ("Ê","E"),("ê","e"),("Ă","A"),("ă","a")]:
        ascii_only = ascii_only.replace(old, new)
    return ascii_only.lower()

BASE_DIR     = os.path.dirname(os.path.abspath(__file__))
INSTRUCTIONS = os.path.join(BASE_DIR, "ai_instructions.txt")
KNOWLEDGE_FP = os.path.join(BASE_DIR, "ai_knowledge.json")
MEMORY_FP   = os.path.join(BASE_DIR, "ai_memory.json")
API_KEY     = "sk-7df540121d72d9bbe64730c4c96f4db488492620644269c88ca817225496839a"
BASE_URL    = "http://pro-x.io.vn"
MODEL       = "claude-sonnet-4-6"

# Memory systems
from ai_memory import (
    get_insights_summary, get_facts_summary,
    save_message, new_session, get_session_context,
    add_fact, load_insights,
)


# ── LOAD FILES ────────────────────────────────────────────────────────────────
def load_text(fp):
    try:
        with open(fp, "r", encoding="utf-8") as f:
            return f.read().strip()
    except Exception:
        return ""


def load_json(fp):
    try:
        with open(fp, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return {}


# ── REALTIME DATA (server.js) ───────────────────────────────────────────────
def fetch_realtime() -> tuple:
    """Goi server.js /data.json — tra ve (orders, error_msg)."""
    try:
        import socket
        socket.setdefaulttimeout(5)
        req = urllib.request.Request(
            "http://localhost:3000/data.json",
            headers={"Accept": "application/json"},
        )
        with urllib.request.urlopen(req) as resp:
            data = json.loads(resp.read().decode("utf-8"))
            return data.get("orders", []), None
    except Exception as e:
        return [], str(e)


# ── EXCEL ENRICHMENT: bo sung ngay_nhan tu File_sach/ ───────────────────────
def load_excel_map() -> dict:
    """
    Doc File_sach/*.xlsx — tra ve dict:
      { ma_dh: datetime(ngay_nhan) }
    De enrich realtime orders voi ngay_nhan that.
    """
    import openpyxl
    excel_map = {}
    for fp in sorted(Path(BASE_DIR).glob("File_sach/*.xlsx")):
        if "node_modules" in str(fp):
            continue
        try:
            wb = openpyxl.load_workbook(fp, data_only=True, read_only=True)
        except Exception:
            continue

        # Tim sheet "Don hang"
        don_sheet = None
        for shname in wb.sheetnames:
            if "don hang" in unaccent(shname):
                don_sheet = shname
                break
        if not don_sheet:
            wb.close()
            continue

        rows = list(wb[don_sheet].iter_rows(values_only=True))
        if not rows:
            wb.close()
            continue

        # Build index by unaccented column name
        h = [str(v).strip() if v is not None else "" for v in rows[0]]
        idx = {unaccent(name): i for i, name in enumerate(h)}
        ma_col = idx.get("ma dh", -1)
        nn_col = idx.get("nhan luc", -1)

        for row in rows[1:]:
            if ma_col < 0 or ma_col >= len(row):
                continue
            ma = str(row[ma_col]).strip() if row[ma_col] is not None else ""
            if not ma or ma == "Ma DH":
                continue
            if ma in excel_map:
                continue
            # Parse ngay_nhan
            ngay_val = row[nn_col] if nn_col >= 0 and nn_col < len(row) else None
            if isinstance(ngay_val, datetime):
                excel_map[ma] = ngay_val
            elif isinstance(ngay_val, (int, float)):
                try:
                    from openpyxl.utils.datetime import from_excel
                    excel_map[ma] = from_excel(ngay_val)
                except Exception:
                    pass
            elif isinstance(ngay_val, str):
                s = ngay_val.strip()
                if s and s not in ("-", "None", "nan"):
                    for fmt in ["%d/%m/%Y %H:%M:%S","%Y-%m-%d %H:%M:%S",
                                "%d/%m/%Y %H:%M","%Y-%m-%d %H:%M",
                                "%d/%m/%Y","%Y-%m-%d"]:
                        try:
                            excel_map[ma] = datetime.strptime(s, fmt)
                            break
                        except Exception:
                            pass
        wb.close()
    return excel_map


def parse_dt(s: str):
    """Parse '2026-04-01 12:00:00' or '01/04/2026 12:00:00' -> datetime."""
    if not s: return None
    for fmt in ["%Y-%m-%d %H:%M:%S","%d/%m/%Y %H:%M:%S",
                "%Y-%m-%d %H:%M","%d/%m/%Y %H:%M",
                "%Y-%m-%d","%d/%m/%Y"]:
        try: return datetime.strptime(s.strip()[:19], fmt)
        except Exception: pass
    return None


def build_realtime_context(orders: list, error: str = None,
                           excel_map: dict = None) -> str:
    """
    Build Realtime data chunk.
    excel_map: { ma_dh: datetime(ngay_nhan) } — tu File_sach/*.xlsx
               Dung de tinh lead time thuc te & do tre.
    """
    if error:
        return f"[Nguon: realtime] Khong lay duoc (server.js: {error})"
    if not orders:
        return "[Nguon: realtime] Khong co du lieu"

    today      = datetime.now()
    today_str  = today.strftime("%Y-%m-%d")
    excel_map  = excel_map or {}

    due_today   = []
    overdue     = []   # qua han, chua xong
    future      = []   # yc_giao > hom nay
    done_today  = []
    pending     = []
    by_kh       = {}
    total       = len(orders)
    done_total  = 0

    # Chi tiet don hom nay
    today_detail = []

    for o in orders:
        ma   = o.get("ma_dh", "?")
        kh   = o.get("kh") or "?"
        bn   = o.get("bn") or "?"
        yc_g = o.get("yc_giao") or ""
        pct  = o.get("pct", 0)
        is_done = pct >= 100

        by_kh[kh] = by_kh.get(kh, 0) + 1
        yc_giao_date = yc_g[:10] if len(yc_g) >= 10 else ""

        ngay_nhan = excel_map.get(ma)

        # Tinh last confirmed stage time
        last_t = None
        stages = o.get("stages", [])
        for s in stages:
            if s.get("x"):
                t = parse_dt(s.get("t") or "")
                if t and (last_t is None or t > last_t):
                    last_t = t

        # Tinh lead time thuc te (neu co ngay_nhan & last_t)
        lead_actual = None
        if ngay_nhan and last_t:
            lead_actual = (last_t - ngay_nhan).total_seconds() / 86400.0

        # Lead time yeu cau: yc_giao - ngay_nhan
        yc_dt = parse_dt(yc_g)
        lead_needed = None
        is_late = False
        if ngay_nhan and yc_dt:
            lead_needed = (yc_dt - ngay_nhan).total_seconds() / 86400.0
            # Late = chua xong + deadline da qua so voi now
            if not is_done and today > yc_dt:
                is_late = True

        item = {
            "ma": ma, "kh": kh, "bn": bn,
            "pct": pct, "is_done": is_done,
            "yc_giao": yc_giao_date,
            "ngay_nhan": ngay_nhan.strftime("%d/%m") if ngay_nhan else "?",
            "lead_actual": round(lead_actual, 1) if lead_actual is not None else None,
            "lead_needed": round(lead_needed, 1) if lead_needed is not None else None,
            "is_late": is_late,
            "curKtv": o.get("curKtv") or "",
        }

        if is_done:
            done_total += 1
            if yc_giao_date == today_str:
                done_today.append(item)
                due_today.append(item)
            elif yc_giao_date > today_str:
                future.append(item)
            # DA XONG roi tre thi van tinh (last_t > yc_dt) nhung khong dua vao overdue
        else:
            if yc_giao_date == today_str:
                due_today.append(item)
                pending.append(item)
                today_detail.append(item)
            elif yc_giao_date and yc_giao_date < today_str:
                overdue.append(item)
                item["is_late"] = True
            else:
                future.append(item)

    # Chi tiet 5 don tre nhat hom nay (chua xong + bi tre so vs lead needed)
    at_risk = [x for x in pending if x.get("is_late")]
    at_risk.sort(key=lambda x: x.get("lead_needed") or 999)
    risk_detail = ""
    if at_risk[:5]:
        risk_lines = []
        for x in at_risk[:5]:
            lt_need = x.get("lead_needed")
            lt_act  = x.get("lead_actual")
            late_tag = " [TRE]" if x.get("is_late") else ""
            if lt_need:
                risk_lines.append(
                    f"  - {x['ma']} | {x['kh']} | yc_giao={x['yc_giao']} "
                    f"| nhan={x['ngay_nhan']} | can={lt_need:.0f}ngay "
                    f"(tre={lt_need:.0f}ngay neu chi con 0){late_tag}"
                )
            else:
                risk_lines.append(
                    f"  - {x['ma']} | {x['kh']} | yc_giao={x['yc_giao']} | nhan={x['ngay_nhan']}{late_tag}"
                )
        risk_detail = "\n  At-risk (het lead time):\n" + "\n".join(risk_lines)

    top_kh = sorted(by_kh.items(), key=lambda x: -x[1])[:3]
    kh_txt = " | ".join([f"{k}={v}don" for k, v in top_kh])

    has_excel = bool(excel_map)
    source_note = "[File_sach: bo sung ngay_nhan]" if has_excel else "[server.js — chua co du lieu Excel]"

    # Don giao truoc 12h hom nay (urgent)
    urgent_done = []
    urgent_pending = []
    for o in orders:
        yc_g = o.get("yc_giao") or ""
        if len(yc_g) >= 10 and yc_g[:10] == today_str and len(yc_g) >= 13:
            try:
                hour = int(yc_g[11:13])
                if hour < 12:
                    pct = o.get("pct", 0)
                    if pct >= 100:
                        urgent_done.append(o)
                    else:
                        urgent_pending.append(o)
            except ValueError:
                pass
    urgent_txt = ""
    if urgent_done or urgent_pending:
        urgent_txt = (
            f"\n-- Urgent (giao truoc 12h {today_str}): {len(urgent_done)+len(urgent_pending)} don"
            f"\n  Da xong: {len(urgent_done)} don"
            f"\n  Chua xong: {len(urgent_pending)} don"
        )
        if urgent_pending:
            urgent_txt += "\n  Chi tiet (chua xong):"
            for o in urgent_pending[:5]:
                urgent_txt += (
                    f"\n    - {o.get('ma_dh','?')} | {o.get('kh','?')} "
                    f"| yc_giao={o.get('yc_giao','?')} | pct={o.get('pct','?')}% | KTV={o.get('curKtv','?')}"
                )

    return f"""[Nguon: realtime] {source_note} (hom nay: {today_str})
Tong don: {total} | Da xong: {done_total} | Dang lam: {len(pending)}
-- Don can giao hom nay ({today_str}): {len(due_today)} don
  Da xong, san sang giao: {len(done_today)} don
  Chua xong, can hoan thanh hom nay: {len(due_today) - len(done_today)} don
-- Don qua han (yc_giao < {today_str}, chua xong): {len(overdue)} don
-- Don tuong lai (yc_giao > {today_str}): {len(future)} don
Top KH: {kh_txt}{risk_detail}{urgent_txt}"""


# ── RAG: RETRIEVE RELEVANT KNOWLEDGE ─────────────────────────────────────────
REALTIME_TRIGGERS = [
    "hôm nay","hom nay","hiện tại","hien tai","realtime",
    "đang làm","dang lam","cần giao","can giao","chưa xong",
    "sắp giao","sap giao","đơn hôm nay","don hom nay",
    "tình trạng","tinh trang","hien co","hiện có",
    "trạng thái","trang thai","tra cứu","tra cuu","kiểm tra",
    "đơn nào","don nao","mấy đơn","may don",
    "giao khi","giao luc","nhận khi","nhan khi","ngày nào",
    "mã đơn","ma don","ma_dh","đơn ","don ",
    "12h","trưa","sáng","chiều","buổi","giờ giao","gio giao",
    # Specific order IDs (8-12 digit patterns)
    "263003050","263003035","263003036","263003037","263003038",
]

STATS_TRIGGERS = [
    "tổng đơn","tong don","tháng","thang","tổng cộng",
    "top khách","top khach","top ktv","ktv nhiều",
    "nút thắt","nut that","bottleneck","lead time",
    "remake","làm lại","xu hướng","xu huong","trend",
    "số lượng","so luong","bao nhiêu","bao nhieu",
    "trung bình","trung binh","tổng răng","tong rang",
    "phân tích","phan tich","thống kê","thong ke",
]

KNOWLEDGE_TRIGGERS = [
    "quy trình","quy trinh","công đoạn","cong doan","cbm",
    "sáp","sườn","đắp","mài","cadcam",
    "vật liệu","vat lieu","zirconia","titanium",
    "làm lại","sửa","bảo hành","bảo hành","làm tiếp",
    "thử sườn","thu suon","ts","đơn sửa","don sua",
    "khách hàng","khach hang","ktv là gì","ai là ktv",
]


def detect_sources(user_msg: str) -> list:
    """Xac dinh nguon nao can doc cho cau hoi nay."""
    msg = user_msg.lower()
    sources = []

    if any(t in msg for t in REALTIME_TRIGGERS):
        sources.append("realtime")
    if any(t in msg for t in STATS_TRIGGERS):
        sources.append("stats")
    if any(t in msg for t in KNOWLEDGE_TRIGGERS):
        sources.append("knowledge")

    # Mac dinh: stats + knowledge
    if not sources:
        sources = ["stats", "knowledge"]

    return list(dict.fromkeys(sources))  # keep order, remove dup


def build_stats_context(memory: dict) -> str:
    ls = memory.get("learnedStats", {})
    if not ls:
        return None

    dr  = ls.get("dataRange", {})
    lt  = ls.get("avgLeadTime", {})
    lt_txt = " | ".join([f"{k}={v}" for k, v in lt.items() if v and v != "—"])

    tc  = ls.get("topCustomers", [])[:3]
    tc_txt = " | ".join([f"{c['name']}={c['count']}" for c in tc])

    kp  = ls.get("ktvPerformance", [])[:3]
    kp_txt = " | ".join([f"{k['name']}={k['xnCount']}don" for k in kp])

    dv  = ls.get("dailyVolume", {})
    dv_txt = f"TB={dv.get('avg','?')}/ngay Max={dv.get('max','?')} Min={dv.get('min','?')}"

    return f"""[Nguon: learnedStats] (Data: {dr.get('from','?')} → {dr.get('to','?')})
Tong don: {ls.get('totalOrdersAnalyzed','?')} | Tong rang: {ls.get('totalTeeth','?')}
Lead time: {lt_txt or '?'}
Nut that: {ls.get('bottleneckStage','?')} ({ls.get('bottleneckPct','?')}% don chua xong)
Remake: {ls.get('remakeRate','?')} ({ls.get('remakeCount','?')} don)
Daily vol: {dv_txt}
Top KH: {tc_txt or '?'}
Top KTV: {kp_txt or '?'}"""


def build_knowledge_context(knowledge: dict) -> str:
    lab = knowledge.get("lab", {})
    stages = lab.get("stages", [])

    stage_txt = "\n".join([
        f"  {s}: {lab.get('stage_descriptions',{}).get(s,'')}"
        for s in stages
    ])

    mats = knowledge.get("materials", {})
    mat_txt = "\n".join([
        f"  {m}: {v.get('time','?')} ({', '.join(v.get('keywords',[])[:3])})"
        for m, v in mats.items()
    ])

    ktvs = knowledge.get("ktvs", [])
    ktv_txt = ", ".join(ktvs[:10]) + ("..." if len(ktvs) > 10 else "")

    sos = knowledge.get("special_orders", {})
    so_txt = "\n".join([f"  {k}: {v.get('note','')}" for k, v in sos.items()])

    return f"""[Nguon: knowledge]
Stages: {', '.join(stages)}
{stage_txt}

Vat lieu:
{mat_txt or '  (khong co du lieu)'}

KTV ({len(ktvs)} nguoi): {ktv_txt}

Don dac biet:
{so_txt or '  (khong co du lieu)'}"""


# ── ASSEMBLE PROMPT ──────────────────────────────────────────────────────────
def build_prompt(user_msg: str, sources: list, memory: dict, knowledge: dict,
                 session_id: str = None) -> str:
    instructions = load_text(INSTRUCTIONS)

    lines = []
    lines.append(instructions)
    lines.append("")

    # Session context (neu co lich su)
    if session_id:
        ctx = get_session_context(session_id)
        if ctx:
            lines.append(ctx)
            lines.append("")

    # Data chunks
    lines.append("=== DATA CHUNKS ===")

    has_realtime = False
    for src in sources:
        if src == "realtime":
            orders, err = fetch_realtime()
            # Enrich voi ngay_nhan tu Excel
            excel_map = load_excel_map() if not err else {}
            ctx = build_realtime_context(orders, err, excel_map)
            lines.append(ctx)
            has_realtime = True

        elif src == "stats":
            ctx = build_stats_context(memory)
            if ctx:
                lines.append(ctx)

        elif src == "knowledge":
            ctx = build_knowledge_context(knowledge)
            if ctx:
                lines.append(ctx)

    # Insights timeline + Learned facts
    insights_txt = get_insights_summary()
    facts_txt    = get_facts_summary()
    if insights_txt and "(chua co" not in insights_txt:
        lines.append("")
        lines.append(insights_txt)
    if facts_txt and "(chua co" not in facts_txt:
        lines.append("")
        lines.append(facts_txt)

    # Neu hoi ve don cu the (ma don 8-12 chu so) — trich xuat chi tiet
    specific_ids = re.findall(r"\b\d{8,13}\b", user_msg)
    if specific_ids and orders:
        lines.append("")
        lines.append("=== CHI TIET DON THEO MA ===")
        for oid in specific_ids:
            for o in orders:
                if o.get("ma_dh") == oid:
                    nn_excel = excel_map.get(oid)
                    nn_str = nn_excel.strftime("%d/%m/%Y") if nn_excel else (o.get("nhan") or "?")
                    stages = o.get("stages", [])
                    stage_txt = " | ".join([
                        f"{s.get('n','?')}:{s.get('k','?')}:{'✓' if s.get('x') else '✗'}"
                        for s in stages
                    ]) if stages else "?"
                    lines.append(
                        f"  {oid} | KH={o.get('kh','?')} | BN={o.get('bn','?')} "
                        f"| nhan={nn_str} | yc_giao={o.get('yc_giao','?')} "
                        f"| pct={o.get('pct','?')}% | KTV={o.get('curKtv','?')} "
                        f"| stages=[{stage_txt}]"
                    )
                    break

    if not has_realtime:
        lines.append("")
        lines.append("[Luu y] Cau hoi ve du lieu hien tai — xem chunk [Nguon: realtime] o tren")

    lines.append("")
    lines.append(f"=== CAU HOI ===\n{user_msg}")

    return "\n".join(lines)


# ── CHAT ────────────────────────────────────────────────────────────────────
def chat(user_msg: str, verbose: bool = False, session_id: str = None) -> str:
    print("  [Dang suy nghi... ]", end="", flush=True)
    try:
        import anthropic
        client    = anthropic.Anthropic(api_key=API_KEY, base_url=BASE_URL)
        memory    = load_json(MEMORY_FP)
        knowledge = load_json(KNOWLEDGE_FP)
        sources   = detect_sources(user_msg)
        prompt    = build_prompt(user_msg, sources, memory, knowledge, session_id)

        if verbose:
            print(f"\n[SOURCES: {sources}]", file=sys.stderr)
            print(f"[PROMPT]\n{prompt[:500]}...", file=sys.stderr)

        response = client.messages.create(
            model=MODEL,
            max_tokens=1024,
            messages=[{"role": "user", "content": prompt}],
            tools=[],  # tat tool use — chi tra loi, khong tu browse
        )

        parts = []
        for block in response.content:
            if hasattr(block, "text"):
                parts.append(block.text.strip())

        answer = "\n".join(parts).strip()

        # Luu lich su session
        if session_id:
            save_message(session_id, "user",   user_msg)
            save_message(session_id, "assistant", answer)

        print("\r  [Xong]            \n", flush=True)
        return answer

    except Exception as e:
        print(f"\r  [Loi: {e}]            ", flush=True)
        return f"LOI: {e}"


# ── INTERACTIVE ─────────────────────────────────────────────────────────────
def interactive():
    print("=" * 50)
    print("  ASIA LAB AI Chat -- Claude Sonnet 4 (NotebookLM)")
    print("  3 nguon: realtime | learnedStats | knowledge")
    print("  Tu hoc: insights timeline | learned facts | session memory")
    print("  Go 'exit'/'quit' de thoat | 'new' de reset session")
    print("=" * 50)

    memory     = load_json(MEMORY_FP)
    knowledge  = load_json(KNOWLEDGE_FP)
    session_id = new_session()

    if memory.get("learnedStats"):
        dr = memory["learnedStats"].get("dataRange", {})
        print(f"\n  [learnedStats] Data: {dr.get('from','?')} -> {dr.get('to','?')}")
    if knowledge.get("lab"):
        print(f"  [knowledge] Lab: {knowledge['lab'].get('name','?')}")
    insights_txt = get_insights_summary()
    if insights_txt and "(chua co" not in insights_txt:
        print(f"  [insights] Timeline co {load_insights()['entries'][-1]['at'][:10]}")
    print()

    while True:
        try:
            user = input("Ban: ").strip()
        except (EOFError, KeyboardInterrupt):
            print("\nTam biet!")
            break
        if not user:
            continue
        if user.lower() in ("exit", "quit", "q"):
            print("Tam biet!")
            break
        if user.lower() == "new":
            session_id = new_session()
            print("  Session moi — da reset lich su")
            continue
        if user.lower() == "reload":
            memory    = load_json(MEMORY_FP)
            knowledge = load_json(KNOWLEDGE_FP)
            print("  Da reload memory + knowledge")
            continue
        if user.lower() == "sources":
            print(f"  Nguon: {detect_sources(input('Cau hoi: ').strip())}")
            continue

        print()
        result = chat(user, session_id=session_id)
        print("  AI:", result)
        print()


if __name__ == "__main__":
    if len(sys.argv) > 1:
        question = " ".join(sys.argv[1:])
        print(chat(question))
    else:
        interactive()
